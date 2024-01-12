import sys
import os
import django
from django.utils import timezone
import datetime
from openpyxl import load_workbook

if os.getenv('DJANGO_DEVELOPMENT') == 'true':
    print('DEV')
    path = os.path.join('C:\\Users\\bulmaj\\OneDrive - Cambridge\\Desktop\\Dev\\Nova')
    sys.path.append(path)
    os.environ['DJANGO_SETTINGS_MODULE'] = 'redepplan.settings_dev'
elif os.getenv('DJANGO_PRODUCTION') == 'true':
    print('PROD')
    path = os.path.join('C:\\Dev\\Nova')
    sys.path.append(path)
    os.environ['DJANGO_SETTINGS_MODULE'] = 'redepplan.settings_prod'
else:
    print('UAT - Check')
    path = os.path.join('C:\\Dev\\nova')
    sys.path.append(path)
    os.environ['DJANGO_SETTINGS_MODULE'] = 'redepplan.settings'

django.setup()

from enquiries.models import TaskManager, EnquiryComponents, CentreEnquiryRequests, TaskTypes, MarkTolerances
from django.contrib.auth.models import User

def run_algo():
    for task in TaskManager.objects.filter(task_id='GRDNEG', task_completion_date__isnull=True):
        # TODO: actually check if grade is negative...
        enquiry_id = task.enquiry_id.enquiry_id


        TaskManager.objects.create(
            enquiry_id = CentreEnquiryRequests.objects.get(enquiry_id=enquiry_id),
            ec_sid = None,
            task_id = TaskTypes.objects.get(task_id = 'NEGCON'),
            task_assigned_to = None,
            task_assigned_date = None,
            task_completion_date = None
        ) 

        TaskManager.objects.filter(pk=task.pk,task_id='GRDNEG').update(task_completion_date=timezone.now())   


    filename=os.path.join("\\\\filestorage\cie\Operations\Results Team\Enquiries About Results\\0.Nova Downloads\\Tolerances.xlsx")
    workbook = load_workbook(filename)
    sheet = workbook.active

    # Iterating through All rows with all columns...
    for i in range(1, sheet.max_row+1):
        row = [cell.value for cell in sheet[i]] # sheet[n] gives nth row (list of cells)
        if MarkTolerances.objects.filter(eps_ass_code = row[0],eps_com_id = row[1]).exists():
            MarkTolerances.objects.filter(eps_ass_code = row[0],eps_com_id = row[1]).update(mark_tolerance = row[2])
        else: 
            MarkTolerances.objects.create(
                eps_ass_code = row[0],
                eps_com_id = row[1],
                mark_tolerance = row[2]
            )
        

run_algo()