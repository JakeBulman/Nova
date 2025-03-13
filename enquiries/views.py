from django.http import FileResponse, HttpResponseRedirect
from . import models
from django.shortcuts import render, redirect
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.core.paginator import Paginator
from django.db.models import Q, QuerySet
from django.conf import settings
import csv, os, datetime
from django.db.models import Sum, Count
from django.contrib.auth.models import User
from django.db.models.functions import Cast, Substr
import dateutil.parser
from django.db.models import OuterRef, Subquery, Max
from datetime import datetime, timedelta, time
from django.db import connection, reset_queries
from openpyxl import load_workbook

#special imports
#from . import script_ServerResetShort as srs
#from . import script_ServerResetFull as srf

PageNotAnInteger = None
EmptyPage = None

# Create your views here.
@login_required
def ear_home_view(request,*args, **kwargs):
	user = None
	if request.user.is_authenticated:
		user = request.user

	alpha_tasks = ['INITCH']
	gamma_tasks = ['ESMCSV','OMRCHE','MANAPP','BOTAPF','MISVRM','MISVRF','MARCHE','LOCMAR','PEXMCH','EXMSLA','REMAPP','REMAPF','MUPREX','NRMSCS','BOTMAF',]
	delta_tasks = ['NRMACC','S3SEND','S3CONF']
	kappa_tasks = ['CLERIC',]
	sigma_tasks = ['ESMSCR','ESMSC2','SCRCHE','SCRREQ','OMRSCR']
	omega_tasks = ['NEGCON','PDACON','PEACON','GRDREJ','MRKAMD','GRDCHG',]
	lambda_tasks = ['BOTAPP','BOTMAR',]

	mytask_count = models.TaskManager.objects.filter(task_assigned_to=user, task_completion_date__isnull=True)
	alpha_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id__in=alpha_tasks, enquiry_tasks__task_completion_date__isnull=True)
	gamma_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id__in=gamma_tasks, enquiry_tasks__task_completion_date__isnull=True)
	delta_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id__in=delta_tasks, enquiry_tasks__task_completion_date__isnull=True)
	kappa_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id__in=kappa_tasks, enquiry_tasks__task_completion_date__isnull=True)
	sigma_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id__in=sigma_tasks, enquiry_tasks__task_completion_date__isnull=True)
	omega_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id__in=omega_tasks, enquiry_tasks__task_completion_date__isnull=True)
	lambda_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id__in=lambda_tasks, enquiry_tasks__task_completion_date__isnull=True)

	session_desc = models.EarServerSettings.objects.first().session_description
	delta_load_status = models.EarServerSettings.objects.first().delta_load_status
	context = {"session_desc":session_desc, "delta_load_status":delta_load_status, "mytask_count":mytask_count,
			"alpha_count":alpha_count, "gamma_count":gamma_count, "delta_count":delta_count, "kappa_count":kappa_count, 
			"sigma_count":sigma_count, "omega_count":omega_count, "lambda_count":lambda_count,
		}

	user_status = models.TaskUserPrimary.objects.get(task_user_id=user).primary_status
	
	if request.htmx:
		return render(request, 'enquiries/htmx_partials/homepage_pill_count.html', context)
	else:
		if user_status == 'CO':
			return render(request, "enquiries/main_templates/home_ear_coordinator.html", context=context, )
		elif user_status == 'TL':
			return render(request, "enquiries/main_templates/home_ear_tl.html", context=context, )
		elif user_status == 'AD':
			return render(request, "enquiries/main_templates/home_ear_admin.html", context=context, )

		return render(request, "enquiries/main_templates/home_ear_restricted.html", context=context, )


def ear_home_view_team_alpha(request,*args, **kwargs):
	user = None
	if request.user.is_authenticated:
		user = request.user
	
	mytask_count = models.TaskManager.objects.filter(task_assigned_to=user, task_completion_date__isnull=True)
	cer_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='INITCH', enquiry_tasks__task_completion_date__isnull=True,enquiries__enquiry_parts__isnull=False).distinct()
	cer_count_compless = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='INITCH', enquiry_tasks__task_completion_date__isnull=True,enquiries__enquiry_parts__isnull=True)

	session_desc = models.EarServerSettings.objects.first().session_description
	context = {"session_desc":session_desc, "mytask":mytask_count,"cer":cer_count, "cer_count_compless":cer_count_compless, 
		}

	return render(request, "enquiries/main_templates/home_ear_alpha.html", context=context, )

def ear_home_view_team_delta(request,*args, **kwargs):
	user = None
	if request.user.is_authenticated:
		user = request.user
	
	mytask_count = models.TaskManager.objects.filter(task_assigned_to=user, task_completion_date__isnull=True)
	nrmacc_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='NRMACC', enquiry_tasks__task_completion_date__isnull=True)
	nrmacca_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='NRMACC', enquiry_tasks__task_completion_date__isnull=True, enquiry_tasks__task_assigned_to__isnull=False)
	s3send_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='S3SEND', enquiry_tasks__task_completion_date__isnull=True)
	s3senda_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='S3SEND', enquiry_tasks__task_completion_date__isnull=True, enquiry_tasks__task_assigned_to__isnull=False)
	s3conf_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='S3CONF', enquiry_tasks__task_completion_date__isnull=True)
	s3confa_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='S3CONF', enquiry_tasks__task_completion_date__isnull=True, enquiry_tasks__task_assigned_to__isnull=False)


	session_desc = models.EarServerSettings.objects.first().session_description
	context = {"session_desc":session_desc, "mytask":mytask_count, "nrmacc":nrmacc_count, "nrmacca":nrmacca_count,  
			"s3send":s3send_count, "s3senda":s3senda_count, "s3conf":s3conf_count, "s3confa":s3confa_count}

	return render(request, "enquiries/main_templates/home_ear_delta.html", context=context, )

def ear_home_view_team_gamma(request,*args, **kwargs):
	user = None
	if request.user.is_authenticated:
		user = request.user
	
	mytask_count = models.TaskManager.objects.filter(task_assigned_to=user, task_completion_date__isnull=True)
	esmcsv_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='ESMCSV', enquiry_tasks__task_completion_date__isnull=True)
	omrche_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='OMRCHE', enquiry_tasks__task_completion_date__isnull=True)
	nrmscs_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='NRMSCS', enquiry_tasks__task_completion_date__isnull=True)
	nrmscsa_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='NRMSCS', enquiry_tasks__task_completion_date__isnull=True, enquiry_tasks__task_assigned_to__isnull=False)	
	manapp_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='MANAPP', enquiry_tasks__task_completion_date__isnull=True)
	manapp_count_assigned = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='MANAPP', enquiry_tasks__task_completion_date__isnull=True, enquiry_tasks__task_assigned_to__isnull=False)
	botapp_fail_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='BOTAPF', enquiry_tasks__task_completion_date__isnull=True)
	misvrm_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='MISVRM', enquiry_tasks__task_completion_date__isnull=True)
	misvrma_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='MISVRM', enquiry_tasks__task_completion_date__isnull=True, enquiry_tasks__task_assigned_to__isnull=False)
	misvrf_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='MISVRF', enquiry_tasks__task_completion_date__isnull=True)
	misvrfa_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='MISVRF', enquiry_tasks__task_completion_date__isnull=True, enquiry_tasks__task_assigned_to__isnull=False)
	marche_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='MARCHE', enquiry_tasks__task_completion_date__isnull=True)
	marchea_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='MARCHE', enquiry_tasks__task_completion_date__isnull=True, enquiry_tasks__task_assigned_to__isnull=False)
	locmar_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='LOCMAR', enquiry_tasks__task_completion_date__isnull=True)
	locmara_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='LOCMAR', enquiry_tasks__task_completion_date__isnull=True, enquiry_tasks__task_assigned_to__isnull=False)
	pexmch_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='PEXMCH', enquiry_tasks__task_completion_date__isnull=True)
	pexmcha_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='PEXMCH', enquiry_tasks__task_completion_date__isnull=True, enquiry_tasks__task_assigned_to__isnull=False)
	exmsla_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='EXMSLA', enquiry_tasks__task_completion_date__isnull=True)
	exmslaa_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='EXMSLA', enquiry_tasks__task_completion_date__isnull=True, enquiry_tasks__task_assigned_to__isnull=False)
	remapp_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='REMAPP', enquiry_tasks__task_completion_date__isnull=True)
	remappa_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='REMAPP', enquiry_tasks__task_completion_date__isnull=True, enquiry_tasks__task_assigned_to__isnull=False)
	remapf_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='REMAPF', enquiry_tasks__task_completion_date__isnull=True)
	remapfa_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='REMAPF', enquiry_tasks__task_completion_date__isnull=True, enquiry_tasks__task_assigned_to__isnull=False)
	muprex_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='MUPREX', enquiry_tasks__task_completion_date__isnull=True)
	muprexa_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='MUPREX', enquiry_tasks__task_completion_date__isnull=True, enquiry_tasks__task_assigned_to__isnull=False)
	botmar_fail_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='BOTMAF', enquiry_tasks__task_completion_date__isnull=True)

	session_desc = models.EarServerSettings.objects.first().session_description
	context = {"session_desc":session_desc, "mytask":mytask_count, "manapp": manapp_count, "manappa": manapp_count_assigned, "nrmscs":nrmscs_count, "nrmscsa":nrmscsa_count,
	     "botapf":botapp_fail_count, "misvrm":misvrm_count, "misvrma":misvrma_count, 
		"misvrf":misvrf_count, "misvrfa":misvrfa_count,	"pexmch":pexmch_count, "pexmcha":pexmcha_count, "locmar":locmar_count, "locmara":locmara_count, 
		"esmcsv":esmcsv_count, "omrche":omrche_count, "exmsla":exmsla_count, "exmslaa":exmslaa_count, "remapp":remapp_count, "remappa":remappa_count, 
		"remapf":remapf_count, "remapfa":remapfa_count, "muprex":muprex_count, "muprexa":muprexa_count, "marche":marche_count, "marchea":marchea_count,
		"botmaf":botmar_fail_count,
		}

	return render(request, "enquiries/main_templates/home_ear_gamma.html", context=context, )

def ear_home_view_team_kappa(request,*args, **kwargs):
	user = None
	if request.user.is_authenticated:
		user = request.user
	
	mytask_count = models.TaskManager.objects.filter(task_assigned_to=user, task_completion_date__isnull=True)
	cleric_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='CLERIC', enquiry_tasks__task_completion_date__isnull=True)
	clerica_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='CLERIC', enquiry_tasks__task_completion_date__isnull=True, enquiry_tasks__task_assigned_to__isnull=False)

	session_desc = models.EarServerSettings.objects.first().session_description
	context = {"session_desc":session_desc, "mytask":mytask_count, "cleric":cleric_count, "clerica":clerica_count,
		}

	return render(request, "enquiries/main_templates/home_ear_kappa.html", context=context, )

def ear_home_view_team_lambda(request,*args, **kwargs):
	user = None
	if request.user.is_authenticated:
		user = request.user
	
	botapp_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='BOTAPP', enquiry_tasks__task_completion_date__isnull=True)
	botmar_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='BOTMAR', enquiry_tasks__task_completion_date__isnull=True)

	session_desc = models.EarServerSettings.objects.first().session_description
	context = {"session_desc":session_desc, "botapp":botapp_count, "botmar":botmar_count,
		}

	return render(request, "enquiries/main_templates/home_ear_lambda.html", context=context, )

def ear_home_view_team_omega(request,*args, **kwargs):
	user = None
	if request.user.is_authenticated:
		user = request.user
	
	mytask_count = models.TaskManager.objects.filter(task_assigned_to=user, task_completion_date__isnull=True)
	negcon_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='NEGCON', enquiry_tasks__task_completion_date__isnull=True)
	negcona_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='NEGCON', enquiry_tasks__task_completion_date__isnull=True, enquiry_tasks__task_assigned_to__isnull=False)
	pdacon_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='PDACON', enquiry_tasks__task_completion_date__isnull=True)
	pdacona_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='PDACON', enquiry_tasks__task_completion_date__isnull=True, enquiry_tasks__task_assigned_to__isnull=False)
	peacon_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='PEACON', enquiry_tasks__task_completion_date__isnull=True)
	peacona_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='PEACON', enquiry_tasks__task_completion_date__isnull=True, enquiry_tasks__task_assigned_to__isnull=False)
	grdrej_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='GRDREJ', enquiry_tasks__task_completion_date__isnull=True)
	grdreja_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='GRDREJ', enquiry_tasks__task_completion_date__isnull=True, enquiry_tasks__task_assigned_to__isnull=False)
	mrkamd_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='MRKAMD', enquiry_tasks__task_completion_date__isnull=True)
	mrkamda_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='MRKAMD', enquiry_tasks__task_completion_date__isnull=True, enquiry_tasks__task_assigned_to__isnull=False)
	grdchg_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='GRDCHG', enquiry_tasks__task_completion_date__isnull=True)
	grdchga_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='GRDCHG', enquiry_tasks__task_completion_date__isnull=True, enquiry_tasks__task_assigned_to__isnull=False)


	session_desc = models.EarServerSettings.objects.first().session_description
	context = {"session_desc":session_desc, "mytask":mytask_count, "negcon":negcon_count, "negcona":negcona_count, "pdacon":pdacon_count, "pdacona":pdacona_count, 
		"peacon":peacon_count, "peacona":peacona_count, "grdrej":grdrej_count, "grdreja":grdreja_count, "mrkamd":mrkamd_count, 
		"mrkamda":mrkamda_count, "grdchg":grdchg_count, "grdchga":grdchga_count,
		}

	return render(request, "enquiries/main_templates/home_ear_omega.html", context=context, )

def ear_home_view_team_sigma(request,*args, **kwargs):
	user = None
	if request.user.is_authenticated:
		user = request.user
	
	mytask_count = models.TaskManager.objects.filter(task_assigned_to=user, task_completion_date__isnull=True)
	esmscr_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='ESMSCR', enquiry_tasks__task_completion_date__isnull=True)
	esmsc2_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='ESMSC2', enquiry_tasks__task_completion_date__isnull=True)
	omrscr_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='OMRSCR', enquiry_tasks__task_completion_date__isnull=True)
	scrren_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='SCRREN', enquiry_tasks__task_completion_date__isnull=True)
	scrche_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='SCRCHE', enquiry_tasks__task_completion_date__isnull=True)
	scrchea_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='SCRCHE', enquiry_tasks__task_completion_date__isnull=True, enquiry_tasks__task_assigned_to__isnull=False)
	scrreq_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='SCRREQ', enquiry_tasks__task_completion_date__isnull=True)
	scrreqa_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='SCRREQ', enquiry_tasks__task_completion_date__isnull=True, enquiry_tasks__task_assigned_to__isnull=False)

	session_desc = models.EarServerSettings.objects.first().session_description
	context = {"session_desc":session_desc, "mytask":mytask_count,"esmscr":esmscr_count,"esmsc2":esmsc2_count,"omrscr":omrscr_count,"scrren":scrren_count,"scrche":scrche_count, "scrchea":scrchea_count,"scrreq":scrreq_count, "scrreqa":scrreqa_count,
		}

	return render(request, "enquiries/main_templates/home_ear_sigma.html", context=context, )

def server_options_view(request):
	context = {}
	return render(request, "enquiries/main_templates/enquiries_server_options.html", context=context)

def server_settings_view(request):
	serv = models.EarServerSettings.objects.first()
	context = {"serv":serv}
	return render(request, "enquiries/main_templates/enquiries_server_settings.html", context=context)

def server_settings_update_view(request):
	sessions = request.POST.get('session_id_list')
	enquiries = request.POST.get('enquiry_id_list')
	session_desc = request.POST.get('session_desc')
	
	serv = models.EarServerSettings.objects.first()
	models.EarServerSettings.objects.filter(id=serv.pk).update(session_id_list=sessions,enquiry_id_list=enquiries,session_description=session_desc)
	# models.EarServerSettings.objects.create(
	# 	session_id_list=sessions,
	# 	enquiry_id_list=enquiries
	# )
	serv = models.EarServerSettings.objects.all().first()
	context = {"serv":serv}
	return render(request, "enquiries/main_templates/enquiries_server_settings.html", context=context)

def server_short_reset_view(request):
	# srs.clear_tables()
	# srs.load_core_tables()
	context = {}
	return render(request, "enquiries/main_templates/enquiries_server_options.html", context=context)

def server_long_reset_view(request):
# 	srf.clear_tables()
# 	srf.load_core_tables()
		context = {}
		return render(request, "enquiries/main_templates/enquiries_server_options.html", context=context)

def my_tasks_view(request):
		#Get username to filter tasks
	user = None
	if request.user.is_authenticated:
		user = request.user

	excluded_task_list = ['INITCH','SETBIE','AUTAPP','BOTAPP','NEWMIS','RETMIS','JUSCHE','BOTMAR','GRDMAT','ESMCSV','ESMSCR','ESMSC2','OMRCHE','SCRREN','SCRAUD','LETSCR','OMRSCR']
	primary_team = models.TaskUserPrimary.objects.get(task_user=user).primary_team
	secondary_team_set = models.TaskUserSecondary.objects.filter(task_user=user)
	secondary_teams = []
	for team in secondary_team_set:
		secondary_teams.append(team.secondary_team)
	#Get task objects for this user
	task_queryset = models.TaskManager.objects.select_related('ec_sid__erp_sid__cer_sid','enquiry_id').prefetch_related('ec_sid__script_id__eb_sid','enquiry_id__enquiry_deadline').filter(task_assigned_to=user,task_completion_date__isnull=True).order_by('enquiry_id__enquiry_deadline__enquiry_deadline')
	task_count = models.TaskManager.objects.filter(task_assigned_to__isnull=True,task_completion_date__isnull=True,task_id__task_team=primary_team).exclude(task_id__in=excluded_task_list).count() + models.TaskManager.objects.order_by('task_creation_date').filter(task_assigned_to__isnull=True,task_completion_date__isnull=True,task_id__task_team__in=secondary_teams).exclude(task_id__in=excluded_task_list).count() 
	context = {"tasks": task_queryset, "task_count": task_count}
	return render(request, "enquiries/main_templates/my_tasks.html", context=context)

def task_router_view(request, task_id):
	if task_id == None:
		#task_type = request.POST.get('task_id')
		task_id = request.POST.get('task_id')
	task_type = models.TaskManager.objects.get(pk=task_id).task_id.task_id	
	if task_type == "MANAPP":
		return redirect('manual-apportionment-task', task_id=task_id)
	if task_type == "NRMACC":
		return redirect('nrmacc-task', task_id=task_id)
	if task_type == "NRMSCS":
		return redirect('nrmscs-task', task_id=task_id)
	if task_type == "S3SEND":
		return redirect('s3send-task', task_id=task_id)
	if task_type == "S3CONF":
		return redirect('s3conf-task', task_id=task_id)
	if task_type == "MISVRM":
		return redirect('misvrm-task', task_id=task_id)
	if task_type == "MISVRF":
		return redirect('misvrf-task', task_id=task_id)
	if task_type == "MARCHE":
		return redirect('marche-task', task_id=task_id)
	if task_type == "PEXMCH":
		return redirect('pexmch-task', task_id=task_id)
	if task_type == "LOCMAR":
		return redirect('locmar-task', task_id=task_id)
	if task_type == "CLERIC":
		return redirect('cleric-task', task_id=task_id)
	if task_type == "MUPREX":
		return redirect('muprex-task', task_id=task_id)
	if task_type == "SCRCHE":
		return redirect('scrche-task', task_id=task_id)
	if task_type == "SCRREQ":
		return redirect('scrreq-task', task_id=task_id)
	if task_type == "SCRREN":
		return redirect('scrren_list')
	if task_type == "EXMSLA":
		return redirect('exmsla-task', task_id=task_id)
	if task_type == "REMAPP":
		return redirect('remapp-task', task_id=task_id)
	if task_type == "REMAPF":
		return redirect('remapf-task', task_id=task_id)
	if task_type == "BOTAPF":
		return redirect('botapf-task', task_id=task_id)
	if task_type == "BOTMAF":
		return redirect('botmaf-task', task_id=task_id)
	if task_type == "NEGCON":
		return redirect('negcon-task', task_id=task_id)
	if task_type == "PDACON":
		return redirect('pdacon-task', task_id=task_id)
	if task_type == "PEACON":
		return redirect('peacon-task', task_id=task_id)
	if task_type == "GRDREJ":
		return redirect('grdrej-task', task_id=task_id)
	if task_type == "MRKAMD":
		return redirect('mrkamd-task', task_id=task_id)
	if task_type == "GRDCHG":
		return redirect('grdchg-task', task_id=task_id)
	else:
		return redirect('my_tasks')
	
def new_task_view(request):
	#Get username to filter tasks
	username = None
	if request.user.is_authenticated:
		username =request.user
	excluded_task_list = ['INITCH','SETBIE','AUTAPP','BOTAPP','NEWMIS','RETMIS','JUSCHE','BOTMAR','GRDMAT','ESMCSV','ESMSCR','ESMSC2','OMRCHE','SCRREN','SCRAUD','LETSCR','OMRSCR']
	primary_team = models.TaskUserPrimary.objects.get(task_user=username).primary_team
	secondary_team_set = models.TaskUserSecondary.objects.filter(task_user=username)
	secondary_teams = []
	for team in secondary_team_set:
		secondary_teams.append(team.secondary_team)
	#Caclulate next task in the queue
	if models.TaskManager.objects.filter(task_assigned_to__isnull=True,task_completion_date__isnull=True,task_id__task_team=primary_team).exclude(task_id__in=excluded_task_list).exists():
		print('primary')
		next_task_id = models.TaskManager.objects.order_by('enquiry_id__enquiry_deadline__enquiry_deadline').filter(task_assigned_to__isnull=True,task_completion_date__isnull=True,task_id__task_team=primary_team).exclude(task_id__in=excluded_task_list).first().pk
	elif models.TaskManager.objects.filter(task_assigned_to__isnull=True,task_completion_date__isnull=True,task_id__task_team__in=secondary_teams).exclude(task_id__in=excluded_task_list).exists():
		print('secondary')
		next_task_id = models.TaskManager.objects.order_by('enquiry_id__enquiry_deadline__enquiry_deadline').filter(task_assigned_to__isnull=True,task_completion_date__isnull=True,task_id__task_team__in=secondary_teams).exclude(task_id__in=excluded_task_list).first().pk
	else:
		print('none')
		next_task_id = None
	#Set the newest task to this user
	if next_task_id is not None:
		models.TaskManager.objects.filter(id=next_task_id).update(task_assigned_to=username)
		models.TaskManager.objects.filter(id=next_task_id).update(task_assigned_date=timezone.now())
	return redirect('my_tasks')

def set_backlog(request):
	task_id = request.POST.get('task_id')
	task_val = request.POST.get('task_val')
	print(task_val)
	if task_val == '0': 
		set_val = 1
	else: 
		set_val = 0
	print(set_val)
	models.TaskManager.objects.filter(id=task_id).update(task_queued=set_val)
	return redirect('my_tasks')

def new_task_comment_view(request):
	task_id = request.POST.get('task_id')
	task_comment = request.POST.get('task_comment')
	#get current user
	username = None
	if request.user.is_authenticated:
		username =request.user

	#get this task, assuming valid
	task = models.TaskManager.objects.get(pk=task_id)
	if task.task_id.task_id == 'SCRREN':
		try:
			models.TaskComments.objects.filter(task_pk = task).first().task_pk
			models.TaskComments.objects.filter(task_pk = task).update(
				task_pk = task,
				task_comment_text = task_comment,
				task_comment_user = username
			)		
		except:
			models.TaskComments.objects.create(
				task_pk = task,
				task_comment_text = task_comment,
				task_comment_user = username
			)			
	elif task_id is not None:
		models.TaskComments.objects.create(
			task_pk = task,
			task_comment_text = task_comment,
			task_comment_user = username
		)

	return redirect('task-router', task_id)

def remove_task_comment_view(request):
	task_id = request.POST.get('task_id')
	comment_id = request.POST.get('comment_id')
	models.TaskComments.objects.filter(pk=comment_id).update(task_comment_invalid=1)
	return redirect('task-router', task_id)

def task_completer(request, task_pk, task_id):
	print(task_pk)
	print(task_id)
	current_assignee = models.TaskManager.objects.get(pk=task_pk,task_id=task_id).task_assigned_to
	task_complete_user = request.user
	if current_assignee is not None:
		models.TaskManager.objects.filter(pk=task_pk,task_id=task_id).update(task_completion_date=timezone.now()) 
	else:
		models.TaskManager.objects.filter(pk=task_pk,task_id=task_id).update(task_completion_date=timezone.now(),task_assigned_to=task_complete_user,task_assigned_date=timezone.now())

def user_list_view(request):
	# grab the model rows (ordered by id), filter to required task and where not completed.
	queryset = models.User.objects.exclude(user_primary__primary_team__team_name='Server').annotate(task_count=Count("assigned_tasks",distinct=True)).order_by('username','user_primary__primary_team__team_name')
	teams = models.TaskTeams.objects.all().order_by('id')
	context = {"users": queryset, "teams":teams}
	return render(request, "enquiries/main_templates/enquiries_user_list.html", context=context)

def user_tasks_view(request, userid=None):
	# grab the model rows (ordered by id), filter to required task and where not completed.
	ec_queryset = models.EnquiryComponents.objects.filter(script_tasks__task_assigned_to=userid, script_tasks__task_completion_date__isnull=True).order_by('script_tasks__task_assigned_date')

	history_queryset = models.EnquiryComponents.objects.filter(script_tasks__task_assigned_to=userid, script_tasks__task_completion_date__isnull=False).order_by('script_tasks__task_completion_date')
	history_queryset_paged = Paginator(history_queryset,20,0,True)
	page_number = request.GET.get('page')
	page_obj = history_queryset_paged.get_page(page_number)  # returns the desired page object
	context = {"cer": ec_queryset, "history":page_obj, "original_user":int(userid)}
	if request.htmx:
		return render(request, "enquiries/htmx_partials/user_task_history.html", context=context) 
	else:
		return render(request, "enquiries/main_templates/enquiries_task_user_list.html", context=context) 

def self_assign_task_view(request, task_id=None):
	#Get username to filter tasks
	username = None
	if request.user.is_authenticated:
		username =request.user
	#Set the  task to this user
	if task_id is not None:
		models.TaskManager.objects.filter(id=task_id).update(task_assigned_to=username)
		models.TaskManager.objects.filter(id=task_id).update(task_assigned_date=timezone.now())
	redirect_address = request.POST.get('page_location')
	current_page = request.POST.get('current_page')
	if redirect_address == 'task_assignment':
		return redirect('my_tasks')
	if redirect_address == 'enquiry_detail':
		enquiry_id = request.POST.get('enquiry_id')
		return redirect('enquiries_detail', enquiry_id)
	else:
		if redirect_address.split('_')[0].upper() in ['PEACON','PDACON','GRDCHG','MRKAMD','GRDREJ']:
			return redirect(f"{reverse('task_list_enq', kwargs={'task_id':redirect_address.split('_')[0].upper()})}?page={current_page}")
			#return redirect('task_list_enq', redirect_address.split('_')[0].upper())
		elif redirect_address.split('_')[0].upper() in ['NRMACC','S3SEND','S3CONF']:
			return redirect('task_list_unpaged', redirect_address.split('_')[0].upper())
		else:
			return redirect(f"{reverse('task_list', kwargs={'task_id':redirect_address.split('_')[0].upper()})}?page={current_page}")
	# + '?page=%s' + current_page
	
def assign_task_user_view(request, user_id=None, task_id=None):
	#grab the model rows (ordered by id), filter to required task and where not completed.
	queryset = models.User.objects.filter(assigned_tasks__task_completion_date__isnull=True).exclude(user_primary__primary_team__team_name='Server').annotate(task_count=Count("assigned_tasks",distinct=True)).order_by('username','user_primary__primary_team__team_name')
	redirect_address = request.POST.get('page_location')
	enquiry_id = request.POST.get('enquiry_id')
	context = {"users": queryset, "original_user":user_id, "task_id":task_id, "redirect_address":redirect_address, "enquiry_id":enquiry_id}	
	return render(request, "enquiries/main_templates/enquiries_user_select.html", context=context)
	#return redirect('user_tasks', user_id)

def assign_task_user_selected_view(request, user_id=None, task_id=None, selected_user=None):
	models.TaskManager.objects.filter(pk=task_id).update(task_assigned_to=User.objects.get(pk=selected_user),task_queued=0,task_assigned_date=timezone.now())
	redirect_address = request.POST.get('page_location')
	if redirect_address == 'my_tasks':
		return redirect('my_tasks')
	elif redirect_address == 'task_assignment':
		return redirect('user_tasks', user_id)
	elif redirect_address == 'enquiry_detail':
		enquiry_id = request.POST.get('enquiry_id')
		return redirect('enquiries_detail', enquiry_id)
	else:
		return redirect(redirect_address)

def manual_apportionment_task(request, task_id=None):
	task_queryset = models.TaskManager.objects.get(pk=task_id)
	task_ass_code = models.EnquiryComponents.objects.get(script_tasks__pk=task_id).eps_ass_code
	task_comp_code = models.EnquiryComponents.objects.get(script_tasks__pk=task_id).eps_com_id
	examiner_queryset = models.UniqueCreditor.objects.annotate(script_count=Sum("creditors__apportion_examiner__script_marked")).filter(creditors__exm_per_details__ass_code = task_ass_code, creditors__exm_per_details__com_id = task_comp_code, creditors__currently_valid=True).order_by('creditors__exm_per_details__exm_examiner_no')
	panel_notes = ''
	if models.ExaminerPanels.objects.filter(ass_code=task_ass_code,com_id=task_comp_code).exists():
		panel_notes = models.ExaminerPanels.objects.get(ass_code=task_ass_code,com_id=task_comp_code).panel_notes
	issue_reason = None
	if models.SetIssueAudit.objects.filter(enquiry_id=task_queryset.enquiry_id).exists():
		issue_reason = models.SetIssueAudit.objects.filter(enquiry_id=task_queryset.enquiry_id).first().issue_reason
	#Check for comments on task
	task_comments = None
	if models.TaskComments.objects.filter(task_pk=task_queryset.pk).exists():
		task_comments = models.TaskComments.objects.filter(task_pk=task_queryset.pk).order_by('task_comment_creation_date')
	context = {"task_id":task_id, "task":task_queryset, "ep":examiner_queryset, "appor_count":0, "issue_reason":issue_reason, "panel_notes":panel_notes, "task_comments":task_comments}
	return render(request, "enquiries/task_singles/enquiries_task_manual_apportionment.html", context=context)

def manual_apportionment(request):
	apportion_enpe_sid = request.POST.get('enpe_sid')
	apportion_script_id = request.POST.get('script_id')
	apportion_task_id = request.POST.get('task_id')
	apportion_enquiry_id = request.POST.get('enquiry_id')

	examiner_obj = models.EnquiryPersonnel.objects.get(enpe_sid=apportion_enpe_sid)
	script_obj = models.EnquiryComponents.objects.get(ec_sid=apportion_script_id)

	if models.ScriptApportionment.objects.filter(ec_sid=apportion_script_id,apportionment_invalidated=0,script_marked=1).exists():
		print('Script already apportioned')
	else:
		models.ScriptApportionment.objects.create(
			enpe_sid = examiner_obj,
			ec_sid = script_obj
			#script_marked is default to 1
		)
		if models.EnquiryComponents.objects.get(ec_sid=apportion_script_id).erp_sid.service_code == '3':
			if not models.TaskManager.objects.filter(ec_sid=apportion_script_id, task_id='S3SEND',task_completion_date = None).exists():
				models.TaskManager.objects.create(
					enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=apportion_enquiry_id),
					ec_sid = models.EnquiryComponents.objects.get(ec_sid=apportion_script_id),
					task_id = models.TaskTypes.objects.get(task_id = 'S3SEND'),
					task_assigned_to = None,
					task_assigned_date = None,
					task_completion_date = None
				)
		else:
			if models.EnquiryComponents.objects.get(ec_sid=apportion_script_id).script_type == "RM Assessor":
				if not models.TaskManager.objects.filter(ec_sid=apportion_script_id, task_id='BOTAPP',task_completion_date = None).exists():
					models.TaskManager.objects.create(
						enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=apportion_enquiry_id),
						ec_sid = models.EnquiryComponents.objects.get(ec_sid=apportion_script_id),
						task_id = models.TaskTypes.objects.get(task_id = 'BOTAPP'),
						task_assigned_to = User.objects.get(username='RPABOT'),
						task_assigned_date = timezone.now(),
						task_completion_date = None
					)
				if not models.TaskManager.objects.filter(ec_sid=apportion_script_id, task_id='NEWMIS',task_completion_date = None).exists():
					models.TaskManager.objects.create(
						enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=apportion_enquiry_id),
						ec_sid = models.EnquiryComponents.objects.get(ec_sid=apportion_script_id),
						task_id = models.TaskTypes.objects.get(task_id = 'NEWMIS'),
						task_assigned_to = User.objects.get(username='NovaServer'),
						task_assigned_date = timezone.now(),
						task_completion_date = None
				)
				if not models.TaskManager.objects.filter(ec_sid=apportion_script_id, task_id='ESMCSV',task_completion_date = None).exists():
					models.TaskManager.objects.create(
						enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=apportion_enquiry_id),
						ec_sid = models.EnquiryComponents.objects.get(ec_sid=apportion_script_id),
						task_id = models.TaskTypes.objects.get(task_id = 'ESMCSV'),
						task_assigned_to = None,
						task_assigned_date = None,
						task_completion_date = None
				)	
			else:
				if not models.TaskManager.objects.filter(ec_sid=apportion_script_id, task_id='NRMACC',task_completion_date = None).exists():
					models.TaskManager.objects.create(
						enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=apportion_enquiry_id),
						ec_sid = models.EnquiryComponents.objects.get(ec_sid=apportion_script_id),
						task_id = models.TaskTypes.objects.get(task_id = 'NRMACC'),
						task_assigned_to = None,
						task_assigned_date = None,
						task_completion_date = None
					)		

		#complete the task
		task_completer(request,apportion_task_id,'MANAPP') 
		return redirect('my_tasks')

def nrmacc_task(request, task_id=None):
	task_queryset = models.TaskManager.objects.get(pk=task_id)
	task_ass_code = models.EnquiryComponents.objects.get(script_tasks__pk=task_id).eps_ass_code
	task_comp_code = models.EnquiryComponents.objects.get(script_tasks__pk=task_id).eps_com_id
	examiner_queryset = models.UniqueCreditor.objects.annotate(script_count=Sum("creditors__apportion_examiner__script_marked")).filter(creditors__exm_per_details__ass_code = task_ass_code, creditors__exm_per_details__com_id = task_comp_code, creditors__currently_valid=True).order_by('creditors__exm_per_details__exm_examiner_no')
	panel_notes = ''
	remapp_check = False
	if models.ExaminerPanels.objects.filter(ass_code=task_ass_code,com_id=task_comp_code).exists():
		panel_notes = models.ExaminerPanels.objects.get(ass_code=task_ass_code,com_id=task_comp_code).panel_notes
	issue_reason = None
	if models.SetIssueAudit.objects.filter(enquiry_id=task_queryset.enquiry_id).exists():
		issue_reason = models.SetIssueAudit.objects.filter(enquiry_id=task_queryset.enquiry_id).first().issue_reason
	#Check for comments on task
	task_comments = None
	if models.TaskComments.objects.filter(task_pk=task_queryset.pk).exists():
		task_comments = models.TaskComments.objects.filter(task_pk=task_queryset.pk).order_by('task_comment_creation_date')
	if models.TaskManager.objects.filter(ec_sid=task_queryset.ec_sid.ec_sid, task_id='REMAPP').exists():
		remapp_check = True
	context = {"task_id":task_id, "task":task_queryset, "ep":examiner_queryset, "panel_notes":panel_notes, "task_comments":task_comments, "issue_reason":issue_reason, "remapp_check":remapp_check}
	return render(request, "enquiries/task_singles/enquiries_task_nrmacc.html", context=context)

def nrmacc_task_complete(request):
	task_id = request.POST.get('task_id')
	apportion_enpe_sid = request.POST.get('enpe_sid')
	apportion_script_id = request.POST.get('script_id')
	task_queryset = models.TaskManager.objects.get(pk=task_id)
	if apportion_enpe_sid:
		examiner_obj = models.EnquiryPersonnel.objects.get(enpe_sid=apportion_enpe_sid)
		script_obj = models.EnquiryComponents.objects.get(ec_sid=apportion_script_id)

		models.ScriptApportionment.objects.filter(ec_sid=script_obj,apportionment_invalidated=0).update(
			enpe_sid = examiner_obj,
		)
		return redirect('nrmacc-task',task_id=task_id)
	else:
		if not models.TaskManager.objects.filter(ec_sid=apportion_script_id, task_id='NEWMIS',task_completion_date = None).exists():
			models.TaskManager.objects.create(
				enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=task_queryset.enquiry_id.enquiry_id),
				ec_sid = models.EnquiryComponents.objects.get(ec_sid=task_queryset.ec_sid.ec_sid),
				task_id = models.TaskTypes.objects.get(task_id = 'NEWMIS'),
				task_assigned_to = User.objects.get(username='NovaServer'),
				task_assigned_date = timezone.now(),
				task_completion_date = None
			)
		#complete the task
		task_completer(request,task_id,'NRMACC')
		return redirect('my_tasks')
	

def nrmscs_task(request, task_id=None):
	task_queryset = models.TaskManager.objects.get(pk=task_id)
	issue_reason = None
	if models.SetIssueAudit.objects.filter(enquiry_id=task_queryset.enquiry_id).exists():
		issue_reason = models.SetIssueAudit.objects.filter(enquiry_id=task_queryset.enquiry_id).first().issue_reason
	#Check for comments on task
	task_comments = None
	if models.TaskComments.objects.filter(task_pk=task_queryset.pk).exists():
		task_comments = models.TaskComments.objects.filter(task_pk=task_queryset.pk).order_by('task_comment_creation_date')
	context = {"task_id":task_id, "task":task_queryset, "issue_reason":issue_reason, "task_comments":task_comments}
	return render(request, "enquiries/task_singles/enquiries_task_nrmscs.html", context=context)

def nrmscs_task_complete(request):
	script_id = request.POST.get('script_id')
	task_id = request.POST.get('task_id')
	enquiry_id = request.POST.get('enquiry_id')
	script_id = models.TaskManager.objectstask_id
	if not models.TaskManager.objects.filter(ec_sid=script_id, task_id='SCRCHE').exists():
		models.TaskManager.objects.create(
			enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id),
			ec_sid = models.EnquiryComponents.objects.get(ec_sid=script_id),
			task_id = models.TaskTypes.objects.get(task_id = 'SCRCHE'),
			task_assigned_to = None,
			task_assigned_date = None,
			task_completion_date = None
		)
	#complete the task
	task_completer(request,task_id,'NRMSCS')    
	return redirect('my_tasks')


def s3send_task(request, task_id=None):
	task_queryset = models.TaskManager.objects.get(pk=task_id)
	task_ass_code = models.EnquiryComponents.objects.get(script_tasks__pk=task_id).eps_ass_code
	task_comp_code = models.EnquiryComponents.objects.get(script_tasks__pk=task_id).eps_com_id
	examiner_queryset = models.UniqueCreditor.objects.annotate(script_count=Sum("creditors__apportion_examiner__script_marked")).filter(creditors__exm_per_details__ass_code = task_ass_code, creditors__exm_per_details__com_id = task_comp_code, creditors__currently_valid=True).order_by('creditors__exm_per_details__exm_examiner_no')
	#Check for pre-emptive scaled marks
	if models.EnquiryComponentsHistory.objects.filter(ec_sid=task_queryset.ec_sid.ec_sid).exists():
		kbr = models.EnquiryComponentsHistory.objects.filter(ec_sid=task_queryset.ec_sid.ec_sid).first().kbr_code
	else:
		kbr = None
	if (kbr == 'SM' or kbr == 'PSM') and task_queryset.ec_sid.erp_sid.cer_sid.ministry_flag == 'MU':
		muprem = True
	else:
		muprem = False
	panel_notes = ''
	remapp_check = False
	if models.ExaminerPanels.objects.filter(ass_code=task_ass_code,com_id=task_comp_code).exists():
		panel_notes = models.ExaminerPanels.objects.get(ass_code=task_ass_code,com_id=task_comp_code).panel_notes
	issue_reason = None
	if models.SetIssueAudit.objects.filter(enquiry_id=task_queryset.enquiry_id).exists():
		issue_reason = models.SetIssueAudit.objects.filter(enquiry_id=task_queryset.enquiry_id).first().issue_reason
	#Check for comments on task
	task_comments = None
	if models.TaskComments.objects.filter(task_pk=task_queryset.pk).exists():
		task_comments = models.TaskComments.objects.filter(task_pk=task_queryset.pk).order_by('task_comment_creation_date')
	if models.TaskManager.objects.filter(ec_sid=task_queryset.ec_sid.ec_sid, task_id='REMAPP').exists():
		remapp_check = True
	context = {"task_id":task_id, "task":task_queryset, "ep":examiner_queryset, "panel_notes":panel_notes, "task_comments":task_comments, "issue_reason":issue_reason, "remapp_check":remapp_check, "muprem":muprem}
	return render(request, "enquiries/task_singles/enquiries_task_s3send.html", context=context)

def s3send_task_complete(request):
	task_id = request.POST.get('task_id')

	apportion_enpe_sid = request.POST.get('enpe_sid')
	apportion_script_id = request.POST.get('script_id')

	task_queryset = models.TaskManager.objects.get(pk=task_id)
	apportion_enquiry_id = task_queryset.enquiry_id.enquiry_id

	if apportion_enpe_sid:
		examiner_obj = models.EnquiryPersonnel.objects.get(enpe_sid=apportion_enpe_sid)
		script_obj = models.EnquiryComponents.objects.get(ec_sid=apportion_script_id)

		models.ScriptApportionment.objects.filter(ec_sid=script_obj,apportionment_invalidated=0).update(
			enpe_sid = examiner_obj,
		)
		return redirect('s3send-task',task_id=task_id)

	else:
		#Check for pre-emptive scaled marks
		if models.EnquiryComponentsHistory.objects.filter(ec_sid=task_queryset.ec_sid.ec_sid).exists():
			kbr = models.EnquiryComponentsHistory.objects.filter(ec_sid=task_queryset.ec_sid.ec_sid).first().kbr_code
		else:
			kbr = None
		print(kbr)
		print(task_queryset.ec_sid.erp_sid.cer_sid.ministry_flag)
		if (kbr == 'SM' or kbr == 'PSM') and task_queryset.ec_sid.erp_sid.cer_sid.ministry_flag == 'MU':
			if not models.TaskManager.objects.filter(ec_sid=apportion_script_id, task_id='MUPREX',task_completion_date = None).exists():
				models.TaskManager.objects.create(
					enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=apportion_enquiry_id),
					ec_sid = models.EnquiryComponents.objects.get(ec_sid=apportion_script_id),
					task_id = models.TaskTypes.objects.get(task_id = 'MUPREX'),
					task_assigned_to = None,
					task_assigned_date = None,
					task_completion_date = None
				)	
			if not models.TaskManager.objects.filter(ec_sid=apportion_script_id, task_id='S3CONF',task_completion_date = None).exists():
				models.TaskManager.objects.create(
					enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=apportion_enquiry_id),
					ec_sid = models.EnquiryComponents.objects.get(ec_sid=apportion_script_id),
					task_id = models.TaskTypes.objects.get(task_id = 'S3CONF'),
					task_assigned_to = None,
					task_assigned_date = None,
					task_completion_date = None
				)		
			if not models.TaskManager.objects.filter(ec_sid=apportion_script_id, task_id='MKWAIT',task_completion_date = None).exists():
				models.TaskManager.objects.create(
					enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=apportion_enquiry_id),
					ec_sid = models.EnquiryComponents.objects.get(ec_sid=apportion_script_id),
					task_id = models.TaskTypes.objects.get(task_id = 'MKWAIT'),
					task_assigned_to = None,
					task_assigned_date = None,
					task_completion_date = None
				)
		elif models.EnquiryComponents.objects.get(ec_sid=apportion_script_id).script_type == "RM Assessor":
			if not models.TaskManager.objects.filter(ec_sid=apportion_script_id, task_id='BOTAPP',task_completion_date = None).exists():
				models.TaskManager.objects.create(
					enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=apportion_enquiry_id),
					ec_sid = models.EnquiryComponents.objects.get(ec_sid=apportion_script_id),
					task_id = models.TaskTypes.objects.get(task_id = 'BOTAPP'),
					task_assigned_to = User.objects.get(username='RPABOT'),
					task_assigned_date = timezone.now(),
					task_completion_date = None
				)
			if not models.TaskManager.objects.filter(ec_sid=apportion_script_id, task_id='NEWMIS',task_completion_date = None).exists():
				models.TaskManager.objects.create(
					enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=apportion_enquiry_id),
					ec_sid = models.EnquiryComponents.objects.get(ec_sid=apportion_script_id),
					task_id = models.TaskTypes.objects.get(task_id = 'NEWMIS'),
					task_assigned_to = User.objects.get(username='NovaServer'),
					task_assigned_date = timezone.now(),
					task_completion_date = None
			)
			if not models.TaskManager.objects.filter(ec_sid=apportion_script_id, task_id='ESMCSV',task_completion_date = None).exists():
				models.TaskManager.objects.create(
					enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=apportion_enquiry_id),
					ec_sid = models.EnquiryComponents.objects.get(ec_sid=apportion_script_id),
					task_id = models.TaskTypes.objects.get(task_id = 'ESMCSV'),
					task_assigned_to = None,
					task_assigned_date = None,
					task_completion_date = None
			)	
			if not models.TaskManager.objects.filter(ec_sid=apportion_script_id, task_id='S3CONF',task_completion_date = None).exists():
				models.TaskManager.objects.create(
					enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=apportion_enquiry_id),
					ec_sid = models.EnquiryComponents.objects.get(ec_sid=apportion_script_id),
					task_id = models.TaskTypes.objects.get(task_id = 'S3CONF'),
					task_assigned_to = None,
					task_assigned_date = None,
					task_completion_date = None
				)		
		else:
			if not models.TaskManager.objects.filter(ec_sid=apportion_script_id, task_id='NEWMIS',task_completion_date = None).exists():
				models.TaskManager.objects.create(
					enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=apportion_enquiry_id),
					ec_sid = models.EnquiryComponents.objects.get(ec_sid=apportion_script_id),
					task_id = models.TaskTypes.objects.get(task_id = 'NEWMIS'),
					task_assigned_to = User.objects.get(username='NovaServer'),
					task_assigned_date = timezone.now(),
					task_completion_date = None
			)
			if not models.TaskManager.objects.filter(ec_sid=apportion_script_id, task_id='S3CONF',task_completion_date = None).exists():
				models.TaskManager.objects.create(
					enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=apportion_enquiry_id),
					ec_sid = models.EnquiryComponents.objects.get(ec_sid=apportion_script_id),
					task_id = models.TaskTypes.objects.get(task_id = 'S3CONF'),
					task_assigned_to = None,
					task_assigned_date = None,
					task_completion_date = None
				)	

		#complete the task
		task_completer(request,task_id,'S3SEND')    
		return redirect('my_tasks')

def s3conf_task(request, task_id=None):
	task_queryset = models.TaskManager.objects.get(pk=task_id)
	issue_reason = None
	if models.SetIssueAudit.objects.filter(enquiry_id=task_queryset.enquiry_id).exists():
		issue_reason = models.SetIssueAudit.objects.filter(enquiry_id=task_queryset.enquiry_id).first().issue_reason
	#Check for comments on task
	task_comments = None
	if models.TaskComments.objects.filter(task_pk=task_queryset.pk).exists():
		task_comments = models.TaskComments.objects.filter(task_pk=task_queryset.pk).order_by('task_comment_creation_date')
	context = {"task_id":task_id, "task":task_queryset, "issue_reason":issue_reason, "task_comments":task_comments}
	return render(request, "enquiries/task_singles/enquiries_task_s3conf.html", context=context)

def s3conf_task_complete(request):
	task_id = request.POST.get('task_id')
	#complete the task
	task_completer(request,task_id,'S3CONF')    
	return redirect('my_tasks')

def request_mis(request):
	return render(request, "enquiries/task_singles/enquiries_task_request_mis.html")

def request_mis_complete(request):
	sessions = str(models.EarServerSettings.objects.get(pk=1).session_id_list).split(',')
	batch_id = request.POST.get('batch_id')
	error = ""
	try:
		script_obj = models.EnquiryComponentElements.objects.get(eb_sid_id=batch_id).ec_sid
	except Exception as e:
		context = {"current_status":"Batch could not be found"}
		return render(request, "enquiries/task_singles/enquiries_task_request_mis.html", context=context)
	try:
		if models.ScriptApportionment.objects.filter(ec_sid=script_obj.ec_sid, apportionment_invalidated=0).exists():
			syll = script_obj.eps_ass_code
			comp = script_obj.eps_com_id
			centre_no = script_obj.erp_sid.cer_sid.centre_id
			cand_no = script_obj.erp_sid.eps_cand_id
			cand_name = script_obj.erp_sid.stud_name
			original_exm = models.EnquiryComponentsHistory.objects.get(ec_sid=script_obj.ec_sid).exm_position
			rev_exm = models.EnquiryPersonnelDetails.objects.filter(enpe_sid=models.ScriptApportionment.objects.filter(ec_sid=script_obj.ec_sid, apportionment_invalidated=0).first().enpe_sid,session__in=sessions).first().exm_examiner_no
			#This is all to get the scaled mark
			scale_ass_code = script_obj.eps_ass_code
			scale_comp_id = script_obj.eps_com_id
			scale_centre_no = script_obj.erp_sid.eps_centre_id
			scale_cand_no = script_obj.erp_sid.eps_cand_id
			scale_ses_id = script_obj.eps_ses_sid
			print(scale_ass_code + " " + scale_comp_id + " " + scale_centre_no + " " + scale_cand_no + " " + scale_ses_id)
			if models.ScaledMarks.objects.filter(eps_ass_code=scale_ass_code,eps_com_id=scale_comp_id,eps_cnu_id=scale_centre_no,eps_cand_no=scale_cand_no,eps_ses_sid=scale_ses_id).exists():
				original_mark = models.ScaledMarks.objects.filter(eps_ass_code=scale_ass_code,eps_com_id=scale_comp_id,eps_cnu_id=scale_centre_no,eps_cand_no=scale_cand_no,eps_ses_sid=scale_ses_id).first().scaled_mark
				if original_mark is None:
					error = "No Valid Scaled Mark"
				else:
					original_mark = int(original_mark.split('.')[0])
			else:
				error = "No Valid Scaled Mark"
		else:
			error = "Batch has no valid apportionment"
		if error == "":
			cred_no = models.ScriptApportionment.objects.filter(ec_sid=script_obj.ec_sid, apportionment_invalidated=0).first().enpe_sid.per_sid.exm_creditor_no

			new_filename = os.path.realpath("\\\\filestorage\cie\Operations\Results Team\Enquiries About Results\\0.RPA_MIS Returns\EARTemplate1.xlsx")

			workbook = load_workbook(filename=new_filename)
			sheet = workbook.active

			#Syll/Comp
			sheet["A2"] = syll + "/" + comp
			#Batch
			sheet["I2"] = batch_id
			#Centre
			sheet["A4"] = centre_no
			#Cand no
			sheet["B4"] = cand_no
			#Cand name
			sheet["C4"] = cand_name
			#Orig Exm
			sheet["D4"] = original_exm
			#Rev Exm
			sheet["E4"] = rev_exm
			#Scaled (prev) mark
			sheet["F4"] = original_mark

			#Examiners-956955_BATCH_836680_MIS
			new_filename2 = os.path.realpath("\\\\filestorage\cie\Operations\Results Team\Enquiries About Results\\0.RPA_MIS Returns\Outbound Copies\\Examiner-" + cred_no + "_" + batch_id + "_" + centre_no + "_" + cand_no + "_" + syll + "_" + comp + "_MIS.xlsx")
			print(new_filename2)
			workbook.save(filename=new_filename2)

			context = {"current_status":"Mis created for " + batch_id}
			return render(request, "enquiries/task_singles/enquiries_task_request_mis.html", context=context)
		else:
			context = {"current_status":error}
			return render(request, "enquiries/task_singles/enquiries_task_request_mis.html", context=context)

	except Exception as e:
		print(f"{e}")
		context = {"current_status":f"{e}"}
		return render(request, "enquiries/task_singles/enquiries_task_request_mis.html", context=context)


	
def manual_mis(request):
	return render(request, "enquiries/task_singles/enquiries_task_manual_mis.html")

def manual_mis_complete(request):
	batch_id = request.POST.get('batch_id')
	original_exm = request.POST.get('original_exm')
	rev_exm = request.POST.get('rev_exm')
	original_mark = request.POST.get('original_mark')
	mark_status = request.POST.get('mark_status')
	revised_mark = request.POST.get('revised_mark')
	justification_code = request.POST.get('justification_code')
	remark_reason = request.POST.get('remark_reason')
	remark_concern_reason = request.POST.get('remark_concern_reason')

	context = {"current_status":"Successful MIS Upload."}

	ec_sid = None
	if models.EnquiryComponentElements.objects.filter(eb_sid=batch_id).exists():
		ec_sid = models.EnquiryComponentElements.objects.filter(eb_sid=batch_id).first().ec_sid.ec_sid
		task_enquiry_id = models.EnquiryComponentElements.objects.filter(eb_sid=batch_id).first().ec_sid.erp_sid.cer_sid.enquiry_id

		task_pk = None
		expected_exm = models.EnquiryPersonnelDetails.objects.filter(enpe_sid=models.ScriptApportionment.objects.get(ec_sid=ec_sid, apportionment_invalidated=0).enpe_sid).first()
		
		if models.TaskManager.objects.filter(task_id='RETMIS', ec_sid=ec_sid, task_completion_date = None).exists():
			task_pk = models.TaskManager.objects.filter(task_id='RETMIS', ec_sid=ec_sid).first().pk
			if task_pk is not None:
				if models.MisReturnData.objects.filter(ec_sid=ec_sid).exists():
					models.MisReturnData.objects.filter(ec_sid=ec_sid).update(
						eb_sid = models.EnquiryBatches.objects.get(eb_sid=batch_id),
						ec_sid = models.EnquiryComponents.objects.get(ec_sid=ec_sid),
						original_exm = original_exm,
						rev_exm = rev_exm,
						original_mark = original_mark,
						mark_status = mark_status,
						revised_mark = revised_mark,
						justification_code = justification_code,
						remark_reason = remark_reason,
						remark_concern_reason = remark_concern_reason,
					)
				else:
					models.MisReturnData.objects.create(
						eb_sid = models.EnquiryBatches.objects.get(eb_sid=batch_id),
						ec_sid = models.EnquiryComponents.objects.get(ec_sid=ec_sid),
						original_exm = original_exm,
						rev_exm = rev_exm,
						original_mark = original_mark,
						mark_status = mark_status,
						revised_mark = revised_mark,
						justification_code = justification_code,
						remark_reason = remark_reason,
						remark_concern_reason = remark_concern_reason,
					)

				#Create next step in chain (MISVRM), now split if SEAB to allow TL pickup
				if models.CentreEnquiryRequests.objects.get(enquiry_id=task_enquiry_id).ministry_flag == 'S':
					models.TaskManager.objects.create(
						enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=task_enquiry_id),
						ec_sid = models.EnquiryComponents.objects.get(ec_sid=ec_sid),
						task_id = models.TaskTypes.objects.get(task_id = 'MISVRF'),
						task_assigned_to = None,
						task_assigned_date = None,
						task_completion_date = None
					)
					mis_data = models.MisReturnData.objects.filter(ec_sid=ec_sid).first()
					mis_data.error_status = "SEAB Component"
					mis_data.save()
				else:
					models.TaskManager.objects.create(
						enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=task_enquiry_id),
						ec_sid = models.EnquiryComponents.objects.get(ec_sid=ec_sid),
						task_id = models.TaskTypes.objects.get(task_id = 'MISVRM'),
						task_assigned_to = None,
						task_assigned_date = None,
						task_completion_date = None
					)
				#complete the task
				task_completer(request,task_pk,'RETMIS')
				models.ScriptApportionment.objects.filter(ec_sid=ec_sid).update(script_marked=0)

		else:
			context = {"current_status":"No Open RETMIS task for this batch."}
	else:
		context = {"current_status":"Batch ID does not exist."}

	#CHANGE THIS TO RENDER SO CAN PASS ERROR CODES
	
	return render(request, "enquiries/task_singles/enquiries_task_manual_mis.html", context=context)

def scaled_mark_entry(request):
	return render(request, "enquiries/task_singles/enquiries_task_scaled_mark_entry.html")

def scaled_mark_entry_complete(request):
	eps_ass_code = request.POST.get('eps_ass_code')
	eps_com_id = request.POST.get('eps_com_id')
	eps_cnu_id = request.POST.get('eps_cnu_id')
	eps_cand_no = request.POST.get('eps_cand_no')
	eps_ses_sid = request.POST.get('eps_ses_sid')
	scaled_mark = request.POST.get('scaled_mark')
	exm_examiner_no = request.POST.get('exm_examiner_no')
	original_exm_scaled = request.POST.get('original_exm_scaled')

	if eps_ass_code != '' and eps_com_id !='' and eps_cnu_id !='' and eps_cand_no !='' and eps_ses_sid !='' and scaled_mark !='' and exm_examiner_no !='' and original_exm_scaled !='':
		models.ScaledMarks.objects.create(
			eps_ass_code = eps_ass_code,
			eps_com_id = eps_com_id,
			eps_cnu_id = eps_cnu_id,
			eps_cand_no = eps_cand_no,
			eps_ses_sid = eps_ses_sid,
			raw_mark = None,
			assessor_mark  = None,
			final_mark = None,
			exm_examiner_no = exm_examiner_no,
			scaled_mark = scaled_mark,
			original_exm_scaled = original_exm_scaled,
		)

		context = {"current_status":f'Successfully added scaled mark - {eps_ass_code} {eps_com_id} {eps_cnu_id} {eps_cand_no}'}

	else:
		context = {"current_status":f'Missing details, please try again.'}

	#CHANGE THIS TO RENDER SO CAN PASS ERROR CODES
	
	return render(request, "enquiries/task_singles/enquiries_task_scaled_mark_entry.html", context=context)


def misvrm_task(request, task_id=None):
	task_queryset = models.TaskManager.objects.get(pk=task_id)

	issue_reason = None
	if models.SetIssueAudit.objects.filter(enquiry_id=task_queryset.enquiry_id).exists():
		issue_reason = models.SetIssueAudit.objects.filter(enquiry_id=task_queryset.enquiry_id).first().issue_reason
	#Check for comments on task
	task_comments = None
	if models.TaskComments.objects.filter(task_pk=task_queryset.pk).exists():
		task_comments = models.TaskComments.objects.filter(task_pk=task_queryset.pk).order_by('task_comment_creation_date')
	context = {"task_id":task_id, "task":task_queryset, "issue_reason":issue_reason, "task_comments":task_comments}
	return render(request, "enquiries/task_singles/enquiries_task_misvrm.html", context=context)

def misvrm_task_complete(request):
	script_id = request.POST.get('script_id')
	task_id = request.POST.get('task_id')
	enquiry_id = request.POST.get('enquiry_id')
	new_mark = request.POST.get('new_mark')
	new_jc = request.POST.get('new_jc')
	new_status = request.POST.get('new_status')
	new_jc4r = request.POST.get('new_jc4r')
	
	misDataQC = models.MisReturnData.objects.get(ec_sid = script_id)

	if new_mark is None:
		new_mark = misDataQC.revised_mark
		new_jc = misDataQC.justification_code
		new_status = misDataQC.mark_status
		new_jc4r = misDataQC.remark_reason
	
	models.MisReturnData.objects.filter(ec_sid=script_id).update(final_mark=new_mark)
	models.MisReturnData.objects.filter(ec_sid=script_id).update(final_justification_code=new_jc)
	models.MisReturnData.objects.filter(ec_sid=script_id).update(final_mark_status=new_status)
	models.MisReturnData.objects.filter(ec_sid=script_id).update(remark_reason=new_jc4r)

	if not models.TaskManager.objects.filter(ec_sid=script_id, task_id='JUSCHE',task_completion_date = None).exists():
		models.TaskManager.objects.create(
			enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=enquiry_id),
			ec_sid = models.EnquiryComponents.objects.get(ec_sid=script_id),
			task_id = models.TaskTypes.objects.get(task_id = 'JUSCHE'),
			task_assigned_to = User.objects.get(username='NovaServer'),
			task_assigned_date = timezone.now(),
			task_completion_date = None
		)

	#complete the task
	task_completer(request,task_id,'MISVRM')    
	return redirect('my_tasks')


def misvrf_task(request, task_id=None):
	task_queryset = models.TaskManager.objects.get(pk=task_id)

	if models.TaskManager.objects.filter(task_id='MISVRM',ec_sid=task_queryset.ec_sid).exists():
		try:
			original_user = models.TaskManager.objects.get(task_id='MISVRM',ec_sid=task_queryset.ec_sid).task_assigned_to.username
		except:
			original_user = "Unknown"
			pass
	else:
		original_user = "Unknown"

	issue_reason = None
	if models.SetIssueAudit.objects.filter(enquiry_id=task_queryset.enquiry_id).exists():
		issue_reason = models.SetIssueAudit.objects.filter(enquiry_id=task_queryset.enquiry_id).first().issue_reason
	#Check for comments on task
	task_comments = None
	if models.TaskComments.objects.filter(task_pk=task_queryset.pk).exists():
		task_comments = models.TaskComments.objects.filter(task_pk=task_queryset.pk).order_by('task_comment_creation_date')
	context = {"task_id":task_id, "task":task_queryset, "issue_reason":issue_reason, "task_comments":task_comments, "original_user":original_user}
	return render(request, "enquiries/task_singles/enquiries_task_misvrf.html", context=context)

def misvrf_task_complete(request):
	script_id = request.POST.get('script_id')
	task_id = request.POST.get('task_id')
	enquiry_id = request.POST.get('enquiry_id')
	new_mark = request.POST.get('new_mark')
	new_jc = request.POST.get('new_jc')
	new_status = request.POST.get('new_status')
	new_jc4r = request.POST.get('new_jc4r')
	
	misDataQC = models.MisReturnData.objects.get(ec_sid = script_id)

	if new_mark is None:
		new_mark = misDataQC.revised_mark
		new_jc = misDataQC.justification_code
		new_status = misDataQC.mark_status
		new_jc4r = misDataQC.remark_reason
	
	models.MisReturnData.objects.filter(ec_sid=script_id).update(final_mark=new_mark)
	models.MisReturnData.objects.filter(ec_sid=script_id).update(final_justification_code=new_jc)
	models.MisReturnData.objects.filter(ec_sid=script_id).update(final_mark_status=new_status)
	models.MisReturnData.objects.filter(ec_sid=script_id).update(remark_reason=new_jc4r)

	if not models.TaskManager.objects.filter(ec_sid=script_id, task_id='JUSCHE',task_completion_date = None).exists():
		models.TaskManager.objects.create(
			enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=enquiry_id),
			ec_sid = models.EnquiryComponents.objects.get(ec_sid=script_id),
			task_id = models.TaskTypes.objects.get(task_id = 'JUSCHE'),
			task_assigned_to = User.objects.get(username='NovaServer'),
			task_assigned_date = timezone.now(),
			task_completion_date = None
		)

	#complete the task
	task_completer(request,task_id,'MISVRF')   
	return redirect('my_tasks')


def marche_task(request, task_id=None):
	task_queryset = models.TaskManager.objects.get(pk=task_id)

	if models.TaskManager.objects.filter(task_id='MISVRM',ec_sid=task_queryset.ec_sid).exists():
		try:
			original_user = models.TaskManager.objects.get(task_id='MISVRM',ec_sid=task_queryset.ec_sid).task_assigned_to.username
		except:
			original_user = None
	else:
		original_user = None

	issue_reason = None
	if models.SetIssueAudit.objects.filter(enquiry_id=task_queryset.enquiry_id).exists():
		issue_reason = models.SetIssueAudit.objects.filter(enquiry_id=task_queryset.enquiry_id).first().issue_reason
	#Check for comments on task
	task_comments = None
	if models.TaskComments.objects.filter(task_pk=task_queryset.pk).exists():
		task_comments = models.TaskComments.objects.filter(task_pk=task_queryset.pk).order_by('task_comment_creation_date')
	context = {"task_id":task_id, "task":task_queryset, "issue_reason":issue_reason, "task_comments":task_comments, "original_user":original_user}
	return render(request, "enquiries/task_singles/enquiries_task_marche.html", context=context)

def marche_task_complete(request):
	script_id = request.POST.get('script_id')
	task_id = request.POST.get('task_id')
	enquiry_id = request.POST.get('enquiry_id')
	new_mark = request.POST.get('new_mark')
	new_jc = request.POST.get('new_jc')
	new_status = request.POST.get('new_status')
	new_jc4r = request.POST.get('new_jc4r')
	
	misDataQC = models.MisReturnData.objects.get(ec_sid = script_id)

	if new_mark is None:
		new_mark = misDataQC.revised_mark
		new_jc = misDataQC.justification_code
		new_status = misDataQC.mark_status
		new_jc4r = misDataQC.remark_reason
	
	models.MisReturnData.objects.filter(ec_sid=script_id).update(final_mark=new_mark)
	models.MisReturnData.objects.filter(ec_sid=script_id).update(final_justification_code=new_jc)
	models.MisReturnData.objects.filter(ec_sid=script_id).update(final_mark_status=new_status)
	models.MisReturnData.objects.filter(ec_sid=script_id).update(remark_reason=new_jc4r)

	if not models.TaskManager.objects.filter(ec_sid=script_id, task_id='BOTMAR',task_completion_date = None).exists():
		models.TaskManager.objects.create(
			enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=enquiry_id),
			ec_sid = models.EnquiryComponents.objects.get(ec_sid=script_id),
			task_id = models.TaskTypes.objects.get(task_id = 'BOTMAR'),
			task_assigned_to = User.objects.get(username='NovaServer'),
			task_assigned_date = timezone.now(),
			task_completion_date = None
		)

	#complete the task
	task_completer(request,task_id,'MARCHE')    
	return redirect('my_tasks')


def pexmch_task(request, task_id=None):
	task_queryset = models.TaskManager.objects.get(pk=task_id)
	task_ass_code = models.EnquiryComponents.objects.get(script_tasks__pk=task_id).eps_ass_code
	task_comp_code = models.EnquiryComponents.objects.get(script_tasks__pk=task_id).eps_com_id
	ses_sid = models.EnquiryComponents.objects.get(script_tasks__pk=task_id).eps_ses_sid
	examiner_queryset = models.UniqueCreditor.objects.filter(creditors__exm_per_details__ass_code = task_ass_code, creditors__exm_per_details__com_id = task_comp_code, creditors__exm_per_details__sp_ses_sid=ses_sid).order_by('creditors__exm_per_details__exm_examiner_no')
	issue_reason = None
	if models.SetIssueAudit.objects.filter(enquiry_id=task_queryset.enquiry_id).exists():
		issue_reason = models.SetIssueAudit.objects.filter(enquiry_id=task_queryset.enquiry_id).first().issue_reason
	#Check for comments on task
	task_comments = None
	if models.TaskComments.objects.filter(task_pk=task_queryset.pk).exists():
		task_comments = models.TaskComments.objects.filter(task_pk=task_queryset.pk).order_by('task_comment_creation_date')
	context = {"task_id":task_id, "task":task_queryset, "ep":examiner_queryset, "task_comments":task_comments, "issue_reason":issue_reason}
	return render(request, "enquiries/task_singles/enquiries_task_pexmch.html", context=context)

def pexmch_task_complete(request):
	script_id = request.POST.get('script_id')
	task_id = request.POST.get('task_id')
	enquiry_id = request.POST.get('enquiry_id')
	#if not models.TaskManager.objects.filter(ec_sid=script_id, task_id='MANAPP',task_completion_date = None).exists():
	if not models.TaskManager.objects.filter(ec_sid=script_id, task_id='AUTAPP',task_completion_date = None).exists():
		for i in range(1,50):
			pexmch = request.POST.get('pexmch'+str(i))
			if pexmch:
				check = models.EnquiryComponentsPreviousExaminers.objects.create(
					cer_sid = models.CentreEnquiryRequests.objects.get(enquiry_id=enquiry_id),
					ec_sid = models.EnquiryComponents.objects.get(ec_sid=script_id),
					exm_position = pexmch
				)
		models.TaskManager.objects.create(
			enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=enquiry_id),
			ec_sid = models.EnquiryComponents.objects.get(ec_sid=script_id),
			#change to AUTAPP once testing complete
			#task_id = models.TaskTypes.objects.get(task_id = 'MANAPP'),
			task_id = models.TaskTypes.objects.get(task_id = 'AUTAPP'),
			task_assigned_to = None,
			task_assigned_date = None,
			task_completion_date = None
		)

		#complete the task
		task_completer(request,task_id,'PEXMCH')    
	return redirect('my_tasks')


def locmar_task(request, task_id=None):
	task_queryset = models.TaskManager.objects.get(pk=task_id)
	task_ass_code = models.EnquiryComponents.objects.get(script_tasks__pk=task_id).eps_ass_code
	task_comp_code = models.EnquiryComponents.objects.get(script_tasks__pk=task_id).eps_com_id
	examiner_queryset = models.UniqueCreditor.objects.filter(creditors__exm_per_details__ass_code = task_ass_code, creditors__exm_per_details__com_id = task_comp_code).order_by('creditors__exm_per_details__exm_examiner_no')
	issue_reason = None
	if models.SetIssueAudit.objects.filter(enquiry_id=task_queryset.enquiry_id).exists():
		issue_reason = models.SetIssueAudit.objects.filter(enquiry_id=task_queryset.enquiry_id).first().issue_reason
	#Check for comments on task
	task_comments = None
	if models.TaskComments.objects.filter(task_pk=task_queryset.pk).exists():
		task_comments = models.TaskComments.objects.filter(task_pk=task_queryset.pk).order_by('task_comment_creation_date')
	context = {"task_id":task_id, "task":task_queryset, "ep":examiner_queryset, "task_comments":task_comments, "issue_reason":issue_reason}
	return render(request, "enquiries/task_singles/enquiries_task_locmar.html", context=context)

def locmar_task_complete(request):
	script_id = request.POST.get('script_id')
	task_id = request.POST.get('task_id')
	enquiry_id = request.POST.get('enquiry_id')
	script_obj = models.EnquiryComponentElements.objects.get(ec_sid=script_id)
	if script_obj.eb_sid is not None:
		batch_no = script_obj.eb_sid.eb_sid
	if script_id is not None and request.method == 'POST':
		models.ScriptApportionment.objects.create(
			enpe_sid = None,
			ec_sid = models.EnquiryComponents.objects.get(ec_sid=script_id),
			script_marked = 0
		)
		if script_obj.ec_sid.erp_sid.cer_sid.ministry_flag == 'S':
			models.TaskManager.objects.create(
				enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=enquiry_id),
				ec_sid = models.EnquiryComponents.objects.get(ec_sid=script_id),
				#change to AUTAPP once testing complete
				task_id = models.TaskTypes.objects.get(task_id = 'BOTMAR'),
				task_assigned_to = User.objects.get(username='RPABOT2'),
				task_assigned_date = timezone.now(),
				task_completion_date = None
			)
			if models.MisReturnData.objects.filter(ec_sid=script_id).exists():
				models.MisReturnData.objects.filter(ec_sid=script_id).update(
					eb_sid = models.EnquiryBatches.objects.get(eb_sid=batch_no),
					ec_sid = models.EnquiryComponents.objects.get(ec_sid=script_id),
					original_exm = None,
					rev_exm = None,
					original_mark = None,
					mark_status = 'Confirmed',
					final_mark_status = 'Confirmed',
					keyed_mark_status = 'Confirmed',
					revised_mark = None,
					justification_code = None,
					remark_reason = None,
					remark_concern_reason = None,
					error_status = "Locally Marked Component",
				)
			else:
				models.MisReturnData.objects.create(
					eb_sid = models.EnquiryBatches.objects.get(eb_sid=batch_no),
					ec_sid = models.EnquiryComponents.objects.get(ec_sid=script_id),
					original_exm = None,
					rev_exm = None,
					original_mark = None,
					mark_status = 'Confirmed',
					final_mark_status = 'Confirmed',
					keyed_mark_status = 'Confirmed',
					revised_mark = None,
					justification_code = None,
					remark_reason = None,
					remark_concern_reason = None,
					error_status = "Locally Marked Component",
				)
		else:
			models.TaskManager.objects.create(
				enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=enquiry_id),
				ec_sid = models.EnquiryComponents.objects.get(ec_sid=script_id),
				#change to AUTAPP once testing complete
				task_id = models.TaskTypes.objects.get(task_id = 'MISVRF'),
				task_assigned_to = User.objects.get(id=models.TaskManager.objects.get(pk=task_id).task_assigned_to.pk),
				task_assigned_date = timezone.now(),
				task_completion_date = None
			)
			if models.MisReturnData.objects.filter(ec_sid=script_id).exists():
				models.MisReturnData.objects.filter(ec_sid=script_id).update(
					eb_sid = models.EnquiryBatches.objects.get(eb_sid=batch_no),
					ec_sid = models.EnquiryComponents.objects.get(ec_sid=script_id),
					original_exm = None,
					rev_exm = None,
					original_mark = None,
					mark_status = None,
					revised_mark = None,
					justification_code = None,
					remark_reason = None,
					remark_concern_reason = None,
					error_status = "Locally Marked Component",
				)
			else:
				models.MisReturnData.objects.create(
					eb_sid = models.EnquiryBatches.objects.get(eb_sid=batch_no),
					ec_sid = models.EnquiryComponents.objects.get(ec_sid=script_id),
					original_exm = None,
					rev_exm = None,
					original_mark = None,
					mark_status = None,
					revised_mark = None,
					justification_code = None,
					remark_reason = None,
					remark_concern_reason = None,
					error_status = "Locally Marked Component",
				)
		#complete the task
		task_completer(request,task_id,'LOCMAR')    
	return redirect('my_tasks')

def cleric_task(request, task_id=None):
	task_queryset = models.TaskManager.objects.get(pk=task_id)
	issue_reason = None
	if models.SetIssueAudit.objects.filter(enquiry_id=task_queryset.enquiry_id).exists():
		issue_reason = models.SetIssueAudit.objects.filter(enquiry_id=task_queryset.enquiry_id).first().issue_reason

	context = {"task_id":task_id, "task":task_queryset, "issue_reason":issue_reason, }
	return render(request, "enquiries/task_singles/enquiries_task_cleric.html", context=context)

def cleric_task_complete(request):
	script_id = request.POST.get('script_id')
	task_id = request.POST.get('task_id')
	enquiry_id = request.POST.get('enquiry_id')
	if script_id is not None and request.method == 'POST':
		enquiry_id = models.TaskManager.objects.filter(ec_sid=script_id,task_id='CLERIC').first().enquiry_id.enquiry_id
		#check if this enquiry already exists at GDWAIT
		gdwait_task = 0
		gdwait_task = models.TaskManager.objects.filter(enquiry_id=enquiry_id,task_id='GDWAIT').count()
		if gdwait_task == 0:
			if not models.TaskManager.objects.filter(enquiry_id=enquiry_id, task_id='GDWAIT',task_completion_date = None).exists():
				models.TaskManager.objects.create(
					enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=enquiry_id),
					ec_sid = None,
					task_id = models.TaskTypes.objects.get(task_id = 'GDWAIT'),
					task_assigned_to = User.objects.get(username='NovaServer'),
					task_assigned_date = timezone.now(),
					task_completion_date = None
				)
		script = models.EnquiryComponents.objects.get(ec_sid=script_id)
		if not models.TaskManager.objects.filter(enquiry_id=enquiry_id, task_id='SCRAUD',task_completion_date = None).exists():
			models.TaskManager.objects.create(
			enquiry_id = models.CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=enquiry_id),
			ec_sid = None,
			task_id = models.TaskTypes.objects.get(task_id = 'SCRAUD'),
			task_assigned_to = None,
			task_assigned_date = None,
			task_completion_date = None
			)
	#complete the task
	task_completer(request,task_id,'CLERIC')    
	return redirect('my_tasks')

def muprex_task(request, task_id=None):
	task_queryset = models.TaskManager.objects.get(pk=task_id)
	task_ass_code = models.EnquiryComponents.objects.get(script_tasks__pk=task_id).eps_ass_code
	task_comp_code = models.EnquiryComponents.objects.get(script_tasks__pk=task_id).eps_com_id
	examiner_queryset = models.UniqueCreditor.objects.filter(creditors__exm_per_details__ass_code = task_ass_code, creditors__exm_per_details__com_id = task_comp_code).order_by('creditors__exm_per_details__exm_examiner_no')
	issue_reason = None
	if models.SetIssueAudit.objects.filter(enquiry_id=task_queryset.enquiry_id).exists():
		issue_reason = models.SetIssueAudit.objects.filter(enquiry_id=task_queryset.enquiry_id).first().issue_reason
	#Check for comments on task
	task_comments = None
	if models.TaskComments.objects.filter(task_pk=task_queryset.pk).exists():
		task_comments = models.TaskComments.objects.filter(task_pk=task_queryset.pk).order_by('task_comment_creation_date')
	context = {"task_id":task_id, "task":task_queryset, "ep":examiner_queryset, "task_comments":task_comments, "issue_reason":issue_reason}
	return render(request, "enquiries/task_singles/enquiries_task_muprex.html", context=context)

def muprex_task_complete(request):
	task_id = request.POST.get('task_id')
	#complete the task
	task_completer(request,task_id,'MUPREX')   
	return redirect('my_tasks')

def scrche_task(request, task_id=None):
	task_queryset = models.TaskManager.objects.get(pk=task_id)
	issue_reason = None
	if models.SetIssueAudit.objects.filter(enquiry_id=task_queryset.enquiry_id).exists():
		issue_reason = models.SetIssueAudit.objects.filter(enquiry_id=task_queryset.enquiry_id).first().issue_reason

	context = {"task_id":task_id, "task":task_queryset, "issue_reason":issue_reason, }
	return render(request, "enquiries/task_singles/enquiries_task_scrche.html", context=context)

def scrche_task_complete(request):
	script_id = request.POST.get('script_id')
	task_id = request.POST.get('task_id')
	enquiry_id = request.POST.get('enquiry_id')
	if script_id is not None and request.method == 'POST':
		enquiry_id = models.TaskManager.objects.filter(ec_sid=script_id,task_id='SCRCHE').first().enquiry_id.enquiry_id
		script = models.EnquiryComponents.objects.get(ec_sid=script_id)
		if not models.TaskManager.objects.filter(enquiry_id=enquiry_id, task_id='SCRAUD',task_completion_date = None).exists():
			models.TaskManager.objects.create(
			enquiry_id = models.CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=enquiry_id),
			ec_sid = None,
			task_id = models.TaskTypes.objects.get(task_id = 'SCRAUD'),
			task_assigned_to = None,
			task_assigned_date = None,
			task_completion_date = None
			)
	#complete the task
	task_completer(request,task_id,'SCRCHE')    
	return redirect('my_tasks')

def scrreq_task(request, task_id=None):
	task_queryset = models.TaskManager.objects.get(pk=task_id)
	issue_reason = None
	if models.SetIssueAudit.objects.filter(enquiry_id=task_queryset.enquiry_id).exists():
		issue_reason = models.SetIssueAudit.objects.filter(enquiry_id=task_queryset.enquiry_id).first().issue_reason
	context = {"task_id":task_id, "task":task_queryset, "issue_reason":issue_reason, }
	return render(request, "enquiries/task_singles/enquiries_task_scrreq.html", context=context)

def scrreq_task_complete(request):
	task_id = request.POST.get('task_id')
	#complete the task
	task_completer(request,task_id,'SCRREQ')  
	return redirect('my_tasks')

def botapf_task(request, task_id=None):
	task_queryset = models.TaskManager.objects.get(pk=task_id)
	context = {"task_id":task_id, "task":task_queryset, }
	return render(request, "enquiries/task_singles/enquiries_task_botapf.html", context=context)

def botapf_task_complete(request):
	task_id = request.POST.get('task_id')
	#complete the task
	task_completer(request,task_id,'BOTAPF')    
	return redirect('my_tasks')

def botmaf_task(request, task_id=None):
	task_queryset = models.TaskManager.objects.get(pk=task_id)
	context = {"task_id":task_id, "task":task_queryset, }
	return render(request, "enquiries/task_singles/enquiries_task_botmaf.html", context=context)

def botmaf_task_complete(request):
	task_id = request.POST.get('task_id')
	script_id = models.TaskManager.objects.get(pk=task_id,task_id='BOTMAF').ec_sid.ec_sid
	if task_id is not None and request.method == 'POST':
		enquiry_id = models.TaskManager.objects.get(pk=task_id,task_id='BOTMAF').enquiry_id.enquiry_id

		#check if this enquiry already exists at GDWAIT
		gdwait_task = 0
		gdwait_task = models.TaskManager.objects.filter(enquiry_id=enquiry_id,task_id='GDWAIT').count()
		if gdwait_task == 0:
			if not models.TaskManager.objects.filter(ec_sid=models.EnquiryComponents.objects.only('ec_sid').get(ec_sid=script_id), task_id='GDWAIT',task_completion_date = None).exists():
				models.TaskManager.objects.create(
					enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=enquiry_id),
					ec_sid = None,
					task_id = models.TaskTypes.objects.get(task_id = 'GDWAIT'),
					task_assigned_to = User.objects.get(username='NovaServer'),
					task_assigned_date = timezone.now(),
					task_completion_date = None
				)
	#complete the task
	task_completer(request,task_id,'BOTMAF')   
	models.ScriptApportionment.objects.filter(ec_sid=script_id,script_marked=0,apportionment_invalidated=0).update(script_mark_entered=0)
	return redirect('my_tasks')

def exmsla_task(request, task_id=None):
	task_queryset = models.TaskManager.objects.get(pk=task_id)
	script_id = task_queryset.ec_sid
	extension_total = 0
	for e in models.ScriptApportionmentExtension.objects.filter(task_id=models.TaskManager.objects.get(ec_sid=script_id,task_id='RETMIS').pk):
		extension_total = extension_total + int(e.extension_days)
	context = {"task_id":task_id, "task":task_queryset, "ext_days":extension_total }
	return render(request, "enquiries/task_singles/enquiries_task_exmsla.html", context=context)

 
def exmsla_task_complete(request):
	script_id = request.POST.get('script_id')
	task_id = request.POST.get('task_id')
	enquiry_id = request.POST.get('enquiry_id')
	new_sla = request.POST.get('new_sla')
	per_sid = request.POST.get('per_sid')
	request_source = request.POST.get('source')
	if new_sla:
		if not models.TaskManager.objects.filter(ec_sid=models.EnquiryComponents.objects.only('ec_sid').get(ec_sid=script_id), task_id='RETMIS',task_completion_date = None).exists():
			models.ScriptApportionmentExtension.objects.create(
				ec_sid = models.EnquiryComponents.objects.get(ec_sid=script_id),
				task_id = models.TaskManager.objects.get(id=models.TaskManager.objects.get(ec_sid=script_id,task_id='RETMIS').pk),
				extension_days = new_sla
			)	
			#Recreate RETMIS task (or set to not complete)
			models.TaskManager.objects.filter(pk=models.TaskManager.objects.get(ec_sid=script_id,task_id='RETMIS').pk,task_id='RETMIS').update(
				task_completion_date = None
			)
		task_completer(request,task_id,'EXMSLA')  
	else:
		if not models.TaskManager.objects.filter(ec_sid=models.EnquiryComponents.objects.only('ec_sid').get(ec_sid=script_id), task_id='REMAPP',task_completion_date = None).exists():
			if task_id:
				models.TaskManager.objects.create(
					enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=enquiry_id),
					ec_sid = models.EnquiryComponents.objects.get(ec_sid=script_id),
					task_id = models.TaskTypes.objects.get(task_id = 'REMAPP'),
					task_assigned_to = None,
					task_assigned_date = None,
					task_completion_date = None
				)
				#invalidate current apportionement
				models.ScriptApportionment.objects.filter(ec_sid=script_id).update(apportionment_invalidated=1, script_marked=0)
				models.TaskManager.objects.filter(ec_sid=models.EnquiryComponents.objects.get(ec_sid=script_id),task_id__in=['REMAPP','REMAPF','BOTAPP','ESMCSV','NRMACC','S3SEND','EXMSLA','NEWMIS','RETMIS','MISVRM','MISVRF','JUSCHE','MARCHE','MKWAIT','BOTMAR'],task_completion_date=None).update(task_completion_date=timezone.now())
			else:
				#this is for forced reapportionments
				#close all outstanding tasks
				models.TaskManager.objects.filter(ec_sid=models.EnquiryComponents.objects.get(ec_sid=script_id),task_id__in=['REMAPP','REMAPF','BOTAPP','ESMCSV','NRMACC','S3SEND','EXMSLA','NEWMIS','RETMIS','MISVRM','MISVRF','JUSCHE','MARCHE','MKWAIT','BOTMAR'],task_completion_date=None).update(task_completion_date=timezone.now())
				models.ScriptApportionment.objects.filter(ec_sid=script_id).update(apportionment_invalidated=1, script_marked=0)
				#CHECK IF REMAPP ALREADY EXISTS
				models.TaskManager.objects.create(
					enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=enquiry_id),
					ec_sid = models.EnquiryComponents.objects.get(ec_sid=script_id),
					task_id = models.TaskTypes.objects.get(task_id = 'REMAPP'),
					task_assigned_to = None,
					task_assigned_date = None,
					task_completion_date = None
				)
			if request_source == 'enquiry':
				return redirect('enquiries_detail',enquiry_id=enquiry_id)
			else:
				return redirect('examiner_scripts',per_sid=per_sid)
			
	return redirect('my_tasks')



def remapp_task(request, task_id=None):
	task_queryset = models.TaskManager.objects.get(pk=task_id)
	task_ass_code = models.EnquiryComponents.objects.get(script_tasks__pk=task_id).eps_ass_code
	task_comp_code = models.EnquiryComponents.objects.get(script_tasks__pk=task_id).eps_com_id
	examiner_queryset = models.UniqueCreditor.objects.annotate(script_count=Sum("creditors__apportion_examiner__script_marked")).filter(creditors__exm_per_details__ass_code = task_ass_code, creditors__exm_per_details__com_id = task_comp_code, creditors__currently_valid=True).order_by('creditors__exm_per_details__exm_examiner_no')
	# for exm in examiner_queryset:
	# 	print(exm.exm_creditor_no)
	# 	enpe = models.EnquiryPersonnelDetails.objects.get(exm_creditor_no=exm.exm_creditor_no).enpe_sid.enpe_sid
	# 	print(exm.exm_creditor_no + ' ' + enpe)
	panel_notes = ''
	if models.ExaminerPanels.objects.filter(ass_code=task_ass_code,com_id=task_comp_code).exists():
		panel_notes = models.ExaminerPanels.objects.get(ass_code=task_ass_code,com_id=task_comp_code).panel_notes
	issue_reason = None
	if models.SetIssueAudit.objects.filter(enquiry_id=task_queryset.enquiry_id).exists():
		issue_reason = models.SetIssueAudit.objects.filter(enquiry_id=task_queryset.enquiry_id).first().issue_reason
	#Check for comments on task
	task_comments = None
	if models.TaskComments.objects.filter(task_pk=task_queryset.pk).exists():
		task_comments = models.TaskComments.objects.filter(task_pk=task_queryset.pk).order_by('task_comment_creation_date')

	context = {"task_id":task_id, "task":task_queryset, "ep":examiner_queryset, "appor_count":0, "issue_reason":issue_reason, "panel_notes":panel_notes, "task_comments":task_comments}
	return render(request, "enquiries/task_singles/enquiries_task_remapp.html", context=context)

def remapp_task_complete(request):
	apportion_enpe_sid = request.POST.get('enpe_sid')
	apportion_script_id = request.POST.get('script_id')
	apportion_task_id = request.POST.get('task_id')
	apportion_enquiry_id = request.POST.get('enquiry_id')

	examiner_obj = models.EnquiryPersonnel.objects.get(enpe_sid=apportion_enpe_sid)
	script_obj = models.EnquiryComponents.objects.get(ec_sid=apportion_script_id)

	if not models.TaskManager.objects.filter(ec_sid=models.EnquiryComponents.objects.only('ec_sid').get(ec_sid=apportion_script_id), task_id='REMAPF',task_completion_date = None).exists():
		models.ScriptApportionment.objects.create(
			enpe_sid = examiner_obj,
			ec_sid =  script_obj
			#script_marked is default to 1
		)
		if models.EnquiryComponents.objects.get(ec_sid=apportion_script_id).erp_sid.service_code == '3':
			if not models.TaskManager.objects.filter(ec_sid=apportion_script_id, task_id='S3SEND',task_completion_date = None).exists():
				models.TaskManager.objects.create(
					enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=apportion_enquiry_id),
					ec_sid = models.EnquiryComponents.objects.get(ec_sid=apportion_script_id),
					task_id = models.TaskTypes.objects.get(task_id = 'S3SEND'),
					task_assigned_to = None,
					task_assigned_date = None,
					task_completion_date = None
				)
		else:
			if models.EnquiryComponents.objects.get(ec_sid=apportion_script_id).script_type == "RM Assessor":
				#07/05/2024 REMAPF suppressed to only occur for RM scripts
				models.TaskManager.objects.create(
					enquiry_id = models.CentreEnquiryRequests.objects.get(enquiries__enquiry_parts__ec_sid=apportion_script_id),
					ec_sid = models.EnquiryComponents.objects.get(ec_sid=apportion_script_id),
					task_id = models.TaskTypes.objects.get(task_id = 'REMAPF'),
					task_assigned_to = None,
					task_assigned_date = None,
					task_completion_date = None
				)
				models.TaskManager.objects.create(
					enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=apportion_enquiry_id),
					ec_sid = models.EnquiryComponents.objects.get(ec_sid=apportion_script_id),
					task_id = models.TaskTypes.objects.get(task_id = 'NEWMIS'),
					task_assigned_to = User.objects.get(username='NovaServer'),
					task_assigned_date = timezone.now(),
					task_completion_date = None
				)
				models.TaskManager.objects.create(
					enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=apportion_enquiry_id),
					ec_sid = models.EnquiryComponents.objects.get(ec_sid=apportion_script_id),
					task_id = models.TaskTypes.objects.get(task_id = 'ESMCSV'),
					task_assigned_to = None,
					task_assigned_date = None,
					task_completion_date = None
				)	
			else:
				models.TaskManager.objects.create(
					enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=apportion_enquiry_id),
					ec_sid = models.EnquiryComponents.objects.get(ec_sid=apportion_script_id),
					task_id = models.TaskTypes.objects.get(task_id = 'NRMACC'),
					task_assigned_to = None,
					task_assigned_date = None,
					task_completion_date = None
				)		

	#complete the task
	task_completer(request,apportion_task_id,'REMAPP')    
	return redirect('my_tasks')

def remapf_task(request, task_id=None):
	task_queryset = models.TaskManager.objects.get(pk=task_id)
	task_ass_code = models.EnquiryComponents.objects.get(script_tasks__pk=task_id).eps_ass_code
	task_comp_code = models.EnquiryComponents.objects.get(script_tasks__pk=task_id).eps_com_id
	examiner_queryset = models.UniqueCreditor.objects.annotate(script_count=Sum("creditors__apportion_examiner__script_marked")).filter(creditors__exm_per_details__ass_code = task_ass_code, creditors__exm_per_details__com_id = task_comp_code).order_by('creditors__exm_per_details__exm_examiner_no').exclude(creditors__exm_per_details__enpe_sid__apportion_examiner__apportionment_invalidated=1)

	context = {"task_id":task_id, "task":task_queryset, "ep":examiner_queryset}
	return render(request, "enquiries/task_singles/enquiries_task_remapf.html", context=context)

def remapf_task_complete(request):
	apportion_task_id = request.POST.get('task_id')
	#complete the task
	task_completer(request,apportion_task_id,'REMAPF')   
	return redirect('my_tasks')


def negcon_task(request, task_id=None):
	task_queryset = models.TaskManager.objects.get(pk=task_id)
	issue_reason = None
	if models.SetIssueAudit.objects.filter(enquiry_id=task_queryset.enquiry_id).exists():
		issue_reason = models.SetIssueAudit.objects.filter(enquiry_id=task_queryset.enquiry_id).first().issue_reason
	#Check for comments on task
	task_comments = None
	if models.TaskComments.objects.filter(task_pk=task_queryset.pk).exists():
		task_comments = models.TaskComments.objects.filter(task_pk=task_queryset.pk).order_by('task_comment_creation_date')
	context = {"task_id":task_id, "task":task_queryset, "task_comments":task_comments, "issue_reason":issue_reason}
	return render(request, "enquiries/task_singles/enquiries_task_negcon.html", context=context)

def negcon_task_complete(request):
	task_id = request.POST.get('task_id')
	task_status = request.POST.get('task_status')
	enquiry_id = models.TaskManager.objects.get(pk=task_id).enquiry_id.enquiry_id
	if task_status == "Pass":
		if not models.TaskManager.objects.filter(enquiry_id=enquiry_id, task_id='PEACON',task_completion_date = None).exists():
			models.TaskManager.objects.create(
				enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=enquiry_id),
				ec_sid = None,
				task_id = models.TaskTypes.objects.get(task_id = 'PEACON'),
				task_assigned_to = User.objects.get(id=models.TaskManager.objects.get(pk=task_id).task_assigned_to.pk),
				task_assigned_date = None,
				task_completion_date = None
			)
	else:
		if not models.TaskManager.objects.filter(enquiry_id=enquiry_id, task_id='GRDREJ',task_completion_date = None).exists():
			new_task = models.TaskManager.objects.create(
				enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=enquiry_id),
				ec_sid = None,
				task_id = models.TaskTypes.objects.get(task_id = 'GRDREJ'),
				task_assigned_to = User.objects.get(id=models.TaskManager.objects.get(pk=task_id).task_assigned_to.pk),
				task_assigned_date = None,
				task_completion_date = None
			)
			models.GradeFailureAudit.objects.create(
				task_key = models.TaskManager.objects.get(pk=new_task.pk),
				failure_stage = models.TaskTypes.objects.get(task_id='NEGCON'),
				failure_reason = request.POST.get('rpa_fail')
			)		
	#complete the task
	task_completer(request,task_id,'NEGCON')    
	return redirect('my_tasks')

def peacon_task(request, task_id=None):
	task_queryset = models.TaskManager.objects.get(pk=task_id)
	script_requests = models.TaskManager.objects.filter(enquiry_id=task_queryset.enquiry_id.enquiry_id,task_id='SCRREQ')
	req_scripts = []
	for script in script_requests:
		req_scripts.append(script.ec_sid.ec_sid)
	issue_reason = None
	if models.SetIssueAudit.objects.filter(enquiry_id=task_queryset.enquiry_id).exists():
		issue_reason = models.SetIssueAudit.objects.filter(enquiry_id=task_queryset.enquiry_id).first().issue_reason
		#Check for comments on task
	task_comments = None
	if models.TaskComments.objects.filter(task_pk=task_queryset.pk).exists():
		task_comments = models.TaskComments.objects.filter(task_pk=task_queryset.pk).order_by('task_comment_creation_date')
	context = {"task_id":task_id, "task":task_queryset, "task_comments":task_comments, "issue_reason":issue_reason, "req_scripts":req_scripts}
	return render(request, "enquiries/task_singles/enquiries_task_peacon.html", context=context)

def peacon_task_complete(request):
	task_id = request.POST.get('task_id')
	task_status = request.POST.get('task_status')
	enquiry_id = models.TaskManager.objects.get(pk=task_id).enquiry_id.enquiry_id
	if task_status == "Pass":
		if not models.TaskManager.objects.filter(enquiry_id=enquiry_id, task_id='PDACON',task_completion_date = None).exists():
			models.TaskManager.objects.create(
				enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=enquiry_id),
				ec_sid = None,
				task_id = models.TaskTypes.objects.get(task_id = 'PDACON'),
				task_assigned_to = User.objects.get(id=models.TaskManager.objects.get(pk=task_id).task_assigned_to.pk),
				task_assigned_date = None,
				task_completion_date = None
			)
	else:
		if not models.TaskManager.objects.filter(enquiry_id=enquiry_id, task_id='GRDREJ',task_completion_date = None).exists():
			new_task = models.TaskManager.objects.create(
				enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=enquiry_id),
				ec_sid = None,
				task_id = models.TaskTypes.objects.get(task_id = 'GRDREJ'),
				task_assigned_to = User.objects.get(id=models.TaskManager.objects.get(pk=task_id).task_assigned_to.pk),
				task_assigned_date = None,
				task_completion_date = None
			)	
			models.GradeFailureAudit.objects.create(
				task_key = models.TaskManager.objects.get(pk=new_task.pk),
				failure_stage = models.TaskTypes.objects.get(task_id='PEACON'),
				failure_reason = request.POST.get('rpa_fail')
			)
	#complete the task
	task_completer(request,task_id,'PEACON')   
	return redirect('my_tasks')

def new_scrreq(request):
	task_id = request.POST.get('task_id')
	script_id = request.POST.get('script_id')
	enquiry_id = models.TaskManager.objects.get(pk=task_id).enquiry_id.enquiry_id
	if not models.TaskManager.objects.filter(ec_sid=script_id, task_id='SCRREQ',task_completion_date = None).exists():
		models.TaskManager.objects.create(
			enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=enquiry_id),
			ec_sid = models.EnquiryComponents.objects.get(ec_sid=script_id),
			task_id = models.TaskTypes.objects.get(task_id = 'SCRREQ'),
			task_assigned_to = None,
			task_assigned_date = None,
			task_completion_date = None
		)
	return redirect('peacon-task', task_id)

def pdacon_task(request, task_id=None):
	task_queryset = models.TaskManager.objects.get(pk=task_id)
	issue_reason = None
	if models.SetIssueAudit.objects.filter(enquiry_id=task_queryset.enquiry_id).exists():
		issue_reason = models.SetIssueAudit.objects.filter(enquiry_id=task_queryset.enquiry_id).first().issue_reason
	#Check for comments on task
	task_comments = None
	if models.TaskComments.objects.filter(task_pk=task_queryset.pk).exists():
		task_comments = models.TaskComments.objects.filter(task_pk=task_queryset.pk).order_by('task_comment_creation_date')
	context = {"task_id":task_id, "task":task_queryset, "task_comments":task_comments, "issue_reason":issue_reason}
	return render(request, "enquiries/task_singles/enquiries_task_pdacon.html", context=context)

def pdacon_task_complete(request):
	task_id = request.POST.get('task_id')
	task_status = request.POST.get('task_status')
	enquiry_id = models.TaskManager.objects.get(pk=task_id).enquiry_id.enquiry_id
	if task_status == "Pass":
		if not models.TaskManager.objects.filter(enquiry_id=enquiry_id, task_id='GRDCHG',task_completion_date = None).exists():
			models.TaskManager.objects.create(
				enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=enquiry_id),
				ec_sid = None,
				task_id = models.TaskTypes.objects.get(task_id = 'GRDCHG'),
				task_assigned_to = User.objects.get(id=models.TaskManager.objects.get(pk=task_id).task_assigned_to.pk),
				task_assigned_date = None,
				task_completion_date = None
			)
	else:
		if not models.TaskManager.objects.filter(enquiry_id=enquiry_id, task_id='GRDREJ',task_completion_date = None).exists():
			new_task = models.TaskManager.objects.create(
				enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=enquiry_id),
				ec_sid = None,
				task_id = models.TaskTypes.objects.get(task_id = 'GRDREJ'),
				task_assigned_to = User.objects.get(id=models.TaskManager.objects.get(pk=task_id).task_assigned_to.pk),
				task_assigned_date = None,
				task_completion_date = None
			)	
			models.GradeFailureAudit.objects.create(
				task_key = models.TaskManager.objects.get(pk=new_task.pk),
				failure_stage = models.TaskTypes.objects.get(task_id='PDACON'),
				failure_reason = request.POST.get('rpa_fail')
			)
	#complete the task
	task_completer(request,task_id,'PDACON')   
	return redirect('my_tasks')

def pdacon_task_sendback(request):
	print('sendback')
	task_id = request.POST.get('task_id')
	task_type = request.POST.get('task_type')
	enquiry_id = models.TaskManager.objects.get(pk=task_id).enquiry_id.enquiry_id
	if not models.TaskManager.objects.filter(enquiry_id=enquiry_id, task_id='PEACON',task_completion_date = None).exists():
		models.TaskManager.objects.create(
			enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=enquiry_id),
			ec_sid = None,
			task_id = models.TaskTypes.objects.get(task_id = 'PEACON'),
			task_assigned_to = User.objects.get(id=models.TaskManager.objects.get(pk=task_id).task_assigned_to.pk),
			task_assigned_date = None,
			task_completion_date = None
		)
	#complete the task
	task_completer(request,task_id,task_type)   
	return redirect('my_tasks')

def grdrej_task(request, task_id=None):
	task_queryset = models.TaskManager.objects.get(pk=task_id)
	#Check for comments on task
	task_comments = None
	if models.TaskComments.objects.filter(task_pk=task_queryset.pk).exists():
		task_comments = models.TaskComments.objects.filter(task_pk=task_queryset.pk).order_by('task_comment_creation_date')
	context = {"task_id":task_id, "task":task_queryset, "task_comments":task_comments}
	return render(request, "enquiries/task_singles/enquiries_task_grdrej.html", context=context)

def grdrej_task_complete(request):
	task_id = request.POST.get('task_id')
	task_status = request.POST.get('task_status')
	enquiry_id = models.TaskManager.objects.get(pk=task_id).enquiry_id.enquiry_id
	# if task_status == "Pass":
	if not models.TaskManager.objects.filter(enquiry_id=enquiry_id, task_id='MRKAMD',task_completion_date = None).exists():
		models.TaskManager.objects.create(
			enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=enquiry_id),
			ec_sid = None,
			task_id = models.TaskTypes.objects.get(task_id = 'MRKAMD'),
			task_assigned_to = None,
			task_assigned_date = None,
			task_completion_date = None
		)
	#complete the task
	task_completer(request,task_id,'GRDREJ')    
	return redirect('my_tasks')

def mrkamd_task(request, task_id=None):
	task_queryset = models.TaskManager.objects.get(pk=task_id)
	#Check for comments on task
	task_comments = None
	if models.TaskComments.objects.filter(task_pk=task_queryset.pk).exists():
		task_comments = models.TaskComments.objects.filter(task_pk=task_queryset.pk).order_by('task_comment_creation_date')
	context = {"task_id":task_id, "task":task_queryset, "task_comments":task_comments}
	return render(request, "enquiries/task_singles/enquiries_task_mrkamd.html", context=context)

def mrkamd_task_complete(request):
	task_id = request.POST.get('task_id')
	enquiry_id = models.TaskManager.objects.get(pk=task_id).enquiry_id.enquiry_id
	if not models.TaskManager.objects.filter(enquiry_id=enquiry_id, task_id='COMPLT',task_completion_date = None).exists():
		models.TaskManager.objects.create(
            enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=enquiry_id),
            ec_sid = None,
            task_id = models.TaskTypes.objects.get(task_id = 'COMPLT'),
            task_assigned_to = User.objects.get(username='NovaServer'),
            task_assigned_date = timezone.now(),
            task_completion_date = None
        )
	#complete the task
	task_completer(request,task_id,'MRKAMD') 
	return redirect('my_tasks')

def grdchg_task(request, task_id=None):
	task_queryset = models.TaskManager.objects.get(pk=task_id)
	#Check for comments on task
	task_comments = None
	if models.TaskComments.objects.filter(task_pk=task_queryset.pk).exists():
		task_comments = models.TaskComments.objects.filter(task_pk=task_queryset.pk).order_by('task_comment_creation_date')
	context = {"task_id":task_id, "task":task_queryset, "task_comments":task_comments}
	return render(request, "enquiries/task_singles/enquiries_task_grdchg.html", context=context)

def grdchg_task_complete(request):
	task_id = request.POST.get('task_id')
	task_status = request.POST.get('task_status')
	enquiry_id = models.TaskManager.objects.get(pk=task_id).enquiry_id.enquiry_id
	if task_status == "Pass":
		if not models.TaskManager.objects.filter(enquiry_id=enquiry_id, task_id='COMPLT',task_completion_date = None).exists():
			models.TaskManager.objects.create(
				enquiry_id_id = enquiry_id,
				ec_sid = None,
				task_id_id = 'COMPLT',
				task_assigned_to = User.objects.get(username='NovaServer'),
				task_assigned_date = timezone.now(),
				task_completion_date = timezone.now()
			)
	#complete the task
	task_completer(request,task_id,'GRDCHG')    
	return redirect('my_tasks')

def scrren_sendback_view(request):
	task_id = request.POST.get('task_id')
	print(task_id)
	enquiry_id = models.TaskManager.objects.get(pk=task_id).enquiry_id.enquiry_id
	script_id = models.TaskManager.objects.get(pk=task_id).ec_sid.ec_sid
	service_code = models.EnquiryComponents.objects.only('ec_sid').get(ec_sid=script_id).erp_sid.service_code
	if service_code == 'ASC' or service_code == 'ASR' or '1' in service_code:
		if not models.TaskManager.objects.filter(ec_sid=script_id, task_id='ESMSCR',task_completion_date = None).exists():
			models.TaskManager.objects.create(
				enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=enquiry_id),
				ec_sid = models.EnquiryComponents.objects.get(ec_sid=script_id),
				task_id = models.TaskTypes.objects.get(task_id = 'ESMSCR'),
				task_assigned_to = None,
				task_assigned_date = None,
				task_completion_date = None
			)  
	else:
		if not models.TaskManager.objects.filter(ec_sid=script_id, task_id='ESMSC2',task_completion_date = None).exists():
			models.TaskManager.objects.create(
				enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=enquiry_id),
				ec_sid = models.EnquiryComponents.objects.get(ec_sid=script_id),
				task_id = models.TaskTypes.objects.get(task_id = 'ESMSC2'),
				task_assigned_to = None,
				task_assigned_date = None,
				task_completion_date = None
			)
	#complete the task
	task_completer(request,task_id,'SCRREN')   
	return redirect('scrren_list')

def mailing_list(request):
	mailing_list = models.MailingList.objects.all().select_related('exm_creditor_no').order_by('-send_date','exm_creditor_no__exm_surname')
	context = {'mailing_list':mailing_list}
	return render(request,'enquiries/main_templates/mailing_list.html',context=context)


def enquiries_detail(request, enquiry_id=None):
	cer_queryset = None
	if enquiry_id is not None:	
		cer_queryset = models.CentreEnquiryRequests.objects.prefetch_related('enquiries__enquiry_parts__script_id__eb_sid','enquiries__enquiry_parts__apportion_script__enpe_sid__per_sid').get(enquiry_id=enquiry_id)
		task_queryset = models.TaskManager.objects.select_related('task_id','ec_sid','enquiry_id','task_assigned_to').filter(enquiry_id=enquiry_id).order_by('task_creation_date')
		excluded_task_list = ['INITCH','SETBIE','AUTAPP','BOTAPP','NEWMIS','RETMIS','JUSCHE','BOTMAR','GRDMAT','ESMCSV','ESMSCR','ESMSC2','OMRCHE','SCRREN','SCRAUD','LETSCR','OMRSCR','MKWAIT']
		complete_list = ['COMPLT','SETBIE']
		marking_list = ['NEWMIS','CLERIC','LOCMAR']
		apportionment_list = ['AUTAPP','MANAPP']
		enq_progress = None
		enq_stage = None
		if models.TaskManager.objects.filter(enquiry_id=enquiry_id,task_id__in=complete_list).exists():
			enq_progress = 100
			enq_stage = 'Complete'
		elif models.TaskManager.objects.filter(enquiry_id=enquiry_id,task_id='GDWAIT',task_completion_date__isnull=False).exists():
			enq_progress = 75
			enq_stage = 'Grading'
		elif models.TaskManager.objects.filter(enquiry_id=enquiry_id,task_id__in=apportionment_list,task_completion_date__isnull=True).exists():
			enq_progress = 25
			enq_stage = 'Apportionment'
		elif models.TaskManager.objects.filter(enquiry_id=enquiry_id,task_id__in=marking_list).exists():
			enq_progress = 50
			enq_stage = 'Marking'
		else:
			enq_progress = 0
			enq_stage = 'Initial Checks'

		#Get task_id for this enquiry if it has SETBIE
		bie_task_id = None
		if models.TaskManager.objects.filter(task_id='SETBIE',enquiry_id=enquiry_id).exists():
			bie_task_id = models.TaskManager.objects.filter(task_id='SETBIE',enquiry_id=enquiry_id).first().task_id.task_id

		#Get task_id for this enquiry if it is PAUSED
		enquiry_paused = None
		if models.PausedEnquiry.objects.filter(enquiry_id=enquiry_id).exists():
			enquiry_paused = models.PausedEnquiry.objects.get(enquiry_id=enquiry_id)

		#Get task_id for this enquiry if it is PRIORITISED
		enquiry_prioritised = None
		if models.PriorityEnquiry.objects.filter(enquiry_id=enquiry_id).exists():
			enquiry_prioritised = models.PriorityEnquiry.objects.get(enquiry_id=enquiry_id)

		#Get task_id for this enquiry if it has ISSUE
		issue_reason = None
		if models.SetIssueAudit.objects.filter(enquiry_id=enquiry_id).exists():
			issue_reason = models.SetIssueAudit.objects.filter(enquiry_id=enquiry_id).first().issue_reason

	context = {"cer": cer_queryset, "bie_status": bie_task_id, "enquiry_paused":enquiry_paused, "enquiry_prioritised":enquiry_prioritised, 
			"issue_reason":issue_reason, "tasks":task_queryset, "excluded_task_list":excluded_task_list, "enq_progress":enq_progress, "enq_stage":enq_stage}
	return render(request, "enquiries/main_templates/enquiries_detail.html", context=context)

def pause_enquiry(request, enquiry_id=None):
	if enquiry_id is not None and request.method == 'POST':	
		pause_status = request.POST.get('pause_status')
		pause_reason = request.POST.get('pause_reason')
		if pause_status == 'pause':
			models.PausedEnquiry.objects.create(
				enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=enquiry_id),
				pause_reason = pause_reason
			)
			models.TaskManager.objects.filter(enquiry_id=enquiry_id, task_completion_date=None).update(task_assigned_to=User.objects.get(username='PausedEnquiry'),task_completion_date=timezone.now(),task_assigned_date=timezone.now())
		else:
			models.PausedEnquiry.objects.filter(enquiry_id=enquiry_id).delete()
			models.TaskManager.objects.filter(enquiry_id=enquiry_id, task_assigned_to=User.objects.get(username='PausedEnquiry')).update(task_assigned_date=None,task_assigned_to=None,task_completion_date=None)

	return redirect('enquiries_detail', enquiry_id)

def prioritise_enquiry(request, enquiry_id=None):
	if enquiry_id is not None and request.method == 'POST':	
		priority_status = request.POST.get('priority_status')
		priority_reason = request.POST.get('priority_reason')
		print(priority_status)
		print(priority_reason)
		if priority_status == 'prioritise':
			models.PriorityEnquiry.objects.create(
				enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=enquiry_id),
				priority_reason = priority_reason
			)
			models.EnquiryDeadline.objects.filter(enquiry_id=enquiry_id).update(enquiry_deadline=timezone.now())
		else:
			models.PriorityEnquiry.objects.filter(enquiry_id=enquiry_id).delete()
			original_enquiry_deadline = models.EnquiryDeadline.objects.get(enquiry_id=enquiry_id).original_enquiry_deadline
			models.EnquiryDeadline.objects.filter(enquiry_id=enquiry_id).update(enquiry_deadline=original_enquiry_deadline)

	return redirect('enquiries_detail', enquiry_id)

def set_issue_enquiry(request, enquiry_id=None):
	if not models.SetIssueAudit.objects.filter(enquiry_id=enquiry_id).exists():
		models.SetIssueAudit.objects.create(
			enquiry_id = models.CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=enquiry_id),
			issue_reason = request.POST.get('rpa_fail')
		)
	else:
		models.SetIssueAudit.objects.filter(enquiry_id=enquiry_id).update(
			issue_reason = request.POST.get('rpa_fail')
		)
	return redirect('enquiries_detail', enquiry_id)

def enquiries_detail_search(request, id=None):
	enquiry_obj = None
	if request.method == "POST" and 'enquiry_id' in request.POST:
		enquiry_id = request.POST.get('enquiry_id')
		enquiry_obj = models.CentreEnquiryRequests.objects.get(enquiry_id=enquiry_id)
		print(enquiry_obj)
	return redirect('enquiries_detail', enquiry_id)

def enquiries_list_view(request):
	# grab the model rows (ordered by id), filter to required task and where not completed.
	search_q = ""
	if request.GET.get('search_query') is not None:
		search_q = request.GET.get('search_query')
	cer_queryset = models.CentreEnquiryRequests.objects.filter(Q(enquiry_id__icontains = search_q), enquiry_tasks__task_id='INITCH', enquiry_tasks__task_completion_date__isnull=True,enquiries__enquiry_parts__isnull=False).order_by('enquiry_id')
	cer_queryset_paged = Paginator(cer_queryset,10,0,True)
	page_number = request.GET.get('page')
	try:
		page_obj = cer_queryset_paged.get_page(page_number)  # returns the desired page object
	except PageNotAnInteger:
		# if page_number is not an integer then assign the first page
		page_obj = cer_queryset_paged.page(1)
	except EmptyPage:
		# if page is empty then return last page
		page_obj = cer_queryset_paged.page(cer_queryset_paged.num_pages)	
	context = {"cer": page_obj, "sq":search_q, }
	return render(request, "enquiries/task_lists/enquiries_list.html", context=context)

# def enquiries_bie_view(request):
# 	# grab the model rows (ordered by id), filter to required task and where not completed.
# 	ec_queryset = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='SETBIE', enquiry_tasks__task_completion_date__isnull=True).order_by('enquiry_id')
# 	ec_queryset_paged = Paginator(ec_queryset,10,0,True)
# 	page_number = request.GET.get('page')
# 	try:
# 		page_obj = ec_queryset_paged.get_page(page_number)  # returns the desired page object
# 	except PageNotAnInteger:
# 		# if page_number is not an integer then assign the first page
# 		page_obj = ec_queryset_paged.page(1)
# 	except EmptyPage:
# 		# if page is empty then return last page
# 		page_obj = ec_queryset_paged.page(ec_queryset_paged.num_pages)	
# 	context = {"cer": page_obj,}
# 	return render(request, "enquiries/task_lists/enquiries_list_setbie.html", context=context)

def iec_pass_view(request, enquiry_id=None):
	if enquiry_id is not None and request.method == 'POST':
		#Get scripts for this enquiry ID, this is a join from EC to ERP
		Scripts = models.EnquiryComponents.objects.filter(erp_sid__cer_sid = enquiry_id)
		for s in Scripts:
			if s.script_type == 'Multiple Choice':
				services = ['1','1S','ASC','ASR']
				if s.erp_sid.service_code in services:
					if not models.TaskManager.objects.filter(ec_sid=s.ec_sid, task_id='OMRSCR',task_completion_date = None).exists():
						models.TaskManager.objects.create(
							enquiry_id = models.CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=s.erp_sid.cer_sid.enquiry_id),
							ec_sid = models.EnquiryComponents.objects.only('ec_sid').get(ec_sid=s.ec_sid),
							task_id = models.TaskTypes.objects.get(task_id = 'OMRSCR'),
							task_assigned_to = None,
							task_assigned_date = None,
							task_completion_date = None
						)
					continue
				else:
					if not models.TaskManager.objects.filter(ec_sid=s.ec_sid, task_id='OMRCHE',task_completion_date = None).exists():
						models.TaskManager.objects.create(
						enquiry_id = models.CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=enquiry_id),
						ec_sid = models.EnquiryComponents.objects.only('ec_sid').get(ec_sid=s.ec_sid),
						task_id = models.TaskTypes.objects.get(task_id = 'OMRCHE'),
						task_assigned_to = None,
						task_assigned_date = None,
						task_completion_date = None
						)		
					continue
			#Check is ASR/ASC and send to script requesting
			if s.erp_sid.service_code == 'ASC' or s.erp_sid.service_code == 'ASR' or '1' in s.erp_sid.service_code:
				if s.script_type == 'RM Assessor':
					if not models.TaskManager.objects.filter(ec_sid=s.ec_sid, task_id='ESMSCR',task_completion_date = None).exists():
						models.TaskManager.objects.create(
							enquiry_id = models.CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=enquiry_id),
							ec_sid = models.EnquiryComponents.objects.only('ec_sid').get(ec_sid=s.ec_sid),
							task_id = models.TaskTypes.objects.get(task_id = 'ESMSCR'),
							task_assigned_to = None,
							task_assigned_date = None,
							task_completion_date = None
						)
				else:
					if not models.TaskManager.objects.filter(ec_sid=s.ec_sid, task_id='CLERIC',task_completion_date = None).exists():
						models.TaskManager.objects.create(
							enquiry_id = models.CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=enquiry_id),
							ec_sid = models.EnquiryComponents.objects.only('ec_sid').get(ec_sid=s.ec_sid),
							task_id = models.TaskTypes.objects.get(task_id = 'CLERIC'),
							task_assigned_to = None,
							task_assigned_date = None,
							task_completion_date = None
						)
			else:
				#Check for pre-emptive scaled marks
				if models.EnquiryComponentsHistory.objects.filter(ec_sid=s.ec_sid).exists():
					kbr = models.EnquiryComponentsHistory.objects.filter(ec_sid=s.ec_sid).first().kbr_code
				else:
					kbr = None
					continue
				print('KBR GO')
				if (kbr == 'SM' or kbr == 'PSM') and s.erp_sid.cer_sid.ministry_flag == 'MU':
					if s.erp_sid.service_code == '3':
						if models.EnquiryComponentsExaminerChecks.objects.filter(ec_sid = s.ec_sid).count() > 0:
							if not models.TaskManager.objects.filter(ec_sid=s.ec_sid, task_id='PEXMCH',task_completion_date = None).exists():
								models.TaskManager.objects.create(
								enquiry_id = models.CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=enquiry_id),
								ec_sid = models.EnquiryComponents.objects.only('ec_sid').get(ec_sid=s.ec_sid),
								task_id = models.TaskTypes.objects.get(task_id = 'PEXMCH'),
								task_assigned_to = None,
								task_assigned_date = None,
								task_completion_date = None
								)
						else:
						#create a new task for the next step (AUTAPP)
							#if not models.TaskManager.objects.filter(ec_sid=s.ec_sid, task_id='MANAPP',task_completion_date = None).exists():
							if not models.TaskManager.objects.filter(ec_sid=s.ec_sid, task_id='AUTAPP',task_completion_date = None).exists():
								models.TaskManager.objects.create(
									enquiry_id = models.CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=enquiry_id),
									ec_sid = models.EnquiryComponents.objects.only('ec_sid').get(ec_sid=s.ec_sid),
									#task_id = models.TaskTypes.objects.get(task_id = 'MANAPP'),
									task_id = models.TaskTypes.objects.get(task_id = 'AUTAPP'),
									task_assigned_to = None,
									task_assigned_date = None,
									task_completion_date = None
								)
								models.EnquiryComponentsPreviousExaminers.objects.create(
									cer_sid = models.CentreEnquiryRequests.objects.get(enquiry_id=enquiry_id),
									ec_sid = models.EnquiryComponents.objects.get(ec_sid=s.ec_sid),
									exm_position = models.EnquiryComponentsHistory.objects.get(ec_sid=s.ec_sid).exm_position
								)
					else:
						#Create confirmed MIS
						eb_sid = models.EnquiryComponentElements.objects.get(ec_sid=s.ec_sid).eb_sid.eb_sid
						if models.MisReturnData.objects.filter(ec_sid=s.ec_sid).exists():
							models.MisReturnData.objects.filter(ec_sid=s.ec_sid).update(
								eb_sid = models.EnquiryBatches.objects.get(eb_sid=eb_sid),
								ec_sid = models.EnquiryComponents.objects.get(ec_sid=s.ec_sid),
								original_exm = None,
								rev_exm = '01.01',
								original_mark = None,
								mark_status = 'Confirmed',
								revised_mark = 0,
								justification_code = None,
								remark_reason = None,
								remark_concern_reason = None
							)
						else:
							models.MisReturnData.objects.create(
								eb_sid = models.EnquiryBatches.objects.get(eb_sid=eb_sid),
								ec_sid = models.EnquiryComponents.objects.get(ec_sid=s.ec_sid),
								original_exm = None,
								rev_exm = '01.01',
								original_mark = None,
								mark_status = 'Confirmed',
								revised_mark = 0,
								justification_code = None,
								remark_reason = None,
								remark_concern_reason = None
							)
						#Assign script to PE
						try:
							principal_exm = models.EnquiryPersonnelDetails.objects.filter(exm_examiner_no='01.01',ass_code=s.eps_ass_code,com_id=s.eps_com_id,enpe_sid__currently_valid=True).first().enpe_sid
						except:
							principal_exm = None
						if models.ScriptApportionment.objects.filter(ec_sid=s.ec_sid,apportionment_invalidated=0,script_marked=1).exists():
							print('Script already apportioned')
						else:
							models.ScriptApportionment.objects.create(
								enpe_sid = principal_exm,
								ec_sid = s
								#script_marked is default to 1
								)
						#Create BOTAPP and MKWAIT and MUPREX
						if not models.TaskManager.objects.filter(ec_sid=s.ec_sid, task_id='BOTAPP',task_completion_date = None).exists():
							models.TaskManager.objects.create(
							enquiry_id = models.CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=enquiry_id),
							ec_sid = models.EnquiryComponents.objects.only('ec_sid').get(ec_sid=s.ec_sid),
							task_id = models.TaskTypes.objects.get(task_id = 'BOTAPP'),
							task_assigned_to = User.objects.get(username='RPABOT'),
							task_assigned_date = timezone.now(),
							task_completion_date = None
							)	
						if not models.TaskManager.objects.filter(ec_sid=s.ec_sid, task_id='MKWAIT',task_completion_date = None).exists():
							models.TaskManager.objects.create(
							enquiry_id = models.CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=enquiry_id),
							ec_sid = models.EnquiryComponents.objects.only('ec_sid').get(ec_sid=s.ec_sid),
							task_id = models.TaskTypes.objects.get(task_id = 'MKWAIT'),
							task_assigned_to = User.objects.get(username='NovaServer'),
							task_assigned_date = timezone.now(),
							task_completion_date = None
							)	
						if not models.TaskManager.objects.filter(ec_sid=s.ec_sid, task_id='MUPREX',task_completion_date = None).exists():
							models.TaskManager.objects.create(
							enquiry_id = models.CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=enquiry_id),
							ec_sid = models.EnquiryComponents.objects.only('ec_sid').get(ec_sid=s.ec_sid),
							task_id = models.TaskTypes.objects.get(task_id = 'MUPREX'),
							task_assigned_to = None,
							task_assigned_date = None,
							task_completion_date = None
							)			
							continue
				if s.script_type == 'MIC - MU' or s.script_type == 'MIC - SEAB':
					if not models.TaskManager.objects.filter(ec_sid=s.ec_sid, task_id='LOCMAR',task_completion_date = None).exists():
						models.TaskManager.objects.create(
						enquiry_id = models.CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=enquiry_id),
						ec_sid = models.EnquiryComponents.objects.only('ec_sid').get(ec_sid=s.ec_sid),
						task_id = models.TaskTypes.objects.get(task_id = 'LOCMAR'),
						task_assigned_to = None,
						task_assigned_date = None,
						task_completion_date = None
						)		
					continue		
				if models.EnquiryComponentsExaminerChecks.objects.filter(ec_sid = s.ec_sid).count() > 0:
					if not models.TaskManager.objects.filter(ec_sid=s.ec_sid, task_id='PEXMCH',task_completion_date = None).exists():
						models.TaskManager.objects.create(
						enquiry_id = models.CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=enquiry_id),
						ec_sid = models.EnquiryComponents.objects.only('ec_sid').get(ec_sid=s.ec_sid),
						task_id = models.TaskTypes.objects.get(task_id = 'PEXMCH'),
						task_assigned_to = None,
						task_assigned_date = None,
						task_completion_date = None
						)
				else:
				#create a new task for the next step (AUTAPP)
					#if not models.TaskManager.objects.filter(ec_sid=s.ec_sid, task_id='MANAPP',task_completion_date = None).exists():
					if not models.TaskManager.objects.filter(ec_sid=s.ec_sid, task_id='AUTAPP',task_completion_date = None).exists():
						models.TaskManager.objects.create(
							enquiry_id = models.CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=enquiry_id),
							ec_sid = models.EnquiryComponents.objects.only('ec_sid').get(ec_sid=s.ec_sid),
							#task_id = models.TaskTypes.objects.get(task_id = 'MANAPP'),
							task_id = models.TaskTypes.objects.get(task_id = 'AUTAPP'),
							task_assigned_to = None,
							task_assigned_date = None,
							task_completion_date = None
						)
						models.EnquiryComponentsPreviousExaminers.objects.create(
							cer_sid = models.CentreEnquiryRequests.objects.get(enquiry_id=enquiry_id),
							ec_sid = models.EnquiryComponents.objects.get(ec_sid=s.ec_sid),
							exm_position = models.EnquiryComponentsHistory.objects.get(ec_sid=s.ec_sid).exm_position
						)
		#Get username to filter tasks
		username = None
		if request.user.is_authenticated:
			username =request.user
			models.TaskManager.objects.filter(enquiry_id=enquiry_id,task_id='INITCH').update(task_assigned_to=username)
			models.TaskManager.objects.filter(enquiry_id=enquiry_id,task_id='INITCH').update(task_assigned_date=timezone.now())
		#complete the task
		models.TaskManager.objects.filter(enquiry_id=enquiry_id,task_id='INITCH').update(task_completion_date=timezone.now())
	return redirect('enquiries_list')

import itertools

def iec_pass_all_view(request):
	if request.method == 'POST':
		start_time = datetime.now()
		print("Start Time:" + str(datetime.now()))
		#Get scripts for this enquiry ID, this is a join from EC to ERP
		all_initch = models.TaskManager.objects.filter(task_id='INITCH',task_completion_date__isnull=True,enquiry_id__enquiries__enquiry_parts__isnull=False).order_by('enquiry_id').distinct()
		all_initch_ids = models.TaskManager.objects.filter(task_id='INITCH',task_completion_date__isnull=True,enquiry_id__enquiries__enquiry_parts__isnull=False).order_by('enquiry_id').distinct().values_list('enquiry_id',flat=True)
		 
		batch_size = 100
		batches = int(all_initch.count()/batch_size) + 1
		print(range(batches))
		for batch in range(batches):
			task_set = []
			batch_start = batch * batch_size
			batch_end = (batch + 1) * batch_size
			print(batch_start)
			print(batch_end)
			batch_items = list(itertools.islice(all_initch, batch_start, batch_end))
			enquiry_list = list(itertools.islice(all_initch_ids, batch_start, batch_end))
			#enquiry_list = batch.objects.value_list('enquiry_id',flat=True)
			print(list(enquiry_list))

			for task in batch_items:
				enquiry_id = task.enquiry_id.enquiry_id
				print(enquiry_id)
				#Get scripts for this enquiry ID, this is a join from EC to ERP
				Scripts = models.EnquiryComponents.objects.filter(erp_sid__cer_sid = enquiry_id)
				for s in Scripts:
					#Check is ASR/ASC and send to script requesting
					if s.script_type == 'Multiple Choice':
						services = ['1','1S','ASC','ASR']
						if s.erp_sid.service_code in services:
							#print('S1 etc')
							task_set.append(
								models.TaskManager(
									enquiry_id_id = s.erp_sid.cer_sid.enquiry_id,
									ec_sid_id = s.ec_sid,
									task_id_id = 'OMRSCR',
									task_assigned_to = None,
									task_assigned_date = None,
									task_completion_date = None
								)
							)
							continue
						else:
							#print('Multiple C')
							task_set.append(
								models.TaskManager(
								enquiry_id_id = s.erp_sid.cer_sid.enquiry_id,
								ec_sid_id = s.ec_sid,
								task_id_id = 'OMRCHE',
								task_assigned_to = None,
								task_assigned_date = None,
								task_completion_date = None
								)	
							)	
							continue	
					#Check is ASR/ASC and send to script requesting
					if s.erp_sid.service_code == 'ASC' or s.erp_sid.service_code == 'ASR' or '1' in s.erp_sid.service_code:
						if s.script_type == 'RM Assessor':
							#print('RM Script req')
							task_set.append(
							models.TaskManager(
								enquiry_id_id = s.erp_sid.cer_sid.enquiry_id,
								ec_sid_id = s.ec_sid,
								task_id_id = 'ESMSCR',
								task_assigned_to = None,
								task_assigned_date = None,
								task_completion_date = None
							)
							)
						else:
							#print('CLERIC')
							task_set.append(
							models.TaskManager(
								enquiry_id_id = s.erp_sid.cer_sid.enquiry_id,
								ec_sid_id = s.ec_sid,
								task_id_id = 'CLERIC',
								task_assigned_to = None,
								task_assigned_date = None,
								task_completion_date = None
							)
							)
					else:
						#Check for pre-emptive scaled marks
						if models.EnquiryComponentsHistory.objects.filter(ec_sid=s.ec_sid).exists():
							kbr = models.EnquiryComponentsHistory.objects.filter(ec_sid=s.ec_sid).first().kbr_code
						else:
							kbr = None
						if (kbr == 'SM' or kbr == 'PSM') and s.erp_sid.cer_sid.ministry_flag == 'MU':
							if s.erp_sid.service_code == '3':
								if models.EnquiryComponentsExaminerChecks.objects.filter(ec_sid = s.ec_sid).count() > 0:
									#print('S3 Pex')
									task_set.append(
									models.TaskManager(
									enquiry_id_id = s.erp_sid.cer_sid.enquiry_id,
									ec_sid_id = s.ec_sid,
									task_id_id = 'PEXMCH',
									task_assigned_to = None,
									task_assigned_date = None,
									task_completion_date = None
									)
									)
								else:
								#create a new task for the next step (AUTAPP)
									#print('S3 autapp')
									task_set.append(
									models.TaskManager(
										enquiry_id_id = s.erp_sid.cer_sid.enquiry_id,
										ec_sid_id = s.ec_sid,
										#task_id = models.TaskTypes.objects.get(task_id = 'MANAPP'),
										task_id_id = 'AUTAPP',
										task_assigned_to = None,
										task_assigned_date = None,
										task_completion_date = None
									)
									)
									models.EnquiryComponentsPreviousExaminers.objects.create(
										cer_sid = models.CentreEnquiryRequests.objects.get(enquiry_id=enquiry_id),
										ec_sid = models.EnquiryComponents.objects.get(ec_sid=s.ec_sid),
										exm_position = models.EnquiryComponentsHistory.objects.filter(ec_sid=s.ec_sid).first().exm_position
									)
							else:
								#Create confirmed MIS
								#print('Conf MIS')
								eb_sid = models.EnquiryComponentElements.objects.get(ec_sid=s.ec_sid).eb_sid.eb_sid
								if models.MisReturnData.objects.filter(ec_sid=s.ec_sid).exists():
									models.MisReturnData.objects.filter(ec_sid=s.ec_sid).update(
										eb_sid = models.EnquiryBatches.objects.get(eb_sid=eb_sid),
										ec_sid = models.EnquiryComponents.objects.get(ec_sid=s.ec_sid),
										original_exm = None,
										rev_exm = '01.01',
										original_mark = None,
										mark_status = 'Confirmed',
										revised_mark = 0,
										justification_code = None,
										remark_reason = None,
										remark_concern_reason = None
									)
								else:
									models.MisReturnData.objects.create(
										eb_sid = models.EnquiryBatches.objects.get(eb_sid=eb_sid),
										ec_sid = models.EnquiryComponents.objects.get(ec_sid=s.ec_sid),
										original_exm = None,
										rev_exm = '01.01',
										original_mark = None,
										mark_status = 'Confirmed',
										revised_mark = 0,
										justification_code = None,
										remark_reason = None,
										remark_concern_reason = None
									)
								#Assign script to PE
								try:
									principal_exm = models.EnquiryPersonnelDetails.objects.filter(exm_examiner_no='01.01',ass_code=s.eps_ass_code,com_id=s.eps_com_id,enpe_sid__currently_valid=True).first().enpe_sid
								except:
									principal_exm = None
								if models.ScriptApportionment.objects.filter(ec_sid=s.ec_sid,apportionment_invalidated=0,script_marked=1).exists():
									print('Script already apportioned')
								else:
									models.ScriptApportionment.objects.create(
										enpe_sid = principal_exm,
										ec_sid = s
										#script_marked is default to 1
										)
								#Create BOTAPP and MKWAIT and MUPREX
								task_set.append(
								models.TaskManager(
								enquiry_id_id = s.erp_sid.cer_sid.enquiry_id,
								ec_sid_id = s.ec_sid,
								task_id_id = 'BOTAPP',
								task_assigned_to = User.objects.get(username='RPABOT'),
								task_assigned_date = timezone.now(),
								task_completion_date = None
								)	
								)
								task_set.append(
								models.TaskManager(
								enquiry_id_id = s.erp_sid.cer_sid.enquiry_id,
								ec_sid_id = s.ec_sid,
								task_id_id = 'MKWAIT',
								task_assigned_to = User.objects.get(username='NovaServer'),
								task_assigned_date = timezone.now(),
								task_completion_date = None
								)	
								)
								task_set.append(
								models.TaskManager(
								enquiry_id_id = s.erp_sid.cer_sid.enquiry_id,
								ec_sid_id = s.ec_sid,
								task_id_id = 'MUPREX',
								task_assigned_to = None,
								task_assigned_date = None,
								task_completion_date = None
								)	
								)		
								continue
						if s.script_type == 'MIC - MU' or s.script_type == 'MIC - SEAB':
							#print('MIC')
							task_set.append(
							models.TaskManager(
							enquiry_id_id = s.erp_sid.cer_sid.enquiry_id,
							ec_sid_id = s.ec_sid,
							task_id_id = 'LOCMAR',
							task_assigned_to = None,
							task_assigned_date = None,
							task_completion_date = None
							)		
							)
							continue		
						if models.EnquiryComponentsExaminerChecks.objects.filter(ec_sid = s.ec_sid).count() > 0:
							#print('S2 pex')
							task_set.append(
							models.TaskManager(
							enquiry_id_id = s.erp_sid.cer_sid.enquiry_id,
							ec_sid_id = s.ec_sid,
							task_id_id = 'PEXMCH',
							task_assigned_to = None,
							task_assigned_date = None,
							task_completion_date = None
							)
							)
						else:
						#create a new task for the next step (AUTAPP)
							#print('S2 autapp')
							task_set.append(
							models.TaskManager(
								enquiry_id_id = s.erp_sid.cer_sid.enquiry_id,
								ec_sid_id = s.ec_sid,
								#task_id = models.TaskTypes.objects.get(task_id = 'MANAPP'),
								task_id_id = 'AUTAPP',
								task_assigned_to = None,
								task_assigned_date = None,
								task_completion_date = None
							)
							)
							models.EnquiryComponentsPreviousExaminers.objects.create(
								cer_sid = models.CentreEnquiryRequests.objects.get(enquiry_id=enquiry_id),
								ec_sid = models.EnquiryComponents.objects.get(ec_sid=s.ec_sid),
								exm_position = models.EnquiryComponentsHistory.objects.filter(ec_sid=s.ec_sid).first().exm_position
							)
			#load all new tasks to TaskManager
			print(task_set)
			models.TaskManager.objects.bulk_create(task_set)
			#complete the task
			if request.user.is_authenticated:
				username =request.user
				all_initch.filter(enquiry_id__in=list(enquiry_list)).update(task_assigned_to=username,task_assigned_date=timezone.now())
			#complete the task, without user
			all_initch.filter(enquiry_id__in=list(enquiry_list)).update(task_completion_date=timezone.now())
		end_time = datetime.now()
		print(end_time - start_time)

	return redirect('enquiries_list')

def iec_fail_view(request, enquiry_id=None):
	if enquiry_id is not None and request.method == 'POST':	
		#complete the tasks
		models.TaskManager.objects.filter(enquiry_id=enquiry_id, task_completion_date=None).update(task_assigned_to=User.objects.get(username='NovaServer'),task_completion_date=timezone.now(),task_assigned_date=timezone.now())
		#No need for script id, BIE is handled at Enquiry Level
		if not models.TaskManager.objects.filter(enquiry_id=enquiry_id, task_id='SETBIE',task_completion_date = None).exists():
			this_task = models.TaskManager.objects.create(
				enquiry_id = models.CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=enquiry_id),
				ec_sid = None,
				task_id = models.TaskTypes.objects.get(task_id = 'SETBIE'),
				task_assigned_to = User.objects.get(username='NovaServer'),
				task_assigned_date = timezone.now(),
				task_completion_date = timezone.now()
			)
			this_task.refresh_from_db()
			models.SetBIEAudit.objects.create(
				rpa_task_key = models.TaskManager.objects.get(pk=this_task.pk),
				failure_reason = request.POST.get('rpa_fail')
			)
		

	if request.POST.get('page_source')=='detail':
		return redirect('enquiries_detail', enquiry_id)
	else:
		return redirect('enquiries_list')

def iec_issue_view(request, enquiry_id=None):
	if enquiry_id is not None and request.method == 'POST':
		#Get scripts for this enquiry ID, this is a join from EC to ERP
		Scripts = models.EnquiryComponents.objects.filter(erp_sid__cer_sid = enquiry_id)
		for s in Scripts:
			if s.script_type == 'Multiple Choice':
				services = ['1','1S','ASC','ASR']
				if s.erp_sid.service_code in services:
					if not models.TaskManager.objects.filter(ec_sid=s.ec_sid, task_id='OMRSCR',task_completion_date = None).exists():
						models.TaskManager.objects.create(
							enquiry_id = models.CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=s.erp_sid.cer_sid.enquiry_id),
							ec_sid = models.EnquiryComponents.objects.only('ec_sid').get(ec_sid=s.ec_sid),
							task_id = models.TaskTypes.objects.get(task_id = 'OMRSCR'),
							task_assigned_to = None,
							task_assigned_date = None,
							task_completion_date = None
						)
					continue
				else:
					if not models.TaskManager.objects.filter(ec_sid=s.ec_sid, task_id='OMRCHE',task_completion_date = None).exists():
						models.TaskManager.objects.create(
						enquiry_id = models.CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=enquiry_id),
						ec_sid = models.EnquiryComponents.objects.only('ec_sid').get(ec_sid=s.ec_sid),
						task_id = models.TaskTypes.objects.get(task_id = 'OMRCHE'),
						task_assigned_to = None,
						task_assigned_date = None,
						task_completion_date = None
						)		
					continue	
			#Check is ASR/ASC and send to script requesting
			if s.erp_sid.service_code == 'ASC' or s.erp_sid.service_code == 'ASR' or '1' in s.erp_sid.service_code:
				if s.script_type == 'RM Assessor':
					if not models.TaskManager.objects.filter(ec_sid=s.ec_sid, task_id='ESMSCR',task_completion_date = None).exists():
						models.TaskManager.objects.create(
							enquiry_id = models.CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=enquiry_id),
							ec_sid = models.EnquiryComponents.objects.only('ec_sid').get(ec_sid=s.ec_sid),
							task_id = models.TaskTypes.objects.get(task_id = 'ESMSCR'),
							task_assigned_to = None,
							task_assigned_date = None,
							task_completion_date = None
						)
				else:
					if not models.TaskManager.objects.filter(ec_sid=s.ec_sid, task_id='CLERIC',task_completion_date = None).exists():
						models.TaskManager.objects.create(
							enquiry_id = models.CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=enquiry_id),
							ec_sid = models.EnquiryComponents.objects.only('ec_sid').get(ec_sid=s.ec_sid),
							task_id = models.TaskTypes.objects.get(task_id = 'CLERIC'),
							task_assigned_to = None,
							task_assigned_date = None,
							task_completion_date = None
						)
			else:
				#Check for pre-emptive scaled marks
				if models.EnquiryComponentsHistory.objects.filter(ec_sid=s.ec_sid).exists():
					kbr = models.EnquiryComponentsHistory.objects.filter(ec_sid=s.ec_sid).first().kbr_code
				else:
					kbr = None
				print('KBR GO')
				if (kbr == 'SM' or kbr == 'PSM') and s.erp_sid.cer_sid.ministry_flag == 'MU':
					if s.erp_sid.service_code == '3':
						if models.EnquiryComponentsExaminerChecks.objects.filter(ec_sid = s.ec_sid).count() > 0:
							if not models.TaskManager.objects.filter(ec_sid=s.ec_sid, task_id='PEXMCH',task_completion_date = None).exists():
								models.TaskManager.objects.create(
								enquiry_id = models.CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=enquiry_id),
								ec_sid = models.EnquiryComponents.objects.only('ec_sid').get(ec_sid=s.ec_sid),
								task_id = models.TaskTypes.objects.get(task_id = 'PEXMCH'),
								task_assigned_to = None,
								task_assigned_date = None,
								task_completion_date = None
								)
						else:
						#create a new task for the next step (AUTAPP)
							#if not models.TaskManager.objects.filter(ec_sid=s.ec_sid, task_id='MANAPP',task_completion_date = None).exists():
							if not models.TaskManager.objects.filter(ec_sid=s.ec_sid, task_id='AUTAPP',task_completion_date = None).exists():
								models.TaskManager.objects.create(
									enquiry_id = models.CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=enquiry_id),
									ec_sid = models.EnquiryComponents.objects.only('ec_sid').get(ec_sid=s.ec_sid),
									#task_id = models.TaskTypes.objects.get(task_id = 'MANAPP'),
									task_id = models.TaskTypes.objects.get(task_id = 'AUTAPP'),
									task_assigned_to = None,
									task_assigned_date = None,
									task_completion_date = None
								)
								models.EnquiryComponentsPreviousExaminers.objects.create(
									cer_sid = models.CentreEnquiryRequests.objects.get(enquiry_id=enquiry_id),
									ec_sid = models.EnquiryComponents.objects.get(ec_sid=s.ec_sid),
									exm_position = models.EnquiryComponentsHistory.objects.get(ec_sid=s.ec_sid).exm_position
								)
					else:
						#Create confirmed MIS
						eb_sid = models.EnquiryComponentElements.objects.get(ec_sid=s.ec_sid).eb_sid.eb_sid
						if models.MisReturnData.objects.filter(ec_sid=s.ec_sid).exists():
							models.MisReturnData.objects.filter(ec_sid=s.ec_sid).update(
								eb_sid = models.EnquiryBatches.objects.get(eb_sid=eb_sid),
								ec_sid = models.EnquiryComponents.objects.get(ec_sid=s.ec_sid),
								original_exm = None,
								rev_exm = '01.01',
								original_mark = None,
								mark_status = 'Confirmed',
								revised_mark = 0,
								justification_code = None,
								remark_reason = None,
								remark_concern_reason = None
							)
						else:
							models.MisReturnData.objects.create(
								eb_sid = models.EnquiryBatches.objects.get(eb_sid=eb_sid),
								ec_sid = models.EnquiryComponents.objects.get(ec_sid=s.ec_sid),
								original_exm = None,
								rev_exm = '01.01',
								original_mark = None,
								mark_status = 'Confirmed',
								revised_mark = 0,
								justification_code = None,
								remark_reason = None,
								remark_concern_reason = None
							)
						#Assign script to PE
						try:
							principal_exm = models.EnquiryPersonnelDetails.objects.filter(exm_examiner_no='01.01',ass_code=s.eps_ass_code,com_id=s.eps_com_id,enpe_sid__currently_valid=True).first().enpe_sid
						except:
							principal_exm = None
						if models.ScriptApportionment.objects.filter(ec_sid=s.ec_sid,apportionment_invalidated=0,script_marked=1).exists():
							print('Script already apportioned')
						else:
							models.ScriptApportionment.objects.create(
								enpe_sid = principal_exm,
								ec_sid = s
								#script_marked is default to 1
								)
						#Create BOTAPP and MKWAIT and MUPREX
						if not models.TaskManager.objects.filter(ec_sid=s.ec_sid, task_id='BOTAPP',task_completion_date = None).exists():
							models.TaskManager.objects.create(
							enquiry_id = models.CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=enquiry_id),
							ec_sid = models.EnquiryComponents.objects.only('ec_sid').get(ec_sid=s.ec_sid),
							task_id = models.TaskTypes.objects.get(task_id = 'BOTAPP'),
							task_assigned_to = User.objects.get(username='RPABOT'),
							task_assigned_date = timezone.now(),
							task_completion_date = None
							)	
						if not models.TaskManager.objects.filter(ec_sid=s.ec_sid, task_id='MKWAIT',task_completion_date = None).exists():
							models.TaskManager.objects.create(
							enquiry_id = models.CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=enquiry_id),
							ec_sid = models.EnquiryComponents.objects.only('ec_sid').get(ec_sid=s.ec_sid),
							task_id = models.TaskTypes.objects.get(task_id = 'MKWAIT'),
							task_assigned_to = User.objects.get(username='NovaServer'),
							task_assigned_date = timezone.now(),
							task_completion_date = None
							)	
						if not models.TaskManager.objects.filter(ec_sid=s.ec_sid, task_id='MUPREX',task_completion_date = None).exists():
							models.TaskManager.objects.create(
							enquiry_id = models.CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=enquiry_id),
							ec_sid = models.EnquiryComponents.objects.only('ec_sid').get(ec_sid=s.ec_sid),
							task_id = models.TaskTypes.objects.get(task_id = 'MUPREX'),
							task_assigned_to = None,
							task_assigned_date = None,
							task_completion_date = None
							)			
							continue
				if s.script_type == 'MIC - MU' or s.script_type == 'MIC - SEAB':
					if not models.TaskManager.objects.filter(ec_sid=s.ec_sid, task_id='LOCMAR',task_completion_date = None).exists():
						models.TaskManager.objects.create(
						enquiry_id = models.CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=enquiry_id),
						ec_sid = models.EnquiryComponents.objects.only('ec_sid').get(ec_sid=s.ec_sid),
						task_id = models.TaskTypes.objects.get(task_id = 'LOCMAR'),
						task_assigned_to = None,
						task_assigned_date = None,
						task_completion_date = None
						)		
					continue		
				if models.EnquiryComponentsExaminerChecks.objects.filter(ec_sid = s.ec_sid).count() > 0:
					if not models.TaskManager.objects.filter(ec_sid=s.ec_sid, task_id='PEXMCH',task_completion_date = None).exists():
						models.TaskManager.objects.create(
						enquiry_id = models.CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=enquiry_id),
						ec_sid = models.EnquiryComponents.objects.only('ec_sid').get(ec_sid=s.ec_sid),
						task_id = models.TaskTypes.objects.get(task_id = 'PEXMCH'),
						task_assigned_to = None,
						task_assigned_date = None,
						task_completion_date = None
						)
				else:
				#create a new task for the next step (AUTAPP)
					#if not models.TaskManager.objects.filter(ec_sid=s.ec_sid, task_id='MANAPP',task_completion_date = None).exists():
					if not models.TaskManager.objects.filter(ec_sid=s.ec_sid, task_id='AUTAPP',task_completion_date = None).exists():
						models.TaskManager.objects.create(
							enquiry_id = models.CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=enquiry_id),
							ec_sid = models.EnquiryComponents.objects.only('ec_sid').get(ec_sid=s.ec_sid),
							#task_id = models.TaskTypes.objects.get(task_id = 'MANAPP'),
							task_id = models.TaskTypes.objects.get(task_id = 'AUTAPP'),
							task_assigned_to = None,
							task_assigned_date = None,
							task_completion_date = None
						)
						models.EnquiryComponentsPreviousExaminers.objects.create(
							cer_sid = models.CentreEnquiryRequests.objects.get(enquiry_id=enquiry_id),
							ec_sid = models.EnquiryComponents.objects.get(ec_sid=s.ec_sid),
							exm_position = models.EnquiryComponentsHistory.objects.get(ec_sid=s.ec_sid).exm_position
						)
		#Get username to filter tasks
		username = None
		if request.user.is_authenticated:
			username =request.user
			models.TaskManager.objects.filter(enquiry_id=enquiry_id,task_id='INITCH').update(task_assigned_to=username)
			models.TaskManager.objects.filter(enquiry_id=enquiry_id,task_id='INITCH').update(task_assigned_date=timezone.now())
		#complete the task
		models.TaskManager.objects.filter(enquiry_id=enquiry_id,task_id='INITCH').update(task_completion_date=timezone.now())
		if not models.SetIssueAudit.objects.filter(enquiry_id=enquiry_id).exists():
			models.SetIssueAudit.objects.create(
				enquiry_id = models.CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=enquiry_id),
				issue_reason = request.POST.get('rpa_fail')
			)
		#complete the task
		models.TaskManager.objects.filter(enquiry_id=enquiry_id,task_id='INITCH').update(task_completion_date=timezone.now())

		return redirect('enquiries_list')

def task_list_view(request, task_id):
	search_q = ""
	if request.GET.get('search_query') is not None:
		search_q = request.GET.get('search_query')
	task_qs = task_id.upper()
	ec_queryset = models.EnquiryComponentElements.objects.prefetch_related('ec_sid__script_tasks__task_id','ec_sid__script_tasks__task_assigned_to').select_related('ec_sid__erp_sid__cer_sid','eb_sid').filter(ec_sid__script_tasks__task_id=task_qs, ec_sid__script_tasks__task_completion_date__isnull=True).order_by('ec_sid')
	ec_queryset = ec_queryset.filter(Q(ec_sid__erp_sid__cer_sid__enquiry_id__icontains = search_q))
	ec_queryset_paged = Paginator(ec_queryset,10,0,True)
	page_number = request.GET.get('page')
	try:
		page_obj = ec_queryset_paged.get_page(page_number)  # returns the desired page object
	except PageNotAnInteger:
		# if page_number is not an integer then assign the first page
		page_obj = ec_queryset_paged.page(1)
	except EmptyPage:
		# if page is empty then return last page
		page_obj = ec_queryset_paged.page(ec_queryset_paged.num_pages)	
	context = {"qs": page_obj, "task_qs":task_qs, "sq":search_q}
	return render(request, "enquiries/task_lists/task_list.html", context=context)

def task_list_enq_view(request, task_id):
	search_q = ""
	if request.GET.get('search_query') is not None:
		search_q = request.GET.get('search_query')
	task_qs = task_id.upper()
	# grab the model rows (ordered by id), filter to required task and where not completed.
	ec_queryset = models.TaskManager.objects.select_related('ec_sid__erp_sid__cer_sid','enquiry_id').filter(task_id=task_qs, task_completion_date__isnull=True).order_by('enquiry_id')
	print(ec_queryset)
	ec_queryset = ec_queryset.filter(Q(enquiry_id__enquiry_id__icontains = search_q))
	print(ec_queryset)
	ec_queryset_paged = Paginator(ec_queryset,10,0,True)
	page_number = request.GET.get('page')
	try:
		page_obj = ec_queryset_paged.get_page(page_number)  # returns the desired page object
	except PageNotAnInteger:
		# if page_number is not an integer then assign the first page
		page_obj = ec_queryset_paged.page(1)
	except EmptyPage:
		# if page is empty then return last page
		page_obj = ec_queryset_paged.page(ec_queryset_paged.num_pages)	
	context = {"qs": page_obj, "task_qs":task_qs, "sq":search_q}
	return render(request, "enquiries/task_lists/task_list_enq.html", context=context)

def task_list_unpaged_view(request, task_id):
	search_q = ""
	if request.GET.get('search_query') is not None:
		search_q = request.GET.get('search_query')
	task_qs = task_id.upper()
	ec_queryset = models.EnquiryComponentElements.objects.prefetch_related('ec_sid__script_tasks__task_id','ec_sid__script_tasks__task_assigned_to').select_related('ec_sid__erp_sid__cer_sid','eb_sid').filter(ec_sid__script_tasks__task_id=task_qs, ec_sid__script_tasks__task_completion_date__isnull=True).order_by('ec_sid')
	ec_queryset = ec_queryset.filter(Q(ec_sid__erp_sid__cer_sid__enquiry_id__icontains = search_q))
	context = {"qs": ec_queryset, "task_qs":task_qs, "sq":search_q}
	return render(request, "enquiries/task_lists/task_list_unpaged.html", context=context)

def esmcsv_list_view(request):
	ec_queryset = models.EsmcsvDownloads.objects.order_by('-uploaded_at')
	ec_queryset_paged = Paginator(ec_queryset,20,0,True)
	page_number = request.GET.get('page')
	try:
		page_obj = ec_queryset_paged.get_page(page_number)  # returns the desired page object
	except PageNotAnInteger:
		# if page_number is not an integer then assign the first page
		page_obj = ec_queryset_paged.page(1)
	except EmptyPage:
		# if page is empty then return last page
		page_obj = ec_queryset_paged.page(ec_queryset_paged.num_pages)	
	context = {"cer": page_obj,}
	return render(request, "enquiries/task_lists/enquiries_esmcsv.html", context=context)

def esmcsv_create_view(request):
	ec_queryset = models.TaskManager.objects.filter(task_id='ESMCSV', task_completion_date__isnull=True, ec_sid__script_id__eb_sid__eb_sid__isnull=False)
	if ec_queryset.count() > 0:
		file_timestamp = timezone.now().strftime("%m_%d_%Y_%H_%M_%S") + ".csv"
		file_location = os.path.join("downloads", file_timestamp).replace('\\', '/')
		print(file_location)
		#check validity of all rows
		for s in ec_queryset:
			models.ScriptApportionment.objects.get(ec_sid = s.ec_sid, apportionment_invalidated = 0)
		username = None
		if request.user.is_authenticated:
			username =request.user		
		with open(file_location, 'w', newline='') as file:
			file.truncate()
			writer = csv.writer(file)
			for s in ec_queryset:
				syllcomp = s.ec_sid.eps_ass_code + "/" + s.ec_sid.eps_com_id
				batch = models.EnquiryComponentElements.objects.get(ec_sid = s.ec_sid).eb_sid.eb_sid
				session = s.ec_sid.eps_ses_name
				candidate = s.ec_sid.erp_sid.eps_cand_id
				centre = s.ec_sid.erp_sid.eps_centre_id
				examiner_name = models.ScriptApportionment.objects.get(ec_sid = s.ec_sid, apportionment_invalidated = 0).enpe_sid.per_sid.exm_forename + " " + models.ScriptApportionment.objects.get(ec_sid = s.ec_sid, apportionment_invalidated = 0).enpe_sid.per_sid.exm_surname
				examiner_pos = models.EnquiryPersonnelDetails.objects.filter(enpe_sid = models.ScriptApportionment.objects.get(ec_sid = s.ec_sid, apportionment_invalidated = 0).enpe_sid).first().exm_examiner_no
				creditor_number = models.ScriptApportionment.objects.get(ec_sid = s.ec_sid, apportionment_invalidated = 0).enpe_sid.per_sid.exm_creditor_no

				writer.writerow([syllcomp,batch,session,candidate,centre,examiner_name, examiner_pos, creditor_number, ""])
				models.TaskManager.objects.filter(ec_sid=s.ec_sid,task_id='ESMCSV').update(task_completion_date=timezone.now())
				models.TaskManager.objects.filter(ec_sid=s.ec_sid,task_id='ESMCSV').update(task_assigned_date=timezone.now())
				models.TaskManager.objects.filter(ec_sid=s.ec_sid,task_id='ESMCSV').update(task_assigned_to=username)

		models.EsmcsvDownloads.objects.create(
			document = file_location,
			file_name = file_timestamp,
			download_count = 0,
			archive_count = 0
			)
		
			#Get username to filter tasks

		


	return redirect('esmcsv_list')

def esmcsv_download_view(request, download_id=None):
	# 
	document = models.EsmcsvDownloads.objects.get(pk=download_id).document
	downloads = int(models.EsmcsvDownloads.objects.get(pk=download_id).download_count) + 1
	models.EsmcsvDownloads.objects.filter(pk=download_id).update(download_count = str(downloads))

	return FileResponse(document, as_attachment=True)

def omrche_list_view(request):
	ec_queryset = models.OmrcheDownloads.objects.order_by('-uploaded_at')
	ec_queryset_paged = Paginator(ec_queryset,20,0,True)
	page_number = request.GET.get('page')
	try:
		page_obj = ec_queryset_paged.get_page(page_number)  # returns the desired page object
	except PageNotAnInteger:
		# if page_number is not an integer then assign the first page
		page_obj = ec_queryset_paged.page(1)
	except EmptyPage:
		# if page is empty then return last page
		page_obj = ec_queryset_paged.page(ec_queryset_paged.num_pages)	
	context = {"cer": page_obj,}
	return render(request, "enquiries/task_lists/enquiries_omrche.html", context=context)

def omrche_create_view(request):
	ec_queryset = models.TaskManager.objects.filter(task_id='OMRCHE', task_completion_date__isnull=True, ec_sid__script_id__eb_sid__eb_sid__isnull=False)
	if ec_queryset.count() > 0:
		file_timestamp = timezone.now().strftime("%m_%d_%Y_%H_%M_%S") + "_OMR.csv"
		file_location = os.path.join("downloads", file_timestamp).replace('\\', '/')
		print(file_location)
		username = None
		if request.user.is_authenticated:
			username =request.user		
		with open(file_location, 'w', newline='') as file:
			file.truncate()
			writer = csv.writer(file)
			writer.writerow(['enquiry_no','batch','centre','candidate','syllcomp','session','service'])
			for s in ec_queryset:
				enquiry_no = s.ec_sid.erp_sid.cer_sid.enquiry_id
				syllcomp = s.ec_sid.eps_ass_code + "/" + s.ec_sid.eps_com_id
				batch = models.EnquiryComponentElements.objects.get(ec_sid = s.ec_sid).eb_sid.eb_sid
				session = s.ec_sid.eps_ses_name
				candidate = s.ec_sid.erp_sid.eps_cand_id
				centre = s.ec_sid.erp_sid.eps_centre_id
				service = s.ec_sid.erp_sid.service_code
				
				writer.writerow([enquiry_no,batch,centre,candidate,syllcomp,session,service])
				models.TaskManager.objects.filter(ec_sid=s.ec_sid,task_id='OMRCHE').update(task_completion_date=timezone.now())
				models.TaskManager.objects.filter(ec_sid=s.ec_sid,task_id='OMRCHE').update(task_assigned_date=timezone.now())
				models.TaskManager.objects.filter(ec_sid=s.ec_sid,task_id='OMRCHE').update(task_assigned_to=username)

		models.OmrcheDownloads.objects.create(
			document = file_location,
			file_name = file_timestamp,
			download_count = 0,
			archive_count = 0
			)

	return redirect('omrche_list')

def omrche_download_view(request, download_id=None):
	# 
	document = models.OmrcheDownloads.objects.get(pk=download_id).document
	downloads = int(models.OmrcheDownloads.objects.get(pk=download_id).download_count) + 1
	models.OmrcheDownloads.objects.filter(pk=download_id).update(download_count = str(downloads))

	return FileResponse(document, as_attachment=True)


def omrscr_list_view(request):
	ec_queryset = models.OmrscrDownloads.objects.order_by('-uploaded_at')
	ec_queryset_paged = Paginator(ec_queryset,20,0,True)
	page_number = request.GET.get('page')
	try:
		page_obj = ec_queryset_paged.get_page(page_number)  # returns the desired page object
	except PageNotAnInteger:
		# if page_number is not an integer then assign the first page
		page_obj = ec_queryset_paged.page(1)
	except EmptyPage:
		# if page is empty then return last page
		page_obj = ec_queryset_paged.page(ec_queryset_paged.num_pages)	
	context = {"cer": page_obj,}
	return render(request, "enquiries/task_lists/enquiries_omrscr.html", context=context)

def omrscr_create_view(request):
	ec_queryset = models.TaskManager.objects.order_by('enquiry_id__enquiry_deadline__enquiry_deadline').filter(task_id='OMRSCR', task_completion_date__isnull=True, ec_sid__script_id__eb_sid__eb_sid__isnull=False)
	if ec_queryset.count() > 0:
		file_timestamp = timezone.now().strftime("%m_%d_%Y_%H_%M_%S") + "_OMRSCR.csv"
		file_location = os.path.join(settings.MEDIA_URL, "downloads", file_timestamp).replace('\\', '/')
		print(file_location)
		username = None
		if request.user.is_authenticated:
			username =request.user		
		with open(file_location, 'w', newline='') as file:
			file.truncate()
			writer = csv.writer(file)
			#writer.writerow(['enquiry_no','batch','centre','candidate','syllcomp','session','service'])
			writer.writerow(['Date Enquiry Booked','OMR Batch','OMR Position','Syllabus','Comp','Centre','Candidate Number','Batch Number'])
			for s in ec_queryset:
				enquiry_book_date = s.enquiry_id.eps_creation_date.date()
				print(s.enquiry_id.eps_creation_date)
				print(enquiry_book_date)
				omr_batch = models.EnquiryComponentsHistory.objects.get(ec_sid = s.ec_sid).omr_batch
				omr_position = models.EnquiryComponentsHistory.objects.get(ec_sid = s.ec_sid).omr_position
				syll = s.ec_sid.eps_ass_code
				comp = s.ec_sid.eps_com_id
				centre = s.ec_sid.erp_sid.eps_centre_id
				candidate = s.ec_sid.erp_sid.eps_cand_id
				batch = models.EnquiryComponentElements.objects.get(ec_sid = s.ec_sid).eb_sid.eb_sid
				
				writer.writerow([enquiry_book_date,omr_batch,omr_position,syll,comp,centre,candidate,batch])
				
				models.TaskManager.objects.filter(ec_sid=s.ec_sid,task_id='OMRSCR').update(task_completion_date=timezone.now())
				models.TaskManager.objects.filter(ec_sid=s.ec_sid,task_id='OMRSCR').update(task_assigned_date=timezone.now())
				models.TaskManager.objects.filter(ec_sid=s.ec_sid,task_id='OMRSCR').update(task_assigned_to=username)
				#Next chain step is 
				if not models.TaskManager.objects.filter(ec_sid=s.ec_sid, task_id='CLERIC',task_completion_date = None).exists():
					models.TaskManager.objects.create(
						enquiry_id = models.CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=s.enquiry_id.enquiry_id),
						ec_sid = models.EnquiryComponents.objects.only('ec_sid').get(ec_sid=s.ec_sid.ec_sid),
						task_id = models.TaskTypes.objects.get(task_id = 'CLERIC'),
						task_assigned_to = None,
						task_assigned_date = None,
						task_completion_date = None
					)

		models.OmrscrDownloads.objects.create(
			document = file_location,
			file_name = file_timestamp,
			download_count = 0,
			archive_count = 0
			)

	return redirect('omrscr_list')

def omrscr_download_view(request, download_id=None):
	# 
	document = models.OmrscrDownloads.objects.get(pk=download_id).document
	downloads = int(models.OmrscrDownloads.objects.get(pk=download_id).download_count) + 1
	models.OmrscrDownloads.objects.filter(pk=download_id).update(download_count = str(downloads))

	return FileResponse(document, as_attachment=True)

def esmscr_list_view(request):
	ec_queryset = models.EsmscrDownloads.objects.order_by('-uploaded_at')
	ec_queryset_paged = Paginator(ec_queryset,20,0,True)
	esmscr_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='ESMSCR', enquiry_tasks__task_completion_date__isnull=True).count()
	page_number = request.GET.get('page')
	try:
		page_obj = ec_queryset_paged.get_page(page_number)  # returns the desired page object
	except PageNotAnInteger:
		# if page_number is not an integer then assign the first page
		page_obj = ec_queryset_paged.page(1)
	except EmptyPage:
		# if page is empty then return last page
		page_obj = ec_queryset_paged.page(ec_queryset_paged.num_pages)	
	context = {"cer": page_obj,"esmscr_count":esmscr_count}
	return render(request, "enquiries/task_lists/enquiries_esmscr.html", context=context)

def esmscr_create_view(request):
	ec_queryset = models.TaskManager.objects.order_by('enquiry_id__enquiry_deadline__enquiry_deadline').filter(task_id='ESMSCR', task_completion_date__isnull=True)[:50]
	if ec_queryset.count() > 0:
		file_timestamp = timezone.now().strftime("%m_%d_%Y_%H_%M_%S") + ".csv"
		file_location = os.path.join("downloads", file_timestamp).replace('\\', '/')
		print(file_location)
		username = None
		if request.user.is_authenticated:
			username =request.user		
		with open(file_location, 'w', newline='') as file:
			file.truncate()
			writer = csv.writer(file)
			writer.writerow(['Assessment','Component ID','Centre','Candidate No'])
			for s in ec_queryset:
				syll = s.ec_sid.eps_ass_code
				comp = s.ec_sid.eps_com_id
				candidate = s.ec_sid.erp_sid.eps_cand_id
				centre = s.ec_sid.erp_sid.eps_centre_id
				writer.writerow([syll,comp,centre,candidate])
				models.TaskManager.objects.filter(ec_sid=s.ec_sid,task_id='ESMSCR').update(task_completion_date=timezone.now())
				models.TaskManager.objects.filter(ec_sid=s.ec_sid,task_id='ESMSCR').update(task_assigned_date=timezone.now())
				models.TaskManager.objects.filter(ec_sid=s.ec_sid,task_id='ESMSCR').update(task_assigned_to=username)
				if not models.TaskManager.objects.filter(ec_sid=s.ec_sid, task_id='SCRREN',task_completion_date = None).exists():
					models.TaskManager.objects.create(
						enquiry_id = s.enquiry_id,
						ec_sid = s.ec_sid,
						task_id = models.TaskTypes.objects.get(task_id = 'SCRREN'),
						task_assigned_to = None,
						task_assigned_date = None,
						task_completion_date = None
					)
		models.EsmscrDownloads.objects.create(
			document = file_location,
			file_name = file_timestamp,
			download_count = 0,
			archive_count = 0
			)
		
			#Get username to filter tasks

	return redirect('esmscr_list')

def esmscr_download_view(request, download_id=None):
	# 
	document = models.EsmscrDownloads.objects.get(pk=download_id).document
	downloads = int(models.EsmscrDownloads.objects.get(pk=download_id).download_count) + 1
	models.EsmscrDownloads.objects.filter(pk=download_id).update(download_count = str(downloads))

	return FileResponse(document, as_attachment=True)


def esmsc2_list_view(request):
	ec_queryset = models.Esmsc2Downloads.objects.order_by('-uploaded_at')
	ec_queryset_paged = Paginator(ec_queryset,20,0,True)
	esmsc2_count = models.CentreEnquiryRequests.objects.filter(enquiry_tasks__task_id='ESMSC2', enquiry_tasks__task_completion_date__isnull=True).count()
	page_number = request.GET.get('page')
	try:
		page_obj = ec_queryset_paged.get_page(page_number)  # returns the desired page object
	except PageNotAnInteger:
		# if page_number is not an integer then assign the first page
		page_obj = ec_queryset_paged.page(1)
	except EmptyPage:
		# if page is empty then return last page
		page_obj = ec_queryset_paged.page(ec_queryset_paged.num_pages)	
	context = {"cer": page_obj,"esmsc2_count":esmsc2_count}
	return render(request, "enquiries/task_lists/enquiries_esmsc2.html", context=context)

def esmsc2_create_view(request):
	ec_queryset = models.TaskManager.objects.order_by('enquiry_id__enquiry_deadline__enquiry_deadline').filter(task_id='ESMSC2', task_completion_date__isnull=True)[:50]
	if ec_queryset.count() > 0:
		file_timestamp = timezone.now().strftime("%m_%d_%Y_%H_%M_%S") + ".csv"
		file_location = os.path.join("downloads", file_timestamp).replace('\\', '/')
		print(file_location)
		username = None
		if request.user.is_authenticated:
			username =request.user		
		with open(file_location, 'w', newline='') as file:
			file.truncate()
			writer = csv.writer(file)
			writer.writerow(['Assessment','Component ID','Centre','Candidate No'])
			for s in ec_queryset:
				#Find MKWAIT completion, add 24hr delay
				mkwait_complete = models.TaskManager.objects.filter(ec_sid=s.ec_sid.ec_sid,task_id='MKWAIT').order_by('task_completion_date').first().task_completion_date
				wait_days = 0
				if timezone.now() > mkwait_complete + timedelta(days=wait_days):
					syll = s.ec_sid.eps_ass_code
					comp = s.ec_sid.eps_com_id
					candidate = s.ec_sid.erp_sid.eps_cand_id
					centre = s.ec_sid.erp_sid.eps_centre_id
					writer.writerow([syll,comp,centre,candidate])
					models.TaskManager.objects.filter(ec_sid=s.ec_sid,task_id='ESMSC2').update(task_completion_date=timezone.now())
					models.TaskManager.objects.filter(ec_sid=s.ec_sid,task_id='ESMSC2').update(task_assigned_date=timezone.now())
					models.TaskManager.objects.filter(ec_sid=s.ec_sid,task_id='ESMSC2').update(task_assigned_to=username)
					if not models.TaskManager.objects.filter(ec_sid=s.ec_sid, task_id='SCRREN',task_completion_date = None).exists():
						models.TaskManager.objects.create(
							enquiry_id = s.enquiry_id,
							ec_sid = s.ec_sid,
							task_id = models.TaskTypes.objects.get(task_id = 'SCRREN'),
							task_assigned_to = None,
							task_assigned_date = None,
							task_completion_date = None
						)
		models.Esmsc2Downloads.objects.create(
			document = file_location,
			file_name = file_timestamp,
			download_count = 0,
			archive_count = 0
			)
		
			#Get username to filter tasks


	return redirect('esmsc2_list')

def esmsc2_download_view(request, download_id=None):
	# 
	document = models.Esmsc2Downloads.objects.get(pk=download_id).document
	downloads = int(models.Esmsc2Downloads.objects.get(pk=download_id).download_count) + 1
	models.Esmsc2Downloads.objects.filter(pk=download_id).update(download_count = str(downloads))

	return FileResponse(document, as_attachment=True)

def scrren_list_view(request):
	# grab the model rows (ordered by id), filter to required task and where not completed.
	comment = models.TaskComments.objects.filter(task_pk_id=OuterRef('pk'))
	ec_queryset = models.TaskManager.objects.select_related('task_id','ec_sid__erp_sid__cer_sid').prefetch_related('ec_sid__script_id__eb_sid').annotate(comment_field=Subquery(comment.values('task_comment_text'))).filter(task_id='SCRREN', task_completion_date__isnull=True).order_by('task_creation_date')
	context = {"cer": ec_queryset,}
	return render(request, "enquiries/task_lists/enquiries_scrren.html", context=context)

def enquiries_rpa_apportion_view(request):
	# grab the model rows (ordered by id), filter to required task and where not completed.
	session_ids_string = str(models.EarServerSettings.objects.first().session_id_list)
	session_ids = (str(session_ids_string).split(','))
	ec_queryset = models.EnquiryComponentElements.objects.select_related('ec_sid__erp_sid__cer_sid','eb_sid').prefetch_related('ec_sid__apportion_script__enpe_sid__per_sid').filter(ec_sid__eps_ses_sid__in=session_ids,ec_sid__script_tasks__task_id='BOTAPP',ec_sid__script_tasks__task_completion_date__isnull=True,eb_sid__created_date__isnull=False).order_by('ec_sid__erp_sid__cer_sid__enquiry_deadline__enquiry_deadline')
	ec_queryset_paged = Paginator(ec_queryset,10,0,True)
	page_number = request.GET.get('page')
	try:
		page_obj = ec_queryset_paged.get_page(page_number)  # returns the desired page object
	except PageNotAnInteger:
		# if page_number is not an integer then assign the first page
		page_obj = ec_queryset_paged.page(1)
	except EmptyPage:
		# if page is empty then return last page
		page_obj = ec_queryset_paged.page(ec_queryset_paged.num_pages)	
	context = {"ec_queryset": page_obj,}
	return render(request, 'enquiries/rpa/rpa_apportionment.html', context=context)

def rpa_apportion_pass_view(request, script_id=None):
	if script_id is not None and request.method == 'POST':
		#Mark the task with this script ID for BOTAPP as complete
		models.TaskManager.objects.filter(ec_sid=script_id,task_id='BOTAPP').update(task_completion_date=timezone.now())
	return redirect('rpa_apportionment')

def rpa_apportion_fail_view(request, script_id=None):
	if script_id is not None and request.method == 'POST':	
		#Get enquiry for this script ID
		#create a new task for the next step (BOTAPF)
		if not models.TaskManager.objects.filter(ec_sid=models.EnquiryComponents.objects.only('ec_sid').get(ec_sid=script_id), task_id='BOTAPF',task_completion_date = None).exists():
			this_task = models.TaskManager.objects.create(
				enquiry_id = models.CentreEnquiryRequests.objects.get(enquiries__enquiry_parts__ec_sid=script_id),
				ec_sid = models.EnquiryComponents.objects.get(ec_sid=script_id),
				task_id = models.TaskTypes.objects.get(task_id = 'BOTAPF'),
				task_assigned_to = None,
				task_assigned_date = None,
				task_completion_date = None
			)
			this_task.refresh_from_db()
			print(this_task.pk)
			models.RpaFailureAudit.objects.create(
				rpa_task_key = models.TaskManager.objects.get(pk=this_task.pk),
				failure_reason = request.POST.get('rpa_fail')
			)
		#complete the task
		models.TaskManager.objects.filter(ec_sid=script_id,task_id='BOTAPP').update(task_completion_date=timezone.now())
		return redirect('rpa_apportionment')

def enquiries_rpa_marks_keying_view(request):
	# grab the model rows (ordered by id), filter to required task and where not completed.
	session_ids_string = models.EarServerSettings.objects.first().session_id_list
	session_ids = (str(session_ids_string).split(','))
	ec_queryset = models.EnquiryComponentElements.objects.select_related('ec_sid__erp_sid__cer_sid','eb_sid').prefetch_related('ec_sid__apportion_script__enpe_sid__per_sid','ec_sid__script_mis').filter(ec_sid__eps_ses_sid__in=session_ids,ec_sid__script_tasks__task_id='BOTMAR',ec_sid__script_tasks__task_completion_date__isnull=True,eb_sid__created_date__isnull=False).order_by('ec_sid__erp_sid__cer_sid__enquiry_deadline__enquiry_deadline')
	ec_queryset_paged = Paginator(ec_queryset,10,0,True)
	page_number = request.GET.get('page')
	try:
		page_obj = ec_queryset_paged.get_page(page_number)  # returns the desired page object
	except PageNotAnInteger:
		# if page_number is not an integer then assign the first page
		page_obj = ec_queryset_paged.page(1)
	except EmptyPage:
		# if page is empty then return last page
		page_obj = ec_queryset_paged.page(ec_queryset_paged.num_pages)	
	context = {"ec_queryset": page_obj,}
	return render(request, 'enquiries/rpa/rpa_marks_keying.html', context=context)

def rpa_marks_keying_pass_view(request, script_id=None):
	if script_id is not None and request.method == 'POST':
		enquiry_id = models.TaskManager.objects.filter(ec_sid=script_id,task_id='BOTMAR').first().enquiry_id.enquiry_id
		#check if this enquiry already exists at GDWAIT
		gdwait_task = 0
		gdwait_task = models.TaskManager.objects.filter(enquiry_id=enquiry_id,task_id='GDWAIT').count()
		if gdwait_task == 0:
			if not models.TaskManager.objects.filter(enquiry_id=enquiry_id, task_id='GDWAIT',task_completion_date = None).exists():
				models.TaskManager.objects.create(
					enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=enquiry_id),
					ec_sid = None,
					task_id = models.TaskTypes.objects.get(task_id = 'GDWAIT'),
					task_assigned_to = User.objects.get(username='NovaServer'),
					task_assigned_date = timezone.now(),
					task_completion_date = None
				)
		#Mark the task with this script ID for BOTMAR as complete
		models.TaskManager.objects.filter(ec_sid=script_id,task_id='BOTMAR').update(task_completion_date=timezone.now())
		models.ScriptApportionment.objects.filter(ec_sid=script_id,script_marked=0,apportionment_invalidated=0).update(script_mark_entered=0)
	return redirect('rpa_marks_keying')

def rpa_marks_keying_fail_view(request, script_id=None):
	if script_id is not None and request.method == 'POST':	
		enquiry_id = models.TaskManager.objects.filter(ec_sid=script_id,task_id='BOTMAR').first().enquiry_id.enquiry_id
		#create a new task for the next step (BOTMAF)
		if not models.TaskManager.objects.filter(enquiry_id=enquiry_id, task_id='BOTMAF',task_completion_date = None).exists():
			this_task = models.TaskManager.objects.create(
				enquiry_id = models.CentreEnquiryRequests.objects.get(enquiry_id=enquiry_id),
				ec_sid = models.EnquiryComponents.objects.get(ec_sid=script_id),
				task_id = models.TaskTypes.objects.get(task_id = 'BOTMAF'),
				task_assigned_to = None,
				task_assigned_date = None,
				task_completion_date = None
			)
			this_task.refresh_from_db()
			models.RpaFailureAudit.objects.create(
				rpa_task_key = models.TaskManager.objects.get(pk=this_task.pk),
				failure_reason = request.POST.get('rpa_fail')
			)
		#complete the task
		models.TaskManager.objects.filter(ec_sid=script_id,task_id='BOTMAR').update(task_completion_date=timezone.now())
		return redirect('rpa_marks_keying')

def reload_tolerance_view(request):
	qs = models.MarkTolerances.objects.all().order_by('eps_ass_code','eps_com_id')
	context = {"qs":qs}
	return render(request, "enquiries/task_singles/enquiries_tolerances.html", context=context)

def reload_tolerance(request):
	try:
		filename=os.path.join("\\\\filestorage\cie\Operations\Results Team\Enquiries About Results\\0.Nova Downloads\\Tolerances.xlsx")
		workbook = load_workbook(filename)
		sheet = workbook.active
		row_num = 1
		models.MarkTolerances.objects.all().delete()
		# Iterating through All rows with all columns...
		for i in range(2, sheet.max_row+1):
			row_num += 1
			row = [cell.value for cell in sheet[i]] # sheet[n] gives nth row (list of cells)
			if str(row[0]) == "None":
				raise Exception("Syllabus missing from row")
			elif str(row[1]) == "None":
				raise Exception("Component missing from row")
			elif str(row[2]) == "None":
				raise Exception("Tolerance missing from row")
			else:
				models.MarkTolerances.objects.create(
					eps_ass_code = str(row[0]).zfill(4),
					eps_com_id = str(row[1]).zfill(2),
					mark_tolerance = row[2]
				)
		qs = models.MarkTolerances.objects.all().order_by('eps_ass_code','eps_com_id')
		context = {"current_status":f"Tolerances reloaded successfully.","qs":qs}
		return render(request, "enquiries/task_singles/enquiries_tolerances.html", context=context)
	except Exception as e:
		print(f"{e}")
		context = {"current_status":f"{e}, Row: {row_num}"}
		return render(request, "enquiries/task_singles/enquiries_tolerances.html", context=context)

def examiner_list_view(request):
	search_q = ""
	if request.GET.get('search_query') is not None:
		search_q = request.GET.get('search_query')
	split_search_q_1 = "None"
	split_search_q_2 = "None"
	if search_q.find("/") > -1:
		split_search_q_1 = search_q.split("/")[0]
		split_search_q_2 = search_q.split("/")[1]
	ep_queryset = models.UniqueCreditor.objects.prefetch_related('creditors__exm_per_details','exm_availability','exm_notes','exm_conflicts','exm_email_manual').filter(Q(exm_creditor_no__icontains = search_q) | Q(exm_surname__icontains = search_q) | Q(creditors__exm_per_details__ass_code__icontains = split_search_q_1) & Q(creditors__exm_per_details__com_id__icontains = split_search_q_2)).order_by('exm_creditor_no').distinct()
	ep_queryset = ep_queryset.filter(creditors__currently_valid=1)
	ep_queryset_paged = Paginator(ep_queryset,10,0,True)
	page_number = request.GET.get('page')
	try:
		page_obj = ep_queryset_paged.get_page(page_number)  # returns the desired page object
	except PageNotAnInteger:
		# if page_number is not an integer then assign the first page
		page_obj = ep_queryset_paged.page(1)
	except EmptyPage:
		# if page is empty then return last page
		page_obj = ep_queryset_paged.page(ep_queryset_paged.num_pages)	
	context = {"ep": page_obj, "sq":search_q, }
	return render(request, "enquiries/examiners/enquiries_examiner_list.html", context=context)

def examiner_detail(request, per_sid=None):
	exm_queryset = None
	if per_sid is not None:	
		uc_queryset = models.UniqueCreditor.objects.get(per_sid=per_sid)
		exm_queryset = models.EnquiryPersonnel.objects.filter(per_sid=per_sid)
		exm_queryset2 = models.EnquiryPersonnelDetails.objects.filter(enpe_sid__per_sid=per_sid)
		exm_queryset2 = exm_queryset2.filter(enpe_sid__currently_valid=1)
		#examiner email override check
		try:
			email_new = models.ExaminerEmailOverride.objects.get(creditor__per_sid = per_sid).examiner_email_manual
		except:
			email_new = uc_queryset.exm_email


	context = {"uc": uc_queryset, "exm": exm_queryset, "exm2": exm_queryset2, "exm_email": email_new}
	return render(request, 'enquiries/examiners/enquiries_examiner_detail.html', context=context)

def examiner_scripts_view(request, per_sid=None):
	
	uc = models.UniqueCreditor.objects.get(per_sid=per_sid).exm_creditor_no
	
	enpe_list = []
	for exm in models.EnquiryPersonnelDetails.objects.filter(exm_creditor_no=uc):
		if exm.enpe_sid.enpe_sid not in enpe_list:
			enpe_list.append(exm.enpe_sid.enpe_sid)
	# grab the model rows (ordered by id).
	enpe_queryset = models.EnquiryPersonnel.objects.filter(enpe_sid__in=enpe_list).first()
	ec_queryset = models.ScriptApportionment.objects.filter(enpe_sid__in=enpe_list,apportionment_invalidated=0).order_by('-script_marked','-script_mark_entered','ec_sid')
	context = {"cer": ec_queryset,"enpe": enpe_queryset}
	return render(request, "enquiries/examiners/enquiries_examiner_scripts.html", context=context)

def examiner_availability_view(request, per_sid=None):
	uc_queryset = models.UniqueCreditor.objects.get(per_sid=per_sid)
	context = {"enpe": uc_queryset,}
	return render(request, 'enquiries/examiners/enquiries_examiner_availability.html', context=context)

def examiner_availability_edit_view(request, per_sid=None):
	print(per_sid)
	print(request.POST.get('start'))
	if per_sid is not None and request.method == 'POST':
		exm = models.UniqueCreditor.objects.get(per_sid=per_sid)
		avail = models.ExaminerAvailability(
			#ea_sid = models.EnquiryPersonnel.objects.only('enpe_sid').get(enpe_sid=enpe_sid),
			unavailability_start_date = request.POST.get('start'),
			unavailability_end_date = request.POST.get('end'),
			unavailable_2_flag = request.POST.get('esstwo'),
			unavailable_5_flag = request.POST.get('essfive'),
			unavailable_9_flag = request.POST.get('essnine'),
		)
		avail.save()
		avail.creditor.add(exm)
	return redirect('examiner_detail', per_sid)

def examiner_availability_delete(request, note_id=None):
	if note_id is not None:
		per_sid = models.UniqueCreditor.objects.get(exm_availability__pk = note_id).per_sid
		models.ExaminerAvailability.objects.get(id=note_id).delete()
	return redirect('examiner_detail', per_sid)

def examiner_notes_view(request, per_sid=None):
	uc_queryset = models.UniqueCreditor.objects.get(per_sid=per_sid)
	context = {"enpe": uc_queryset,}
	return render(request, 'enquiries/examiners/enquiries_examiner_notes.html', context=context)

def examiner_notes_edit_view(request, per_sid=None):
		#Get username to filter tasks
	username = None
	if request.user.is_authenticated:
		username = request.user

	if per_sid is not None and request.method == 'POST':
		exm = models.UniqueCreditor.objects.get(per_sid=per_sid)
		note_entry = models.ExaminerNotes(
			examiner_notes = request.POST.get('exm_note'),
			note_owner = username,
		)
		note_entry.save()
		note_entry.creditor.add(exm)
	return redirect('examiner_detail', per_sid)

def examiner_notes_delete(request, note_id=None):
	if note_id is not None:
		per_sid = models.UniqueCreditor.objects.get(exm_notes__pk = note_id).per_sid
		models.ExaminerNotes.objects.get(id=note_id).delete()
	return redirect('examiner_detail', per_sid)

def examiner_conflicts_view(request, per_sid=None):
	uc_queryset = models.UniqueCreditor.objects.get(per_sid=per_sid)
	if models.ExaminerConflicts.objects.filter(creditor=uc_queryset).exists():
		current_conflict = models.ExaminerConflicts.objects.get(creditor=uc_queryset)
	else:
		current_conflict = None
	context = {"enpe": uc_queryset, "current_conflict":current_conflict}
	return render(request, 'enquiries/examiners/enquiries_examiner_conflicts.html', context=context)

def examiner_conflicts_edit_view(request, per_sid=None):
	#Get username
	username = None
	if request.user.is_authenticated:
		username = request.user

	if per_sid is not None and request.method == 'POST':
		exm = models.UniqueCreditor.objects.get(per_sid=per_sid)
		#detect if note exists or not
		try:
			existing_note = models.ExaminerConflicts.objects.get(creditor__per_sid = per_sid).pk
		except:
			existing_note = None
		#if note does not exist, create it, else update it
		conflict_string= str(request.POST.get('exm_conflicts')).upper()
		if existing_note is None:
			conflict_entry = models.ExaminerConflicts(
				examiner_conflicts = conflict_string,
				note_owner = username,
			)
			conflict_entry.save()
			conflict_entry.creditor.add(exm)
		else:
			models.ExaminerConflicts.objects.filter(id=existing_note).update(examiner_conflicts=conflict_string)
	return redirect('examiner_detail', per_sid)

def examiner_conflicts_delete(request, note_id=None):
	if note_id is not None:
		per_sid = models.UniqueCreditor.objects.get(exm_conflicts__pk = note_id).per_sid
		models.ExaminerConflicts.objects.get(id=note_id).delete()
	return redirect('examiner_detail', per_sid)

def examiner_email_view(request, per_sid=None):
	uc_queryset = models.UniqueCreditor.objects.get(per_sid=per_sid)
	context = {"enpe": uc_queryset,}
	return render(request, 'enquiries/examiners/enquiries_examiner_email.html', context=context)

def examiner_email_edit_view(request, per_sid=None):
	#Get username
	username = None
	if request.user.is_authenticated:
		username = request.user
	print(username)

	if per_sid is not None and request.method == 'POST':
		exm = models.UniqueCreditor.objects.get(per_sid=per_sid)
		#detect if note exists or not
		try:
			existing_note = models.ExaminerEmailOverride.objects.get(creditor__per_sid = per_sid).pk
		except:
			existing_note = None
		#if note does not exist, create it, else update it
		print(existing_note)
		if existing_note is None:
			conflict_entry = models.ExaminerEmailOverride(
				examiner_email_manual = request.POST.get('exm_new_email'),
			)
			conflict_entry.save()
			conflict_entry.creditor.add(exm)
		else:
			models.ExaminerEmailOverride.objects.filter(id=existing_note).update(examiner_email_manual=request.POST.get('exm_new_email'))
	return redirect('examiner_detail', per_sid)

def case_system_view(request):
	#Get username
	username = None
	if request.user.is_authenticated:
		username = request.user
	my_cases = models.ExaminerOverdueCases.objects.filter(task_assigned_to=username,task_completion_date__isnull=True).order_by('task_creation_date')
	all_cases = models.ExaminerOverdueCases.objects.all().order_by('task_creation_date')
	context = {'all_cases':all_cases, 'my_cases':my_cases}
	return render(request, 'enquiries/examiners/case_system/case_home.html', context=context)

def create_cases_view(request):
	today = datetime.now().date()
	tomorrow = today + timedelta(1)
	today_start = datetime.combine(today, time())
	today_end = datetime.combine(tomorrow, time())
	all_cases = models.ExaminerOverdueCases.objects.all().order_by('task_creation_date')
	exmsla_tasks = models.TaskManager.objects.filter(task_completion_date__isnull=True, task_id='EXMSLA')
	exmsla_scripts = []
	for task in exmsla_tasks:
		exmsla_scripts.append(task.ec_sid.ec_sid)
	examiner_scripts = models.ScriptApportionment.objects.filter(apportionment_invalidated=0,script_marked=1)
	#list of all scripts valid and not marked yet, check if they are against a case
	overdue_filtered_scripts_list = []
	for script in examiner_scripts:
		if script.ec_sid.ec_sid in exmsla_scripts: #has an overdue task open
			if script.overdue_case is None: # is not assigned to a case
				overdue_filtered_scripts_list.append([script.ec_sid.ec_sid,script.enpe_sid.enpe_sid])
	#cycle scripts and check if avaiable case open today for that examiner:
	for script in overdue_filtered_scripts_list:
		examiner = script[1]
		case_script = script[0]
		models.ExaminerOverdueCases.objects.filter()
		today_case_open = models.ExaminerOverdueCases.objects.filter(case_apportioned_scripts__enpe_sid=examiner,task_creation_date__lte=today_end,task_creation_date__gte=today_start).count() 
		#If case exists, assign this script to it
		if today_case_open > 0:
			today_case_open = models.ExaminerOverdueCases.objects.filter(case_apportioned_scripts__enpe_sid=examiner,task_creation_date__lte=today_end,task_creation_date__gte=today_start).first()
			models.ScriptApportionment.objects.filter(apportionment_invalidated=0,script_marked=1,ec_sid=case_script,enpe_sid=examiner).update(overdue_case=today_case_open)
		else:
			new_case = models.ExaminerOverdueCases.objects.create(enpe_sid=models.EnquiryPersonnel.objects.get(enpe_sid=examiner))
			models.ScriptApportionment.objects.filter(apportionment_invalidated=0,script_marked=1,ec_sid=case_script,enpe_sid=examiner).update(overdue_case=new_case)

	return redirect('case_system')

def case_detail_view(request, case_id=None):
	task_queryset = models.ExaminerOverdueCases.objects.get(pk=case_id)
	#Check for comments on task
	task_comments = None
	if models.CaseComments.objects.filter(case_pk=task_queryset.pk).exists():
		task_comments = models.CaseComments.objects.filter(case_pk=task_queryset.pk).order_by('case_comment_creation_date')
	context = {"case_id":case_id, "case":task_queryset, "task_comments":task_comments}
	return render(request, "enquiries/examiners/case_system/case_details.html", context=context)

def panel_list_view(request):
	search_q = ""
	if request.GET.get('search_query') is not None:
		search_q = request.GET.get('search_query')
	split_search_q_1 = "None"
	split_search_q_2 = "None"
	if search_q.find("/") > -1:
		split_search_q_1 = search_q.split("/")[0]
		split_search_q_2 = search_q.split("/")[1]
		ep_queryset = models.ExaminerPanels.objects.filter(Q(ass_code__icontains = split_search_q_1) & Q(com_id__icontains = split_search_q_2)).order_by('ass_code','com_id')
	else:
		ep_queryset = models.ExaminerPanels.objects.all().annotate(exm_count=Count("panel_creditors",distinct=True)).order_by('ass_code','com_id')
	ep_queryset_paged = Paginator(ep_queryset,10,0,True)
	page_number = request.GET.get('page')
	try:
		page_obj = ep_queryset_paged.get_page(page_number)  # returns the desired page object
	except PageNotAnInteger:
		# if page_number is not an integer then assign the first page
		page_obj = ep_queryset_paged.page(1)
	except EmptyPage:
		# if page is empty then return last page
		page_obj = ep_queryset_paged.page(ep_queryset_paged.num_pages)	
	context = {"ep": page_obj, "sq":search_q, }
	return render(request, "enquiries/examiners/enquiries_panel_list.html", context=context)

def panel_set_manual_view(request):
	panel_id = request.POST.get('panel_id')
	models.ExaminerPanels.objects.filter(pk=panel_id).update(manual_apportionment=not models.ExaminerPanels.objects.get(pk=panel_id).manual_apportionment)
	return redirect('panel_list')

def panel_update_note_view(request):
	panel_id = request.POST.get('panel_id')
	panel_notes = request.POST.get('panel_notes')
	models.ExaminerPanels.objects.filter(pk=panel_id).update(panel_notes=panel_notes)
	return redirect('panel_list')

def user_panel_view(request):
	# grab the model rows (ordered by id), filter to required task and where not completed.
	queryset = models.User.objects.filter(is_active=True).exclude(user_primary__primary_team__team_name='Server').order_by('username','user_primary__primary_team__team_name')
	queryset2 = models.User.objects.filter(is_active=True,assigned_tasks__task_completion_date__isnull=True).exclude(user_primary__primary_team__team_name='Server').annotate(task_count=Count("assigned_tasks",distinct=True)).order_by('username','user_primary__primary_team__team_name')
	teams = models.TaskTeams.objects.all().order_by('id')
	context = {"users": queryset, "users2": queryset2, "teams":teams}
	return render(request, "enquiries/main_templates/enquiries_task_user.html", context=context)

def create_user_view(request):
	username = request.POST.get('username')
	password = request.POST.get('password')
	email = request.POST.get('email')
	pteam = request.POST.get('pteam')
	access = request.POST.get('access')

	new_user = User.objects.create_user(username=username,
							email=email,
							password=password)
	
	models.TaskUserPrimary.objects.create(
		task_user = new_user,
		primary_team = models.TaskTeams.objects.get(team_name=pteam),
		primary_status = access
	)

	return redirect("user_panel")

def update_user_view(request):
	username = request.POST.get('username')
	pteam = request.POST.get('pteam')
	access = request.POST.get('access')

	this_user = User.objects.get(username=username)
	
	models.TaskUserPrimary.objects.filter(task_user = this_user).update(
		primary_team = models.TaskTeams.objects.get(team_name=pteam),
		primary_status = access
	)

	if models.TaskUserSecondary.objects.filter(task_user=this_user,secondary_team=models.TaskTeams.objects.get(team_name=pteam)).exists():
		models.TaskUserSecondary.objects.filter(task_user=this_user,
		secondary_team=models.TaskTeams.objects.get(team_name=pteam)).delete()
	

	return redirect("edit_user", this_user.pk)

def user_remove_tasks_view(request):
	username = request.POST.get('username')
	# grab the model rows (ordered by id), filter to required task and where not completed.
	models.TaskManager.objects.filter(task_assigned_to=models.User.objects.get(username=username)).update(task_assigned_to=None, task_assigned_date=None)
	

	return redirect("user_panel")

def edit_user_view(request, userid=None):
	teams = models.TaskTeams.objects.all().order_by('team_name')
	user = User.objects.get(id=userid)

	team_list = []
	for team in models.TaskUserSecondary.objects.filter(task_user=user):
		team_list.append(team.secondary_team.team_name)

	primary_team = models.TaskUserPrimary.objects.get(task_user=user).primary_team.team_name

	context = {"teams":teams, "user":user, "team_list":team_list, "primary_team":primary_team}
	return render(request, "enquiries/main_templates/enquiries_task_user_detail.html", context=context) 

def user_change_secondary(request):
	userid = request.POST.get('userid')
	sec_team = request.POST.get('sec_team')
	type = request.POST.get('type')

	if type == 'enroll':
		models.TaskUserSecondary.objects.create(
			task_user = User.objects.get(id=userid),
			secondary_team = models.TaskTeams.objects.get(team_name=sec_team)
		)
	else:
		models.TaskUserSecondary.objects.filter(
			task_user=User.objects.get(id=userid),
			secondary_team = models.TaskTeams.objects.get(team_name=sec_team)
		).delete()


	return redirect('edit_user', userid)

# def new_mis_request(request):
# 	sessions = str(EarServerSettings.objects.get(pk=1).session_id_list).split(',')
# 	#rev_exm = EnquiryPersonnelDetails.objects.filter(enpe_sid=ScriptApportionment.objects.filter(ec_sid=script_id, apportionment_invalidated=0).first().enpe_sid).first().exm_examiner_no
# 	rev_exm = EnquiryPersonnelDetails.objects.filter(enpe_sid=ScriptApportionment.objects.filter(ec_sid=script_id, apportionment_invalidated=0).first().enpe_sid,session__in=sessions).first().exm_examiner_no
# 	#This is all to get the scaled mark
# 	scale_ass_code = app_task.ec_sid.eps_ass_code
# 	scale_comp_id = app_task.ec_sid.eps_com_id
# 	scale_centre_no = app_task.enquiry_id.centre_id
# 	scale_cand_no = app_task.ec_sid.erp_sid.eps_cand_id
# 	scale_ses_id = app_task.ec_sid.eps_ses_sid
# 	print(scale_ass_code + " " + scale_comp_id + " " + scale_centre_no + " " + scale_cand_no + " " + scale_ses_id)
# 	if ScaledMarks.objects.filter(eps_ass_code=scale_ass_code,eps_com_id=scale_comp_id,eps_cnu_id=scale_centre_no,eps_cand_no=scale_cand_no,eps_ses_sid=scale_ses_id).exists():
# 		original_mark = ScaledMarks.objects.filter(eps_ass_code=scale_ass_code,eps_com_id=scale_comp_id,eps_cnu_id=scale_centre_no,eps_cand_no=scale_cand_no,eps_ses_sid=scale_ses_id).first().scaled_mark
# 		if original_mark is None:
# 			print("No Valid Scaled Mark 1")
# 		else:
# 			original_mark = int(original_mark.split('.')[0])
# 	else:
# 		print("No Valid Scaled Mark 2")
# 	cred_no = ScriptApportionment.objects.filter(ec_sid=script_id, apportionment_invalidated=0).first().enpe_sid.per_sid.exm_creditor_no

# 	#Work to be done by NEWMIS done here 
# 	# new_filename = os.path.realpath("\\\\filestorage\cie\Operations\Results Team\Enquiries About Results\\0.RPA_MIS Returns\EARTemplate1.xlsx")
# 	# print(os.path.dirname(new_filename))

# 	new_filename = os.path.realpath("\\\\filestorage\cie\Operations\Results Team\Enquiries About Results\\0.RPA_MIS Returns\EARTemplate1.xlsx")
# 	print(new_filename)

# 	workbook = load_workbook(filename=new_filename)
# 	sheet = workbook.active

# 	#Syll/Comp
# 	sheet["A2"] = syll_comp
# 	#Batch
# 	sheet["I2"] = batch_no
# 	#Centre
# 	sheet["A4"] = centre_no
# 	#Cand no
# 	sheet["B4"] = cand_no
# 	#Cand name
# 	sheet["C4"] = cand_name
# 	#Orig Exm
# 	sheet["D4"] = original_exm
# 	#Rev Exm
# 	sheet["E4"] = rev_exm
# 	#Scaled (prev) mark
# 	sheet["F4"] = original_mark

# 	new_filename2 = os.path.realpath("\\\\filestorage\cie\Operations\Results Team\Enquiries About Results\\" + mis_folder + "\Outbound\\Examiner-" + cred_no + "_" + batch_no + "_" + centre_no + "_" + cand_no + "_" + syll_comp.split('/')[0] + "_" + syll_comp.split('/')[1] + "_MIS.xlsx")
# 	print(new_filename2)
# 	workbook.save(filename=new_filename2)


