from openpyxl import load_workbook
import sys
import os
import django
import datetime

if os.getenv('DJANGO_DEVELOPMENT') == 'true':
    path = os.path.join('C:\\Users\\bulmaj\\OneDrive - Cambridge\\Desktop\\Dev\\Nova')
    sys.path.append(path)
    os.environ['DJANGO_SETTINGS_MODULE'] = 'redepplan.settings_dev'
else:
    sys.path.append('C:/Dev/redepplan')
    os.environ['DJANGO_SETTINGS_MODULE'] = 'redepplan.settings'

django.setup()

from enquiries.models import TaskTeams, TaskTypes, TaskUserPrimary, TaskUserSecondary
from django.contrib.auth.models import User



def load_core_tables():

    start_time = datetime.datetime.now()
    print("Start Time:" + str(datetime.datetime.now()))

    filename=os.path.join("Y:\Operations\Results Team\Enquiries About Results\\0.Nova Downloads\\Test Users\\TaskTeams.xlsx")
    workbook = load_workbook(filename)
    sheet = workbook.active

    for row in sheet.iter_rows():
        taskteam = str(row[0].value)
        print(taskteam)

        if not TaskTeams.objects.filter(team_name=taskteam).exists:
            TaskTeams.objects.create(
                team_name=taskteam,
            )

    filename=os.path.join("Y:\Operations\Results Team\Enquiries About Results\\0.Nova Downloads\\Test Users\\TaskTypes.xlsx")
    workbook = load_workbook(filename)
    sheet = workbook.active

    for row in sheet.iter_rows():
        taskid = str(row[0].value)
        taskdesc = str(row[1].value)
        taskteam = str(row[2].value)
        taskrank = str(row[3].value)

        try:
            if not TaskTypes.objects.filter(task_id=taskid).exists:
                TaskTypes.objects.create(
                    task_id = taskid,
                    task_description = taskdesc,
                    task_team = TaskTeams.objects.get(team_name=taskteam),
                    task_rank = taskrank
                )
        except:
            pass


    filename=os.path.join("Y:\Operations\Results Team\Enquiries About Results\\0.Nova Downloads\\Test Users\\TestUsers.xlsx")
    workbook = load_workbook(filename)
    sheet = workbook.active

    for row in sheet.iter_rows():
        username = str(row[0].value)
        password = str(row[1].value)

        if not User.objects.filter(username=username).exists:
            User.objects.create_user(username=username,
            email='',
            password=password)

    filename=os.path.join("Y:\Operations\Results Team\Enquiries About Results\\0.Nova Downloads\\Test Users\\PrimaryTeams.xlsx")
    workbook = load_workbook(filename)
    sheet = workbook.active

    for row in sheet.iter_rows():
        username = str(row[0].value)
        teamname = str(row[1].value)
        status = str(row[2].value)

        # try:
        #if not TaskUserPrimary.objects.filter(task_user=User.objects.get(username=username)).exists:
        TaskUserPrimary.objects.create(
            task_user = User.objects.get(username=username),
            primary_team = TaskTeams.objects.get(team_name=teamname),
            primary_status = status
        )
        # except:
        #     pass

    end_time = datetime.datetime.now()
    print(end_time - start_time)

load_core_tables()