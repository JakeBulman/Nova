import sys
import os
import django
from django.utils import timezone
import datetime
from openpyxl import load_workbook

if os.getenv('DJANGO_DEVELOPMENT') == 'true':
    print('DEV')
    path = os.path.join('C:\\Users\\bulmaj\\OneDrive - Cambridge\\Desktop\\Dev\\Nova')
    sys.path.append(path)
    os.environ['DJANGO_SETTINGS_MODULE'] = 'redepplan.settings_dev'
elif os.getenv('DJANGO_PRODUCTION') == 'true':
    print('PROD')
    path = os.path.join('C:\\Dev\\Nova')
    sys.path.append(path)
    os.environ['DJANGO_SETTINGS_MODULE'] = 'redepplan.settings_prod'
else:
    print('UAT - Check')
    path = os.path.join('C:\\Dev\\nova')
    sys.path.append(path)
    os.environ['DJANGO_SETTINGS_MODULE'] = 'redepplan.settings'

django.setup()

from enquiries.models import MarkTolerances, ManualTaskQueue
from django.contrib.auth.models import User

def run_algo():

    for task in ManualTaskQueue.objects.all().filter(task_queued=1, task_running=0):
        task.task_running = 1
        task.save()
        if task.task_type == 'MARTOL':
            filename=os.path.join("\\\\filestorage\cie\Operations\Results Team\Enquiries About Results\\0.Nova Downloads\\Tolerances.xlsx")
            workbook = load_workbook(filename)
            sheet = workbook.active

            # Iterating through All rows with all columns...
            for i in range(1, sheet.max_row+1):
                row = [cell.value for cell in sheet[i]] # sheet[n] gives nth row (list of cells)
                if MarkTolerances.objects.filter(eps_ass_code = str(row[0]).zfill(4),eps_com_id = str(row[1]).zfill(2)).exists():
                    MarkTolerances.objects.filter(eps_ass_code = str(row[0]).zfill(4),eps_com_id = str(row[1]).zfill(2)).update(mark_tolerance = row[2])
                else: 
                    MarkTolerances.objects.create(
                        eps_ass_code = str(row[0]).zfill(4),
                        eps_com_id = str(row[1]).zfill(2),
                        mark_tolerance = row[2]
                    )
        task.task_completion_date = timezone.now()
        task.task_queued = 0
        task.task_running = 0
        task.save()

        # More tasks can be checked for using IF statement here...

run_algo()