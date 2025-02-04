import sys
import os
import django
import datetime
import pyodbc
import pandas as pd
from openpyxl import load_workbook
import re
from django.conf import settings
from django.utils.timezone import make_aware
from dateutil.parser import parse

if os.getenv('DJANGO_DEVELOPMENT') == 'true':
    print('UAT')
    path = os.path.join('C:\\Dev\\Nova')
    sys.path.append(path)
    os.environ['DJANGO_SETTINGS_MODULE'] = 'redepplan.settings'
elif os.getenv('DJANGO_PRODUCTION') == 'true':
    print('PRD')
    path = os.path.join('C:\\Dev\\Nova')
    sys.path.append(path)
    os.environ['DJANGO_SETTINGS_MODULE'] = 'redepplan.settings_prod'
else:
    print('DEV')
    path = os.path.join('C:\\Users\\bulmaj\\OneDrive - Cambridge\\Desktop\\Dev\\Nova')
    sys.path.append(path)
    os.environ['DJANGO_SETTINGS_MODULE'] = 'redepplan.settings_dev'

django.setup()

from enquiries.models import CentreEnquiryRequests, EarServerSettings, EnquiryGrades, MarkTolerances, EnquiryComponentsHistory, EnquiryComponents, TaskManager, TaskTypes

def load_core_tables():

    start_time = datetime.datetime.now()
    print("Start Time:" + str(datetime.datetime.now()))

    #Limits enquiries to session or enquiry list
    session_id = EarServerSettings.objects.first().session_id_list
    enquiry_id_list = EarServerSettings.objects.first().enquiry_id_list
    if enquiry_id_list != '':
        enquiry_id_list = ' and sid in (' + enquiry_id_list + ')'

    print("ENQ:" + enquiry_id_list)


    # # # Get datalake data - Enquiry Request Parts
    # with pyodbc.connect("DSN=hive.ucles.internal", autocommit=True) as conn:
    #     df = pd.read_sql(f'''
    #         select 
    #         cer.sid as enquiry_id,
    #         erp.caom_ass_code as eps_ass_code,
    #         erp.caom_cand_no as eps_cand_no,
    #         qgrp.title as previous_grade,
    #         qgrp.seq_no as previous_seq,
    #         qgrn.title as new_grade,
    #         qgrn.seq_no as new_seq
                         
    #         from ar_meps_req_prd.centre_enquiry_requests cer
    #         left join ar_meps_req_prd.enquiry_request_parts erp
    #         on cer.sid = erp.cer_sid
    #         left join ar_meps_ord_prd.audit_changes ac
    #         on cer.ses_sid = ac.ses_sid
    #         and cer.cnu_id = ac.co_centre_id
    #         and erp.caom_ass_code = ac.ass_code
    #         and erp.caom_cand_no = ac.coc_cand_no
    #         left join ar_meps_prod_prd.qualification_grades qgr
    #         on qgr.qgs_sid = erp.qgr_qgs_sid and qgr.seq_no = erp.qgr_seq_no_derived
    #         left join ar_meps_prod_prd.qualification_grades qgrp
    #         on erp.qgr_qgs_sid = qgrp.qgs_sid and qgrp.seq_no = ac.previous_value
    #         left join ar_meps_prod_prd.qualification_grades qgrn
    #         on erp.qgr_qgs_sid = qgrn.qgs_sid and qgrn.seq_no = ac.new_value
    #         where ac.record_type = "CAOM"
    #         --and ac.confirmation_status = "UNC"
    #         and ac.field_name in ("QGR_SEQ_NO_DERIVED","QGR_SEQ_NO_MANUAL")
    #         and ac.ses_sid in ({session_id}) 
    #                             ''', conn)

    # def insert_to_model_erp(row):
    #     if EnquiryGrades.objects.filter(enquiry_id=row['enquiry_id']).exists():
    #         EnquiryGrades.objects.filter(enquiry_id=row['enquiry_id']).update(
    #             enquiry_id = CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=row['enquiry_id']),
    #             eps_ass_code = row['eps_ass_code'],
    #             eps_cand_no = row['eps_cand_no'],
    #             previous_grade = row['previous_grade'],
    #             previous_seq = row['previous_seq'],
    #             new_grade = row['new_grade'],
    #             new_seq = row['new_seq']
    #         )
    #     else:
    #         try:
    #             EnquiryGrades.objects.create(
    #                 enquiry_id = CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=row['enquiry_id']),
    #                 eps_ass_code = row['eps_ass_code'],
    #                 eps_cand_no = row['eps_cand_no'],
    #                 previous_grade = row['previous_grade'],
    #                 previous_seq = row['previous_seq'],
    #                 new_grade = row['new_grade'],
    #                 new_seq = row['new_seq']
    #             )
    #         except:
    #             pass

    # EnquiryGrades.objects.all().delete()

    # df.apply(insert_to_model_erp, axis=1)


    # print("EG loaded:" + str(datetime.datetime.now()))

    # filename=os.path.join("\\\\filestorage\cie\Operations\Results Team\Enquiries About Results\\0.Nova Downloads\\Tolerances.xlsx")
    # workbook = load_workbook(filename)
    # sheet = workbook.active

    # # Iterating through All rows with all columns...
    # for i in range(1, sheet.max_row+1):
    #     row = [cell.value for cell in sheet[i]] # sheet[n] gives nth row (list of cells)
    #     if str(row[0]) != 'None':
    #         if MarkTolerances.objects.filter(eps_ass_code = str(row[0]).zfill(4),eps_com_id = str(row[1]).zfill(2)).exists():
    #             MarkTolerances.objects.filter(eps_ass_code = str(row[0]).zfill(4),eps_com_id = str(row[1]).zfill(2)).update(mark_tolerance = row[2])
    #         else: 
    #             MarkTolerances.objects.create(
    #                 eps_ass_code = str(row[0]).zfill(4),
    #                 eps_com_id = str(row[1]).zfill(2),
    #                 mark_tolerance = row[2]
    #             )



    # # Get datalake data - ECH
    # with pyodbc.connect("DSN=hive.ucles.internal", autocommit=True) as conn:
    #     df = pd.read_sql(f'''
    #                     SELECT distinct
    #                     cer.sid as cer_sid,
    #                     ec.sid as ec_sid,
    #                     wrm.ses_sid as ses_sid,
    #                     wrm.ass_code as ass_code,
    #                     wrm.com_id as com_id,
    #                     wrm.cnu_id as cnu_id,
    #                     wrm.cand_no as cand_no,
    #                     spp.examiner_no as exm_position,
    #                     wrm.kbr_code as kbr_code,
    #                     kbr.reason as kbr_reason,
    #                     wrm.mark as current_mark,
    #                     ec.ccm_measure as ear_mark,
    #                     ec.ccm_outcome as ear_mark_alt
    #                     from ar_meps_req_prd.enquiry_components ec
    #                     left join ar_meps_req_prd.enquiry_request_parts erp
    #                     on erp.sid = ec.erp_sid
    #                     left join ar_meps_req_prd.centre_enquiry_requests cer
    #                     on cer.sid = erp.cer_sid
    #                     left join ar_meps_mark_prd.working_raw_marks wrm
    #                     on wrm.ses_sid = ec.ccm_ses_sid
    #                     and wrm.ass_code = ec.ccm_ass_code
    #                     and wrm.com_id = ec.ccm_com_id
    #                     and wrm.cand_no = ec.ccm_cand_no
    #                     and wrm.cnu_id = ec.ccm_cnu_id
    #                     left join ar_meps_pan_prd.session_panel_positions spp
    #                     on wrm.spp_sid = spp.sid
    #                     left join ar_meps_mark_prd.keyed_batch_reason kbr
    #                     on kbr.code = wrm.kbr_code
    #                     where ec.ccm_ses_sid in ({session_id}) 
    #                     and kbr_code = 'PSM'
    #                             ''', conn)
        
    # print(df)

    # def insert_to_model_enpee(row):
    #     if EnquiryComponentsHistory.objects.filter(ec_sid=row['ec_sid']).exists():
    #         if row['kbr_code'] == 'PSM':
    #             print(row['cer_sid'])
    #         EnquiryComponentsHistory.objects.filter(ec_sid=row['ec_sid']).update(
    #         cer_sid = CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=row['cer_sid']),
    #         ec_sid = EnquiryComponents.objects.only('ec_sid').get(ec_sid=row['ec_sid']),
    #         eps_ses_sid = row['ses_sid'],
    #         eps_ass_code = row['ass_code'],
    #         eps_com_id = row['com_id'],
    #         eps_cnu_id = row['cnu_id'],
    #         eps_cand_no = row['cand_no'],
    #         exm_position = row['exm_position'],
    #         kbr_code = row['kbr_code'],
    #         kbr_reason = row['kbr_reason'],
    #         current_mark = row['current_mark'],
    #         ear_mark = row['ear_mark'],
    #         ear_mark_alt = row['ear_mark_alt'],
    #         )

    #     else:
    #         try:
    #             EnquiryComponentsHistory.objects.create(
    #             cer_sid = CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=row['cer_sid']),
    #             ec_sid = EnquiryComponents.objects.only('ec_sid').get(ec_sid=row['ec_sid']),
    #             eps_ses_sid = row['ses_sid'],
    #             eps_ass_code = row['ass_code'],
    #             eps_com_id = row['com_id'],
    #             eps_cnu_id = row['cnu_id'],
    #             eps_cand_no = row['cand_no'],
    #             exm_position = row['exm_position'],
    #             kbr_code = row['kbr_code'],
    #             kbr_reason = row['kbr_reason'],
    #             current_mark = row['current_mark'],
    #             ear_mark = row['ear_mark'],
    #             ear_mark_alt = row['ear_mark_alt'],
    #             )
    #         except:
    #             pass
        
    # df.apply(insert_to_model_enpee, axis=1)

    # print("ECH loaded:" + str(datetime.datetime.now()))


    # Get datalake data - ECH
    with pyodbc.connect("DSN=hive.ucles.internal", autocommit=True) as conn:
        df = pd.read_sql(f'''
                        SELECT distinct
                        cer.sid as cer_sid,
                        ec.sid as ec_sid,
                        wrm.ses_sid as ses_sid,
                        wrm.ass_code as ass_code,
                        wrm.com_id as com_id,
                        wrm.cnu_id as cnu_id,
                        wrm.cand_no as cand_no,
                        spp.examiner_no as exm_position,
                        wrm.kbr_code as kbr_code,
                        kbr.reason as kbr_reason,
                        wrm.mark as current_mark,
                        ec.ccm_measure as ear_mark,
                        ec.ccm_outcome as ear_mark_alt,
                        mld.batch as omr_batch,
                        mld.position as omr_position
                        from ar_meps_req_prd.enquiry_components ec
                        left join ar_meps_req_prd.enquiry_request_parts erp
                        on erp.sid = ec.erp_sid
                        left join ar_meps_req_prd.centre_enquiry_requests cer
                        on cer.sid = erp.cer_sid
                        left join ar_meps_mark_prd.working_raw_marks wrm
                        on wrm.ses_sid = ec.ccm_ses_sid
                        and wrm.ass_code = ec.ccm_ass_code
                        and wrm.com_id = ec.ccm_com_id
                        and wrm.cand_no = ec.ccm_cand_no
                        and wrm.cnu_id = ec.ccm_cnu_id
                        left join ar_meps_pan_prd.session_panel_positions spp
                        on wrm.spp_sid = spp.sid
                        left join ar_meps_mark_prd.keyed_batch_reason kbr
                        on kbr.code = wrm.kbr_code
                        left join ar_meps_mark_prd.mark_load_details mld
                        on wrm.mld_sid = mld.sid
                        where ec.ccm_ses_sid in ({session_id}) 
                                ''', conn)

    def insert_to_model_enpee(row):
        try:
            kbr_reason = row['kbr_reason'][:50]
        except:
            kbr_reason = ''
        if EnquiryComponentsHistory.objects.filter(ec_sid=row['ec_sid']).exists():
            EnquiryComponentsHistory.objects.filter(ec_sid=row['ec_sid']).update(
            cer_sid = CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=row['cer_sid']),
            ec_sid = EnquiryComponents.objects.only('ec_sid').get(ec_sid=row['ec_sid']),
            eps_ses_sid = row['ses_sid'],
            eps_ass_code = row['ass_code'],
            eps_com_id = row['com_id'],
            eps_cnu_id = row['cnu_id'],
            eps_cand_no = row['cand_no'],
            exm_position = row['exm_position'],
            kbr_code = row['kbr_code'],
            kbr_reason = kbr_reason,
            current_mark = row['current_mark'],
            ear_mark = row['ear_mark'],
            ear_mark_alt = row['ear_mark_alt'],
            omr_batch = row['omr_batch'],
            omr_position = row['omr_position'],
            )

        else:
            try:
                EnquiryComponentsHistory.objects.create(
                cer_sid = CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=row['cer_sid']),
                ec_sid = EnquiryComponents.objects.only('ec_sid').get(ec_sid=row['ec_sid']),
                eps_ses_sid = row['ses_sid'],
                eps_ass_code = row['ass_code'],
                eps_com_id = row['com_id'],
                eps_cnu_id = row['cnu_id'],
                eps_cand_no = row['cand_no'],
                exm_position = row['exm_position'],
                kbr_code = row['kbr_code'],
                kbr_reason = kbr_reason,
                current_mark = row['current_mark'],
                ear_mark = row['ear_mark'],
                ear_mark_alt = row['ear_mark_alt'],
                )
            except:
                pass
                
    df.apply(insert_to_model_enpee, axis=1)

    print("ECH loaded:" + str(datetime.datetime.now()))

    end_time = datetime.datetime.now()
    print(end_time - start_time)

load_core_tables()
