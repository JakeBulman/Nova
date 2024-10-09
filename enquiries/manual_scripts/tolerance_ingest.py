import sys
import os
import django
from openpyxl import load_workbook

if os.getenv('DJANGO_DEVELOPMENT') == 'true':
    print('DEV')
    path = os.path.join('C:\\Users\\bulmaj\\OneDrive - Cambridge\\Desktop\\Dev\\Nova')
    sys.path.append(path)
    os.environ['DJANGO_SETTINGS_MODULE'] = 'redepplan.settings_dev'
elif os.getenv('DJANGO_PRODUCTION') == 'true':
    print('PROD')
    path = os.path.join('C:\\Dev\\Nova')
    sys.path.append(path)
    os.environ['DJANGO_SETTINGS_MODULE'] = 'redepplan.settings_prod'
else:
    print('UAT')
    sys.path.append('C:/Dev/redepplan')
    os.environ['DJANGO_SETTINGS_MODULE'] = 'redepplan.settings'

django.setup()


from enquiries.models import MarkTolerances


filename=os.path.join("\\\\filestorage\cie\Operations\Results Team\Enquiries About Results\\0.Nova Downloads\\Tolerances.xlsx")
workbook = load_workbook(filename)
sheet = workbook.active

# Iterating through All rows with all columns...
for i in range(1, sheet.max_row+1):
    row = [cell.value for cell in sheet[i]] # sheet[n] gives nth row (list of cells)
    print(str(row[0]) + str(row[1]))
    if MarkTolerances.objects.filter(eps_ass_code = str(row[0]).zfill(4),eps_com_id = str(row[1]).zfill(2)).exists():
        MarkTolerances.objects.filter(eps_ass_code = str(row[0]).zfill(4),eps_com_id = str(row[1]).zfill(2)).update(mark_tolerance = row[2])
    else: 
        MarkTolerances.objects.create(
            eps_ass_code = str(row[0]).zfill(4),
            eps_com_id = str(row[1]).zfill(2),
            mark_tolerance = row[2]
        )