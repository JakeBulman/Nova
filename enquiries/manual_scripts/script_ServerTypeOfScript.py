from openpyxl import load_workbook
import sys
import os
import django
import datetime

if os.getenv('DJANGO_DEVELOPMENT') == 'true':
    print('UAT')
    path = os.path.join('C:\\Dev\\Nova')
    sys.path.append(path)
    os.environ['DJANGO_SETTINGS_MODULE'] = 'redepplan.settings'
elif os.getenv('DJANGO_PRODUCTION') == 'true':
    print('PRD')
    path = os.path.join('C:\\Dev\\Nova')
    sys.path.append(path)
    os.environ['DJANGO_SETTINGS_MODULE'] = 'redepplan.settings_prod'
else:
    print('DEV')
    path = os.path.join('C:\\Users\\bulmaj\\OneDrive - Cambridge\\Desktop\\Dev\\Nova')
    sys.path.append(path)
    os.environ['DJANGO_SETTINGS_MODULE'] = 'redepplan.settings_dev'

django.setup()

from enquiries.models import EnquiryComponents, MarkTolerances



def load_core_tables():

    start_time = datetime.datetime.now()
    print("Start Time:" + str(datetime.datetime.now()))
    filename=os.path.join("Y:\Operations\Results Team\Enquiries About Results\\0.Nova Downloads\\Type Of Script.xlsx")
    workbook = load_workbook(filename)
    sheet = workbook.active

    for row in sheet.iter_rows():
        ass_code = str(row[0].value).zfill(4)
        comp_id = str(row[2].value).zfill(2)
        script_type = row[5].value

        print(ass_code + comp_id + script_type)
        a = EnquiryComponents.objects.filter(eps_ass_code=ass_code,eps_com_id=comp_id).update(script_type=script_type)
        print(a)
    #EnquiryComponents.objects.filter(script_type=None).update(script_type='Unknown')

    filename=os.path.join("Y:\Operations\Results Team\Enquiries About Results\\0.Nova Downloads\\Tolerances.xlsx")
    workbook = load_workbook(filename)
    sheet = workbook.active

    MarkTolerances.objects.all().delete()

    for row in sheet.iter_rows():
        ass_code = str(row[0].value).zfill(4)
        comp_id = row[1].value
        mark_tolerance = row[2].value

        MarkTolerances.objects.create(
            eps_ass_code = ass_code,
            eps_com_id = comp_id,
            mark_tolerance = mark_tolerance
        )

    end_time = datetime.datetime.now()
    print(end_time - start_time)

load_core_tables()