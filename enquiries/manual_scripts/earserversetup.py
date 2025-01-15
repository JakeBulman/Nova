import sys
import os
import django
from openpyxl import load_workbook

if os.getenv('DJANGO_DEVELOPMENT') == 'true':
    print('DEV')
    path = os.path.join('C:\\Users\\bulmaj\\OneDrive - Cambridge\\Desktop\\Dev\\Nova')
    sys.path.append(path)
    os.environ['DJANGO_SETTINGS_MODULE'] = 'redepplan.settings_dev'
elif os.getenv('DJANGO_PRODUCTION') == 'true':
    print('PROD')
    path = os.path.join('C:\\Dev\\Nova')
    sys.path.append(path)
    os.environ['DJANGO_SETTINGS_MODULE'] = 'redepplan.settings_prod'
else:
    print('UAT')
    sys.path.append('C:/Dev/nova') #Jake Local
    #sys.path.append('C:/Users/rfrancisco/Desktop/Dev Folder/EAR/Nova/')
    os.environ['DJANGO_SETTINGS_MODULE'] = 'redepplan.settings'

    
django.setup()

from django.contrib.auth.models import User
from enquiries.models import EarServerSettings, TaskUserPrimary, TaskTeams, TaskTypes

if EarServerSettings.objects.all().count() == 0:
    EarServerSettings.objects.create(
            session_id_list = "19848",
            enquiry_id_list = "",
        )
if TaskTeams.objects.all().count() == 0:
    TaskTeams.objects.create(
        team_name='Server',
    )

if TaskUserPrimary.objects.filter(task_user=1).exists():
    pass
else:
    TaskUserPrimary.objects.create(
        primary_status = "CO",
        primary_team_id = TaskTeams.objects.get(team_name='Server').pk,
        task_user_id = 1
    )

filename=os.path.join("\\\\filestorage\cie\Operations\Results Team\Enquiries About Results\\0.Nova Downloads\\ServerData\\TaskTeams.xlsx")
workbook = load_workbook(filename)
sheet = workbook.active

for row in sheet.iter_rows():
    team_name = row[1].value
    if not TaskTeams.objects.filter(team_name=team_name).exists():
        TaskTeams.objects.create(
            team_name=team_name
        )

filename=os.path.join("\\\\filestorage\cie\Operations\Results Team\Enquiries About Results\\0.Nova Downloads\\ServerData\\TaskTypes.xlsx")
workbook = load_workbook(filename)
sheet = workbook.active

for row in sheet.iter_rows():
    task_id = row[1].value
    task_team_id = row[2].value
    task_description = row[3].value
    task_rank = row[4].value
    if not TaskTypes.objects.filter(task_id=task_id).exists():
        TaskTypes.objects.create(
            task_id = task_id,
            task_team_id = task_team_id,
            task_description = task_description,
            task_rank = task_rank
        )
    else:
        TaskTypes.objects.filter(task_id=task_id).update(
            task_team_id = task_team_id,
            task_description = task_description,
            task_rank = task_rank
        )