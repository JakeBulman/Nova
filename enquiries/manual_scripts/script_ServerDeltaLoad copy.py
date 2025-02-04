import sys
import os
import django
import datetime
import pyodbc
import pandas as pd
from openpyxl import load_workbook

if os.getenv('DJANGO_DEVELOPMENT') == 'true':
    print('UAT')
    path = os.path.join('C:\\Dev\\Nova')
    sys.path.append(path)
    os.environ['DJANGO_SETTINGS_MODULE'] = 'redepplan.settings'
elif os.getenv('DJANGO_PRODUCTION') == 'true':
    print('PRD')
    path = os.path.join('C:\\Dev\\Nova')
    sys.path.append(path)
    os.environ['DJANGO_SETTINGS_MODULE'] = 'redepplan.settings_prod'
else:
    print('DEV')
    path = os.path.join('C:\\Users\\bulmaj\\OneDrive - Cambridge\\Desktop\\Dev\\Nova')
    sys.path.append(path)
    os.environ['DJANGO_SETTINGS_MODULE'] = 'redepplan.settings_dev'

django.setup()

from enquiries.models import CentreEnquiryRequests, EnquiryRequestParts, EnquiryComponents, EnquiryPersonnel, EnquiryPersonnelDetails, EnquiryBatches, EnquiryComponentElements, TaskManager, UniqueCreditor, EnquiryComponentsHistory, EnquiryComponentsExaminerChecks, TaskTypes, EarServerSettings, EnquiryGrades, EnquiryDeadline, ExaminerPanels, MarkTolerances, ScaledMarks, CentreEnquiryRequestsExtra, EnquiryRequestPartsExtra, EnquiryComponentsExtra

def load_core_tables():

    start_time = datetime.datetime.now()
    print("Start Time:" + str(datetime.datetime.now()))

    #Limits enquiries to session or enquiry list
    session_id = EarServerSettings.objects.first().session_id_list
    enquiry_id_list = EarServerSettings.objects.first().enquiry_id_list
    if enquiry_id_list != '':
        enquiry_id_list = ' and sid in (' + enquiry_id_list + ')'

    print("ENQ:" + enquiry_id_list)

    filename=os.path.join("Y:\Operations\Results Team\Enquiries About Results\\0.Nova Downloads\\Tolerances.xlsx")
    workbook = load_workbook(filename)
    sheet = workbook.active

    # Iterating through All rows with all columns...
    for i in range(1, sheet.max_row+1):
        row = [cell.value for cell in sheet[i]] # sheet[n] gives nth row (list of cells)
        if MarkTolerances.objects.filter(eps_ass_code = row[0],eps_com_id = row[1]).exists():
            MarkTolerances.objects.filter(eps_ass_code = row[0],eps_com_id = row[1]).update(mark_tolerance = row[2])
        else: 
            MarkTolerances.objects.create(
                eps_ass_code = row[0],
                eps_com_id = row[1],
                mark_tolerance = row[2]
            )
        

    end_time = datetime.datetime.now()
    print(end_time - start_time)

load_core_tables()
