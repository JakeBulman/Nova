import sys
import os
import django
import datetime
import pyodbc
import pandas as pd
from openpyxl import load_workbook
import re
from django.conf import settings
from django.utils.timezone import make_aware
from dateutil.parser import parse
from email.message import EmailMessage
import smtplib
import traceback

import warnings
warnings.filterwarnings("ignore")

if os.getenv('DJANGO_DEVELOPMENT') == 'true':
    print('DEV')
    path = os.path.join('C:\\Users\\bulmaj\\OneDrive - Cambridge\\Desktop\\Dev\\Nova')
    sys.path.append(path)
    os.environ['DJANGO_SETTINGS_MODULE'] = 'redepplan.settings_dev'
elif os.getenv('DJANGO_PRODUCTION') == 'true':
    print('PROD')
    path = os.path.join('C:\\Dev\\Nova')
    sys.path.append(path)
    os.environ['DJANGO_SETTINGS_MODULE'] = 'redepplan.settings_prod'
else:
    print('UAT - Check')
    path = os.path.join('C:\\Dev\\nova')
    sys.path.append(path)
    os.environ['DJANGO_SETTINGS_MODULE'] = 'redepplan.settings'

django.setup()

from enquiries.models import CentreEnquiryRequests, EnquiryRequestParts, EnquiryComponents, EnquiryPersonnel, EnquiryPersonnelDetails, EnquiryBatches, EnquiryComponentElements, TaskManager, UniqueCreditor, EnquiryComponentsHistory, EnquiryComponentsExaminerChecks, TaskTypes, EarServerSettings, EnquiryGrades, EnquiryDeadline, ExaminerPanels, MarkTolerances, ScaledMarks, CentreEnquiryRequestsExtra, EnquiryRequestPartsExtra, EnquiryComponentsExtra

def load_core_tables():
    try:
        start_time = datetime.datetime.now()
        print("Start Time:" + str(datetime.datetime.now()))
        EarServerSettings.objects.update(delta_load_status='Primary Load Starting, first table loading')

        #Limits enquiries to session or enquiry list
        session_id = EarServerSettings.objects.first().session_id_list
        enquiry_id_list = EarServerSettings.objects.first().enquiry_id_list
        if enquiry_id_list != '':
            enquiry_id_list = ' and cer.sid in (' + enquiry_id_list + ')'

        # # Get datalake data - Centre Enquiry Requests
        with pyodbc.connect("DSN=Impala DL", autocommit=True) as conn:
            df = pd.read_sql(f'''
                select distinct
                cer.sid as enquiry_id,
                cer.status as enquiry_status,
                cer.created_datetime as eps_creation_date,
                cer.completed_datetime as eps_completion_date,
                cer.acknowledge_letter_ind as eps_ack_letter_ind,
                cer.ses_sid as eps_ses_sid,
                cer.cnu_id as centre_id,
                cer.created_by as created_by,
                cer.cie_direct_id as cie_direct_id
                from ar_meps_req_prd.centre_enquiry_requests cer
                left join ar_meps_req_prd.enquiry_request_parts erp
                on erp.cer_sid = cer.sid
                where ses_sid in ({session_id}) 
                --and erp.es_service_code in ('1','1S','2','2P','2PS','2S','ASC','ASR','3')
                {enquiry_id_list}
                                    ''', conn)

        rolling_insert = []
        def insert_to_model(row):
            centre_id = row['centre_id'][:5]
            ministry_flag = None
            if re.search("^S[0-9]",centre_id):
                ministry_flag = 'S'
            if re.search("^MU",centre_id):
                ministry_flag = 'MU'
            rolling_insert.append(
                CentreEnquiryRequests(
                        enquiry_id = int(row['enquiry_id']),
                        enquiry_status = row['enquiry_status'],
                        eps_creation_date = make_aware(parse(str(row['eps_creation_date']))),
                        eps_completion_date = make_aware(parse(str(row['eps_completion_date']))),
                        eps_ack_letter_ind = row['eps_ack_letter_ind'],
                        eps_ses_sid = str(row['eps_ses_sid'])[:5],
                        centre_id = row['centre_id'][:5],
                        created_by = row['created_by'][:50],
                        cie_direct_id = None, 
                        ministry_flag = ministry_flag
                    )
            )

        df.apply(insert_to_model, axis=1)
        CentreEnquiryRequests.objects.bulk_create(rolling_insert,update_conflicts=True,unique_fields=['enquiry_id'],
                                                    update_fields=['enquiry_status','eps_creation_date','eps_completion_date','eps_ack_letter_ind',
                                                                    'eps_ses_sid','centre_id','created_by','cie_direct_id','ministry_flag']
                                                    )

        print("CER loaded:" + str(datetime.datetime.now()))
        EarServerSettings.objects.update(delta_load_status='Primary - Table 1 of 11 loaded, Centre Enquiry Requests')


        # # Get datalake data - Enquiry Request Parts
        with pyodbc.connect("DSN=Impala DL", autocommit=True) as conn:
            df = pd.read_sql(f'''
                select 
                erp.sid as erp_sid,
                erp.cer_sid as cer_sid,
                erp.es_service_code as service_code,
                erp.caom_ses_sid as eps_ses_sid,
                erp.caom_ass_code as eps_ass_code,
                erp.caom_asv_ver_no as eps_ass_ver_no,
                erp.caom_opt_code as eps_option_code,
                erp.caom_can_unique_identifier as eps_cand_unique_id,
                erp.caom_cand_no as eps_cand_id,
                erp.caom_cnu_id as eps_centre_id,
                if(erp.component_ind="Y","C","S") as eps_comp_ind,
                erp.caom_measure as eps_script_measure,
                erp.booked_in_error_ind as booked_in_error_ind,
                can.full_name as stud_name,
                erp.grade_confirmed_ind as grade_confirmed_ind,
                erp.grade_changed_ind as grade_changed_ind
                from ar_meps_req_prd.enquiry_request_parts erp
                left join cie.ca_candidates can
                on erp.caom_can_unique_identifier = can.unique_id
                where caom_ses_sid in ({session_id})
                and erp.cer_sid is not null 
                                    ''', conn)

        rolling_insert = []

        def insert_to_model(row):
            rolling_insert.append(
                EnquiryRequestParts(
                        erp_sid = int(row['erp_sid']),
                        #cer_sid = CentreEnquiryRequests.objects.get(enquiry_id=row['cer_sid']),
                        #cer_sid = all_cer.get(enquiry_id=row['cer_sid']),
                        cer_sid_id = str(int(row['cer_sid'])),
                        service_code = row['service_code'],
                        eps_ses_sid = int(row['eps_ses_sid']),
                        eps_ass_code = row['eps_ass_code'],
                        eps_ass_ver_no = int(row['eps_ass_ver_no']),
                        eps_option_code = row['eps_option_code'],
                        eps_cand_unique_id = row['eps_cand_unique_id'],
                        eps_cand_id = int(row['eps_cand_id']),
                        eps_centre_id = row['eps_centre_id'],
                        eps_comp_ind = row['eps_comp_ind'],
                        eps_script_measure = row['eps_script_measure'],
                        booked_in_error_ind = row['booked_in_error_ind'],
                        stud_name = row['stud_name'][:100],
                        grade_confirmed_ind = row['grade_confirmed_ind'],
                        grade_changed_ind = row['grade_changed_ind']
                    )
            )

        df.apply(insert_to_model, axis=1)
        EnquiryRequestParts.objects.bulk_create(rolling_insert,update_conflicts=True,unique_fields=['erp_sid'],
                                                    update_fields=['cer_sid','service_code','eps_ses_sid','eps_ass_code',
                                                                    'eps_ass_ver_no','eps_option_code','eps_cand_unique_id','eps_cand_id','eps_centre_id',
                                                                    'eps_comp_ind','eps_script_measure','booked_in_error_ind','stud_name','grade_confirmed_ind','grade_changed_ind'])

        EarServerSettings.objects.update(delta_load_status='Primary - Table 2 of 11 loaded, Enquiry Request Parts')
        print("ERP loaded:" + str(datetime.datetime.now()))



        # # Get datalake data - Enquiry Components
        with pyodbc.connect("DSN=Impala DL", autocommit=True) as conn:
            df = pd.read_sql(f'''
                    select distinct
                    ec.sid as ec_sid,
                    ec.erp_sid as erp_sid,
                    ec.ccm_ses_sid as eps_ses_sid,
                    ses.sessionname as eps_ses_name,
                    ec.ccm_ass_code as eps_ass_code,
                    ec.ccm_asv_ver_no as eps_ass_ver_no,
                    ec.ccm_com_id as eps_com_id,
                    prod.qualfication as eps_qual_id,
                    prod.qualification_short_text as eps_qual_name,
                    prod.assessment_short_text as eps_ass_name,
                    prod.component_text as eps_comp_name,
                    ec.ccm_measure as ccm_measure
                    from ar_meps_req_prd.enquiry_components ec
                    left join ar_meps_req_prd.enquiry_request_parts erp
                    on ec.erp_sid = erp.sid
                    left join cie.ca_products prod
                    on cast(ec.ccm_ass_code as string) = prod.assessment
                    and cast(ec.ccm_asv_ver_no as string) = prod.assessment_version_no
                    and ec.ccm_com_id = prod.component
                    and erp.caom_opt_code = prod.option_code
                    left join cie.ods_sessions ses
                    on ec.ccm_ses_sid = ses.sessionid
                    where ccm_ses_sid in ({session_id}) 
                                    ''', conn)
            
            
        workbook = load_workbook("\\\\filestorage\cie\Operations\Results Team\Enquiries About Results\\0.Nova Downloads\\Type Of Script.xlsx")
        sheet = workbook.active

        rolling_insert = []
        def insert_to_model(row):
            try:
                rolling_insert.append(
                    EnquiryComponents(
                        ec_sid = int(row['ec_sid']),
                        erp_sid_id = int(row['erp_sid']),
                        eps_ses_sid = int(row['eps_ses_sid']),
                        eps_ses_name = row['eps_ses_name'],
                        eps_ass_code = row['eps_ass_code'],
                        eps_ass_ver_no = int(row['eps_ass_ver_no']),
                        eps_com_id = row['eps_com_id'],
                        eps_qual_id = row['eps_qual_id'],
                        eps_qual_name = row['eps_qual_name'][:50],
                        eps_ass_name = row['eps_ass_name'][:50],
                        eps_comp_name = row['eps_comp_name'][:50],
                        ccm_measure = row['ccm_measure'],
                    )
                )
            except:
                pass

        df.apply(insert_to_model, axis=1)
        EnquiryComponents.objects.bulk_create(rolling_insert,update_conflicts=True,unique_fields=['ec_sid'],
                                                    update_fields=['erp_sid_id','eps_ses_sid','eps_ses_name',
                                                                    'eps_ass_code','eps_ass_ver_no','eps_com_id','eps_qual_id','eps_qual_name',
                                                                    'eps_ass_name','eps_comp_name','ccm_measure'])


        for row in sheet.iter_rows():
            ass_code = str(row[0].value).zfill(4)
            comp_id = str(row[2].value).zfill(2)
            script_type = row[5].value

            EnquiryComponents.objects.filter(eps_ass_code=ass_code,eps_com_id=comp_id).update(script_type=script_type)

        EarServerSettings.objects.update(delta_load_status='Primary - Table 3 of 11 loaded, Enquiry Components')
        print("EC loaded:" + str(datetime.datetime.now()))



            # # Get datalake data - Scaled Marks
        with pyodbc.connect("DSN=Impala DL", autocommit=True) as conn:
            df = pd.read_sql(f'''
            select 
            distinct 
                c.sessionassessmentcomponentid,
                a.assessmentcode as eps_ass_code,
                a.componentid as eps_com_id,
                c.centrenumber as eps_cnu_id,
                c.candidatenumber as eps_cand_no,
                s.sessionid as eps_ses_id,
                c.mark as raw_mark,
                c.assessormark as assessor_mark,
                c.finalmark as final_mark,
                c.examinernumber as exm_examiner_no,
                case
                when c.assessormark is not null then assessormark 
                else c.mark
                end as scaled_mark,
                case
                when assessormark is not null then "Scaled" 
                when assessormark is null then "No scaling"
                end as original_exm_scaled
            from cie.ods_candidatemarkelementmarks as c
            left join cie.ods_sessionassessmentcomponents as s
            on c.sessionassessmentcomponentid=s.sessionassessmentcomponentid
            left join cie.ods_assessmentcomponents as a
            on s.assessmentcomponentid=a.assessmentcomponentid

            inner join 
            (select 
            cer.cnu_id,
            ec.ccm_ass_code,
            ec.ccm_com_id,
            erp.caom_cand_no,
            cer.ses_sid
            from ar_meps_req_prd.centre_enquiry_requests cer
            left join ar_meps_req_prd.enquiry_request_parts erp
            on cer.sid=erp.cer_sid
            left join ar_meps_req_prd.enquiry_components ec
            on erp.sid=ec.erp_sid
            ) req
            on req.cnu_id = c.centrenumber
            and req.ccm_ass_code = a.assessmentcode
            and req.ccm_com_id = a.componentid
            and req.caom_cand_no = c.candidatenumber
            and req.ses_sid = s.sessionid
            
            where c.businessstreamid='02'
                and s.isdeletedfromsource!=1
                and c.isdeletedfromsource!=1
                and a.isdeletedfromsource!=1
                and c.examinernumber!=''
                and s.sessionid in ({session_id}) 
                                    ''', conn)
            
        rolling_insert = []
        def insert_to_model(row):
            rolling_insert.append(
            ScaledMarks(
                eps_ass_code = row['eps_ass_code'],
                eps_com_id = int(row['eps_com_id']),
                eps_cnu_id = row['eps_cnu_id'],
                eps_cand_no = row['eps_cand_no'],
                eps_ses_sid = row['eps_ses_id'],
                raw_mark = row['raw_mark'],
                assessor_mark  = row['assessor_mark'],
                final_mark = row['final_mark'],
                exm_examiner_no = row['exm_examiner_no'],
                scaled_mark = row['scaled_mark'],
                original_exm_scaled = row['original_exm_scaled'],
                )
            )

        df.apply(insert_to_model, axis=1)
        print("Scaled Marks prepped:" + str(datetime.datetime.now())) 
        ScaledMarks.objects.all().delete()
        print("Scaled Marks deleted:" + str(datetime.datetime.now()))
        ScaledMarks.objects.bulk_create(rolling_insert)
        print("Scaled Marks loaded:" + str(datetime.datetime.now()))
        EarServerSettings.objects.update(delta_load_status='Primary - Table 4 of 11 loaded, Scaled Marks')

        # # Get datalake data - Enquiry Batches
        with pyodbc.connect("DSN=Impala DL", autocommit=True) as conn:
            df = pd.read_sql(f'''
                select distinct
                sb.sid as eb_sid,
                sb.created_date as created_date,
                sb.enpe_eper_per_sid as eper_per_sid
                from ar_meps_req_prd.summary_batches sb
                left join ar_meps_req_prd.enquiry_component_elements ece
                on sb.sid = ece.eb_sid
                left join ar_meps_req_prd.enquiry_components ec
                on ec.sid = ece.ec_sid
                where ec.ccm_ses_sid in ({session_id}) 
                                    ''', conn)


        rolling_insert = []
        def insert_to_model(row):
            try:
                enpe = int(row['eper_per_sid'])
            except:
                enpe = None
            rolling_insert.append(
                EnquiryBatches(
                    eb_sid = int(row['eb_sid']),
                    created_date = make_aware(parse(str(row['created_date']))),
                    enpe_eper_per_sid = enpe,
                )
            )

        df.apply(insert_to_model, axis=1)
        EnquiryBatches.objects.bulk_create(rolling_insert,update_conflicts=True,unique_fields=['eb_sid'],
                                                    update_fields=['created_date','enpe_eper_per_sid'])

        print("EB loaded:" + str(datetime.datetime.now()))
        EarServerSettings.objects.update(delta_load_status='Primary - Table 5 of 11 loaded, Enquiry Batches')

        #Get datalake data - Enquiry Component Elements
        with pyodbc.connect("DSN=Impala DL", autocommit=True) as conn:
            df = pd.read_sql(f'''
                select distinct
                ece.ec_sid as ec_sid,
                ece.status as ece_status,
                ece.eb_sid as eb_sid,
                ece.clerical_mark as clerical_mark,
                ece.mark_after_enquiry as mark_after_enquiry,
                ece.justification_code as justification_code,
                ece.omr_mark_changed_ind as omr_mark_changed_ind,
                ece.omr_mark_confirmed_ind as omr_mark_confirmed_ind,
                ece.clerical_mark_changed_ind as clerical_mark_changed_ind,
                ece.clerical_mark_confirmed_ind as clerical_mark_confirmed_ind,
                ece.me_id as me_id             
                from  ar_meps_req_prd.enquiry_component_elements ece
                left join ar_meps_req_prd.enquiry_components ec
                on ece.ec_sid = ec.sid
                where ec.ccm_ses_sid in ({session_id}) 
                                    ''', conn)
            
        rolling_insert = []
        def insert_to_model(row):
            try:
                eb = int(row['eb_sid'])
            except:
                eb = None
            rolling_insert.append(
                EnquiryComponentElements(
                    ec_sid_id = int(row['ec_sid']),
                    ece_status = row['ece_status'][:20],
                    eb_sid_id = eb,
                    clerical_mark = row['clerical_mark'],
                    mark_after_enquiry = row['mark_after_enquiry'],
                    justification_code = row['justification_code'],
                    omr_mark_changed_ind = row['omr_mark_changed_ind'],
                    omr_mark_confirmed_ind = row['omr_mark_confirmed_ind'],
                    clerical_mark_changed_ind = row['clerical_mark_changed_ind'],
                    clerical_mark_confirmed_ind = row['clerical_mark_confirmed_ind'],
                    me_id = row['me_id'],
                )
            )

        df.apply(insert_to_model, axis=1)
        EnquiryComponentElements.objects.all().delete()
        EnquiryComponentElements.objects.bulk_create(rolling_insert)
        print("ECE loaded:" + str(datetime.datetime.now()))
        EarServerSettings.objects.update(delta_load_status='Primary - Table 6 of 11 loaded, Enquiry Component Elements')

        # Get datalake data - ECH
        with pyodbc.connect("DSN=Impala DL", autocommit=True) as conn:
            df = pd.read_sql(f'''
                            SELECT distinct
                            cer.sid as cer_sid,
                            ec.sid as ec_sid,
                            wrm.ses_sid as ses_sid,
                            wrm.ass_code as ass_code,
                            wrm.com_id as com_id,
                            wrm.cnu_id as cnu_id,
                            wrm.cand_no as cand_no,
                            spp.examiner_no as exm_position,
                            wrm.kbr_code as kbr_code,
                            kbr.reason as kbr_reason,
                            wrm.mark as current_mark,
                            ec.ccm_measure as ear_mark,
                            ec.ccm_outcome as ear_mark_alt,
                            mld.batch as omr_batch,
                            mld.position as omr_position
                            from ar_meps_req_prd.enquiry_components ec
                            left join ar_meps_req_prd.enquiry_request_parts erp
                            on erp.sid = ec.erp_sid
                            left join ar_meps_req_prd.centre_enquiry_requests cer
                            on cer.sid = erp.cer_sid
                            left join ar_meps_mark_prd.working_raw_marks wrm
                            on wrm.ses_sid = ec.ccm_ses_sid
                            and wrm.ass_code = ec.ccm_ass_code
                            and wrm.com_id = ec.ccm_com_id
                            and wrm.cand_no = ec.ccm_cand_no
                            and wrm.cnu_id = ec.ccm_cnu_id
                            left join ar_meps_pan_prd.session_panel_positions spp
                            on wrm.spp_sid = spp.sid
                            left join ar_meps_mark_prd.keyed_batch_reason kbr
                            on kbr.code = wrm.kbr_code
                            left join ar_meps_mark_prd.mark_load_details mld
                            on wrm.mld_sid = mld.sid
                            where ec.ccm_ses_sid in ({session_id}) 
                                    ''', conn)

        rolling_insert = []
        def insert_to_model(row):
            try:
                kbr_reason = row['kbr_reason'][:50]
            except:
                kbr_reason = ''
            rolling_insert.append(
                EnquiryComponentsHistory(
                    cer_sid_id = int(row['cer_sid']),
                    ec_sid_id = int(row['ec_sid']),
                    eps_ses_sid = int(row['ses_sid']),
                    eps_ass_code = row['ass_code'],
                    eps_com_id = row['com_id'],
                    eps_cnu_id = row['cnu_id'],
                    eps_cand_no = int(row['cand_no']),
                    exm_position = row['exm_position'],
                    kbr_code = row['kbr_code'],
                    kbr_reason = kbr_reason,
                    current_mark = row['current_mark'],
                    ear_mark = row['ear_mark'],
                    ear_mark_alt = row['ear_mark_alt'],
                )
            )

        df.apply(insert_to_model, axis=1)
        EnquiryComponentsHistory.objects.all().delete()
        EnquiryComponentsHistory.objects.bulk_create(rolling_insert)

        print("ECH loaded:" + str(datetime.datetime.now()))
        EarServerSettings.objects.update(delta_load_status='Primary - Table 7 of 11 loaded, Enquiry Component History')



        # Get datalake data - ECEC
        with pyodbc.connect("DSN=Impala DL", autocommit=True) as conn:
            df = pd.read_sql(f'''
                        SELECT distinct
                        cer.sid as cer_sid,
                        ec.sid as ec_sid,
                        wrm.ses_sid as eps_ses_id,
                        wrm.ass_code as eps_ass_code,
                        wrm.com_id as eps_com_id,
                        wrm.cnu_id as eps_cnu_id,
                        wrm.cand_no as eps_cand_no,
                        spp.examiner_no as exm_position,
                        wrm.kbr_code as kbr_code,
                        kbr.reason as kbr_reason
                        from ar_meps_req_prd.enquiry_components ec
                        left join ar_meps_req_prd.enquiry_request_parts erp
                        on erp.sid = ec.erp_sid
                        left join ar_meps_req_prd.centre_enquiry_requests cer
                        on cer.sid = erp.cer_sid
                        left join (select ses_sid,ass_code,com_id,cnu_id,cand_no,kbr_code,mark,spp_sid from ar_meps_mark_prd.working_raw_marks where ses_sid in ({session_id}) 
                        union select ses_sid,ass_code,com_id,cnu_id,cand_no,kbr_code,mark,spp_sid from ar_meps_mark_prd.historical_raw_marks where ses_sid in ({session_id})) wrm
                        on wrm.ses_sid = ec.ccm_ses_sid
                        and wrm.ass_code = ec.ccm_ass_code
                        and wrm.com_id = ec.ccm_com_id
                        and wrm.cand_no = ec.ccm_cand_no
                        and wrm.cnu_id = ec.ccm_cnu_id
                        left join ar_meps_pan_prd.session_panel_positions spp
                        on wrm.spp_sid = spp.sid
                        left join ar_meps_mark_prd.keyed_batch_reason kbr
                        on kbr.code = wrm.kbr_code
                        where wrm.kbr_code in ('TL','GR')
                        and ec.ccm_ses_sid in ({session_id}) 
                                    ''', conn)

        rolling_insert = []
        def insert_to_model(row):
            try:
                kbr_reason = row['kbr_reason'][:50]
            except:
                kbr_reason = ''
            rolling_insert.append(
                EnquiryComponentsExaminerChecks(
                    cer_sid_id = int(row['cer_sid']),
                    ec_sid_id = int(row['ec_sid']),
                    eps_ses_sid = int(row['eps_ses_id']),
                    eps_ass_code = row['eps_ass_code'],
                    eps_com_id = row['eps_com_id'],
                    eps_cnu_id = row['eps_cnu_id'],
                    eps_cand_no = int(row['eps_cand_no']),
                    exm_position = row['exm_position'],
                    kbr_code = row['kbr_code'],
                    kbr_reason = kbr_reason,
                )
            )

        df.apply(insert_to_model, axis=1)
        EnquiryComponentsExaminerChecks.objects.all().delete()
        EnquiryComponentsExaminerChecks.objects.bulk_create(rolling_insert)

        print("ECEC loaded:" + str(datetime.datetime.now()))
        EarServerSettings.objects.update(delta_load_status='Primary - Table 8 of 11 loaded, Enquiry Component Examiner Checks')

        # # Get datalake data - Enquiry Grades
        with pyodbc.connect("DSN=Impala DL", autocommit=True) as conn:
            df = pd.read_sql(f'''
                select 
                cer.sid as enquiry_id,
                erp.caom_ass_code as eps_ass_code,
                erp.caom_cand_no as eps_cand_no,
                qgrp.title as previous_grade,
                qgrp.seq_no as previous_seq,
                qgrn.title as new_grade,
                qgrn.seq_no as new_seq
                            
                from ar_meps_req_prd.centre_enquiry_requests cer
                left join ar_meps_req_prd.enquiry_request_parts erp
                on cer.sid = erp.cer_sid
                left join ar_meps_ord_prd.audit_changes ac
                on cer.ses_sid = ac.ses_sid
                and cer.cnu_id = ac.co_centre_id
                and erp.caom_ass_code = ac.ass_code
                and erp.caom_cand_no = ac.coc_cand_no
                left join ar_meps_prod_prd.qualification_grades qgr
                on qgr.qgs_sid = erp.qgr_qgs_sid and qgr.seq_no = erp.qgr_seq_no_derived
                left join ar_meps_prod_prd.qualification_grades qgrp
                on erp.qgr_qgs_sid = qgrp.qgs_sid and cast(qgrp.seq_no as string) = ac.previous_value
                left join ar_meps_prod_prd.qualification_grades qgrn
                on erp.qgr_qgs_sid = qgrn.qgs_sid and cast(qgrp.seq_no as string) = ac.new_value
                where ac.record_type = "CAOM"
                --and ac.confirmation_status = "UNC"
                and ac.field_name in ("QGR_SEQ_NO_DERIVED","QGR_SEQ_NO_MANUAL")
                and ac.ses_sid in ({session_id}) 
                                    ''', conn)

        rolling_insert = []
        def insert_to_model(row):
            try:
                previous_seq = int(row['previous_seq'])
            except:
                previous_seq = None
            try:
                new_seq = int(row['new_seq'])
            except:
                new_seq = None
            rolling_insert.append(
                EnquiryGrades(
                    enquiry_id_id = int(row['enquiry_id']),
                    eps_ass_code = row['eps_ass_code'],
                    eps_cand_no = int(row['eps_cand_no']),
                    previous_grade = row['previous_grade'],
                    previous_seq = previous_seq,
                    new_grade = row['new_grade'],
                    new_seq = new_seq
                )
            )

        df.apply(insert_to_model, axis=1)
        EnquiryGrades.objects.all().delete()
        EnquiryGrades.objects.bulk_create(rolling_insert)
        EarServerSettings.objects.update(delta_load_status='Primary - Table 9 of 11 loaded, Enquiry Grades')
        print("EG loaded:" + str(datetime.datetime.now()))

        rolling_insert = []
        for enquiry in CentreEnquiryRequests.objects.filter(enquiry_deadline__isnull=True,enquiries__service_code__contains='P'):
            rolling_insert.append(
            EnquiryDeadline(
                enquiry_id = enquiry,
                unique_enquiry_id = enquiry.enquiry_id,
                enquiry_deadline = enquiry.eps_creation_date + datetime.timedelta(days=18),
                original_enquiry_deadline = enquiry.eps_creation_date + datetime.timedelta(days=18)
            )
            )

        EnquiryDeadline.objects.bulk_create(rolling_insert,ignore_conflicts=True)

        rolling_insert = []
        for enquiry in CentreEnquiryRequests.objects.filter(enquiry_deadline__isnull=True,ministry_flag='S'):
            rolling_insert.append(
            EnquiryDeadline(
                enquiry_id = enquiry,
                unique_enquiry_id = enquiry.enquiry_id,
                enquiry_deadline = enquiry.eps_creation_date + datetime.timedelta(days=21),
                original_enquiry_deadline = enquiry.eps_creation_date + datetime.timedelta(days=21)
            )
            )

        EnquiryDeadline.objects.bulk_create(rolling_insert,ignore_conflicts=True)

        rolling_insert = []
        for enquiry in CentreEnquiryRequests.objects.filter(enquiry_deadline__isnull=True):
            rolling_insert.append(
            EnquiryDeadline(
                enquiry_id = enquiry,
                unique_enquiry_id = enquiry.enquiry_id,
                enquiry_deadline = enquiry.eps_creation_date + datetime.timedelta(days=30),
                original_enquiry_deadline = enquiry.eps_creation_date + datetime.timedelta(days=30)
            )
            )

        EnquiryDeadline.objects.bulk_create(rolling_insert,ignore_conflicts=True)


        print("ED loaded:" + str(datetime.datetime.now()))
        EarServerSettings.objects.update(delta_load_status='Primary - Table 10 of 11 loaded, Enquiry Deadlines')

        filename=os.path.join("\\\\filestorage\cie\Operations\Results Team\Enquiries About Results\\0.Nova Downloads\\Tolerances.xlsx")
        workbook = load_workbook(filename)
        sheet = workbook.active

        # Iterating through All rows with all columns...
        for i in range(1, sheet.max_row+1):
            row = [cell.value for cell in sheet[i]] # sheet[n] gives nth row (list of cells)
            if str(row[0]) != 'None':
                if MarkTolerances.objects.filter(eps_ass_code = str(row[0]).zfill(4),eps_com_id = str(row[1]).zfill(2)).exists():
                    MarkTolerances.objects.filter(eps_ass_code = str(row[0]).zfill(4),eps_com_id = str(row[1]).zfill(2)).update(mark_tolerance = row[2])
                else: 
                    MarkTolerances.objects.create(
                        eps_ass_code = str(row[0]).zfill(4),
                        eps_com_id = str(row[1]).zfill(2),
                        mark_tolerance = row[2]
                    )

        print("Tols loaded:" + str(datetime.datetime.now()))
        EarServerSettings.objects.update(delta_load_status='Primary - Table 11 of 11 loaded, Enquiry Tolerances')


        service_list = ['1','1S','2','2P','2PS','2S','ASC','ASR','3']
        to_insert = []
        for enquiry in CentreEnquiryRequests.objects.exclude(enquiry_tasks__task_id='INITCH').filter(enquiries__service_code__in=service_list):
            enquiry_id = enquiry.enquiry_id
            enquiry_obj = enquiry
            to_insert.append(
                TaskManager(
                    enquiry_id = enquiry_obj,
                    ec_sid = None,
                    task_id_id = 'INITCH',
                    task_assigned_to = None,
                    task_assigned_date = None,
                    task_completion_date = None
                )
            )
        TaskManager.objects.bulk_create(to_insert)
        print("IEC loaded:" + str(datetime.datetime.now()))

        if os.getenv('DJANGO_PRODUCTION') == 'true':
            EarServerSettings.objects.update(delta_load_status='Primary Load completed at '+str(datetime.datetime.now().strftime("%d/%m/%Y, %H:%M:%S")))
            email = EmailMessage()
            email["From"] = "results.enquiries@cambridge.org"
            email["To"] = "results.enquiries@cambridge.org, jacob.bulman@cambridge.org,jonathon.east@cambridge.org,ben.herbert@cambridge.org,charlotte.weedon@cambridge.org,morgan.jones@cambridge.org,lab.d@cambridgeassessment.org.uk"
            email["Subject"] = "Primary Data Load - SUCCESS"
            email.set_content("Primary load was successful, IECs are ready to be completed.", subtype='html')

            sender = "results.enquiries@cambridge.org"
            smtp = smtplib.SMTP("smtp0.ucles.internal", port=25) 
            smtp.sendmail(sender, ["results.enquiries@cambridge.org", "jacob.bulman@cambridge.org","jonathon.east@cambridge.org","ben.herbert@cambridge.org","charlotte.weedon@cambridge.org","morgan.jones@cambridge.org","lab.d@cambridgeassessment.org.uk"], email.as_string())
            smtp.quit()


            end_time = datetime.datetime.now()
            print(end_time - start_time)
    except Exception as e:
        print(traceback.format_exc())
        if os.getenv('DJANGO_PRODUCTION') == 'true':
            EarServerSettings.objects.update(delta_load_status='Primary Load failed at '+str(datetime.datetime.now().strftime("%d/%m/%Y, %H:%M:%S")))
            email = EmailMessage()
            email["From"] = "results.enquiries@cambridge.org"
            email["To"] = "results.enquiries@cambridge.org, jacob.bulman@cambridge.org,jonathon.east@cambridge.org,ben.herbert@cambridge.org,charlotte.weedon@cambridge.org,morgan.jones@cambridge.org,lab.d@cambridgeassessment.org.uk"
            email["Subject"] = "Primary Data Load - ERROR"
            email.set_content(f"Primary load has failed, please contact the system administrator for further details. Error: {e}", subtype='html')

            sender = "results.enquiries@cambridge.org"
            smtp = smtplib.SMTP("smtp0.ucles.internal", port=25) 
            smtp.sendmail(sender, ["results.enquiries@cambridge.org", "jacob.bulman@cambridge.org","jonathon.east@cambridge.org","ben.herbert@cambridge.org","charlotte.weedon@cambridge.org","morgan.jones@cambridge.org","lab.d@cambridgeassessment.org.uk"], email.as_string())
            smtp.quit()
            end_time = datetime.datetime.now()
            print(end_time - start_time)



load_core_tables()
