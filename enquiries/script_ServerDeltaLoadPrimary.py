import sys
import os
import django
import datetime
import pyodbc
import pandas as pd
from openpyxl import load_workbook
import re
from django.conf import settings
from django.utils.timezone import make_aware
from dateutil.parser import parse
from email.message import EmailMessage
import smtplib
import traceback

if os.getenv('DJANGO_DEVELOPMENT') == 'true':
    print('DEV')
    path = os.path.join('C:\\Users\\bulmaj\\OneDrive - Cambridge\\Desktop\\Dev\\Nova')
    sys.path.append(path)
    os.environ['DJANGO_SETTINGS_MODULE'] = 'redepplan.settings_dev'
elif os.getenv('DJANGO_PRODUCTION') == 'true':
    print('PROD')
    path = os.path.join('C:\\Dev\\Nova')
    sys.path.append(path)
    os.environ['DJANGO_SETTINGS_MODULE'] = 'redepplan.settings_prod'
else:
    print('UAT - Check')
    path = os.path.join('C:\\Dev\\nova')
    sys.path.append(path)
    os.environ['DJANGO_SETTINGS_MODULE'] = 'redepplan.settings'

django.setup()

from enquiries.models import CentreEnquiryRequests, EnquiryRequestParts, EnquiryComponents, EnquiryPersonnel, EnquiryPersonnelDetails, EnquiryBatches, EnquiryComponentElements, TaskManager, UniqueCreditor, EnquiryComponentsHistory, EnquiryComponentsExaminerChecks, TaskTypes, EarServerSettings, EnquiryGrades, EnquiryDeadline, ExaminerPanels, MarkTolerances, ScaledMarks, CentreEnquiryRequestsExtra, EnquiryRequestPartsExtra, EnquiryComponentsExtra

def load_core_tables():

    start_time = datetime.datetime.now()
    print("Start Time:" + str(datetime.datetime.now()))
    EarServerSettings.objects.update(delta_load_status='Delta Load Starting, first table loading')

    try:

        #Limits enquiries to session or enquiry list
        session_id = EarServerSettings.objects.first().session_id_list
        enquiry_id_list = EarServerSettings.objects.first().enquiry_id_list
        if enquiry_id_list != '':
            enquiry_id_list = ' and cer.sid in (' + enquiry_id_list + ')'

        print("ENQ:" + enquiry_id_list)

        # # Get datalake data - Centre Enquiry Requests
        with pyodbc.connect("DSN=hive.ucles.internal", autocommit=True) as conn:
            df = pd.read_sql(f'''
                select distinct
                cer.sid as enquiry_id,
                cer.status as enquiry_status,
                cer.created_datetime as eps_creation_date,
                cer.completed_datetime as eps_completion_date,
                cer.acknowledge_letter_ind as eps_ack_letter_ind,
                cer.ses_sid as eps_ses_sid,
                cer.cnu_id as centre_id,
                cer.created_by as created_by,
                cer.cie_direct_id as cie_direct_id
                from ar_meps_req_prd.centre_enquiry_requests cer
                left join ar_meps_req_prd.enquiry_request_parts erp
                on erp.cer_sid = cer.sid
                where ses_sid in ({session_id}) 
                and erp.es_service_code in ('1','1S','2','2P','2PS','2S','ASC','ASR','3')
                {enquiry_id_list}
                                    ''', conn)
            print(df)
        
        def insert_to_model_cer(row):
            if CentreEnquiryRequests.objects.filter(enquiry_id=row['enquiry_id']).exists():
                CentreEnquiryRequests.objects.filter(enquiry_id=row['enquiry_id']).update(
                    enquiry_id = row['enquiry_id'],
                    enquiry_status = row['enquiry_status'],
                    eps_creation_date = make_aware(parse(row['eps_creation_date'])),
                    eps_completion_date = row['eps_completion_date'],
                    eps_ack_letter_ind = row['eps_ack_letter_ind'],
                    eps_ses_sid = row['eps_ses_sid'],
                    centre_id = row['centre_id'],
                    created_by = row['created_by'][:50],
                    cie_direct_id = row['cie_direct_id'], 
                )
            else:
                CentreEnquiryRequests.objects.create(
                    enquiry_id = row['enquiry_id'],
                    enquiry_status = row['enquiry_status'],
                    eps_creation_date = make_aware(parse(row['eps_creation_date'])),
                    eps_completion_date = row['eps_completion_date'],
                    eps_ack_letter_ind = row['eps_ack_letter_ind'],
                    eps_ses_sid = row['eps_ses_sid'],
                    centre_id = row['centre_id'],
                    created_by = row['created_by'][:50],
                    cie_direct_id = row['cie_direct_id'],
                    ministry_flag = None
                )

        df.apply(insert_to_model_cer, axis=1)
        print("CER loaded:" + str(datetime.datetime.now()))
        EarServerSettings.objects.update(delta_load_status='Table 1 of 15 loaded, Centre Enquiry Requests')

        for enquiry in CentreEnquiryRequests.objects.all():
            centre_id = enquiry.centre_id
            if re.search("^S[0-9]",centre_id):
                enquiry.ministry_flag = 'S'
                enquiry.save()
            if re.search("^MU",centre_id):
                enquiry.ministry_flag = 'MU'
                enquiry.save()



        # # Get datalake data - Enquiry Request Parts
        with pyodbc.connect("DSN=hive.ucles.internal", autocommit=True) as conn:
            df = pd.read_sql(f'''
                select 
                erp.sid as erp_sid,
                erp.cer_sid as cer_sid,
                erp.es_service_code as service_code,
                erp.caom_ses_sid as eps_ses_sid,
                erp.caom_ass_code as eps_ass_code,
                erp.caom_asv_ver_no as eps_ass_ver_no,
                erp.caom_opt_code as eps_option_code,
                erp.caom_can_unique_identifier as eps_cand_unique_id,
                erp.caom_cand_no as eps_cand_id,
                erp.caom_cnu_id as eps_centre_id,
                if(erp.component_ind="Y","C","S") as eps_comp_ind,
                erp.caom_measure as eps_script_measure,
                erp.booked_in_error_ind as booked_in_error_ind,
                can.full_name as stud_name,
                erp.grade_confirmed_ind as grade_confirmed_ind,
                erp.grade_changed_ind as grade_changed_ind
                from ar_meps_req_prd.enquiry_request_parts erp
                left join cie.ca_candidates can
                on erp.caom_can_unique_identifier = can.unique_id
                where caom_ses_sid in ({session_id}) 
                                    ''', conn)

        def insert_to_model_erp(row):
            if EnquiryRequestParts.objects.filter(erp_sid = row['erp_sid']).exists():
                EnquiryRequestParts.objects.filter(erp_sid = row['erp_sid']).update(
                    erp_sid = row['erp_sid'],
                    cer_sid = CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=row['cer_sid']),
                    service_code = row['service_code'],
                    eps_ses_sid = row['eps_ses_sid'],
                    eps_ass_code = row['eps_ass_code'],
                    eps_ass_ver_no = row['eps_ass_ver_no'],
                    eps_option_code = row['eps_option_code'],
                    eps_cand_unique_id = row['eps_cand_unique_id'],
                    eps_cand_id = row['eps_cand_id'],
                    eps_centre_id = row['eps_centre_id'],
                    eps_comp_ind = row['eps_comp_ind'],
                    eps_script_measure = row['eps_script_measure'],
                    booked_in_error_ind = row['booked_in_error_ind'],
                    stud_name = row['stud_name'][:100],
                    grade_confirmed_ind = row['grade_confirmed_ind'],
                    grade_changed_ind = row['grade_changed_ind']
            )
            else:
                try:
                    EnquiryRequestParts.objects.create(
                        erp_sid = row['erp_sid'],
                        cer_sid = CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=row['cer_sid']),
                        service_code = row['service_code'],
                        eps_ses_sid = row['eps_ses_sid'],
                        eps_ass_code = row['eps_ass_code'],
                        eps_ass_ver_no = row['eps_ass_ver_no'],
                        eps_option_code = row['eps_option_code'],
                        eps_cand_unique_id = row['eps_cand_unique_id'],
                        eps_cand_id = row['eps_cand_id'],
                        eps_centre_id = row['eps_centre_id'],
                        eps_comp_ind = None,
                        eps_script_measure = 1,
                        booked_in_error_ind = row['booked_in_error_ind'],
                        stud_name = row['stud_name'][:100],
                        grade_confirmed_ind = row['grade_confirmed_ind']
                    )
                except:
                    pass

        df.apply(insert_to_model_erp, axis=1)

        EarServerSettings.objects.update(delta_load_status='Table 2 of 15 loaded, Enquiry Request Parts')
        print("ERP loaded:" + str(datetime.datetime.now()))

        # # Get datalake data - Enquiry Components
        with pyodbc.connect("DSN=hive.ucles.internal", autocommit=True) as conn:
            df = pd.read_sql(f'''
                    select distinct
                    ec.sid as ec_sid,
                    ec.erp_sid as erp_sid,
                    ec.ccm_ses_sid as eps_ses_sid,
                    ses.sessionname as eps_ses_name,
                    ec.ccm_ass_code as eps_ass_code,
                    ec.ccm_asv_ver_no as eps_ass_ver_no,
                    ec.ccm_com_id as eps_com_id,
                    prod.qualfication as eps_qual_id,
                    prod.qualification_short_text as eps_qual_name,
                    prod.assessment_short_text as eps_ass_name,
                    prod.component_text as eps_comp_name,
                    ec.ccm_measure as ccm_measure
                    from ar_meps_req_prd.enquiry_components ec
                    left join ar_meps_req_prd.enquiry_request_parts erp
                    on ec.erp_sid = erp.sid
                    left join cie.ca_products prod
                    on ec.ccm_ass_code = prod.assessment
                    and ec.ccm_asv_ver_no = prod.assessment_version_no
                    and ec.ccm_com_id = prod.component
                    and erp.caom_opt_code = prod.option_code
                    left join cie.ods_sessions ses
                    on ec.ccm_ses_sid = ses.sessionid
                    where ccm_ses_sid in ({session_id}) 
                                    ''', conn)

        def insert_to_model_ec(row):
            if EnquiryComponents.objects.filter(ec_sid = row['ec_sid']).exists():
                EnquiryComponents.objects.filter(ec_sid = row['ec_sid']).update(
                    ec_sid = row['ec_sid'],
                    erp_sid = EnquiryRequestParts.objects.only('erp_sid').get(erp_sid=row['erp_sid']),
                    eps_ses_sid = row['eps_ses_sid'],
                    eps_ses_name = row['eps_ses_name'],
                    eps_ass_code = row['eps_ass_code'],
                    eps_ass_ver_no = row['eps_ass_ver_no'],
                    eps_com_id = row['eps_com_id'],
                    eps_qual_id = row['eps_qual_id'],
                    eps_qual_name = row['eps_qual_name'][:50],
                    eps_ass_name = row['eps_ass_name'][:50],
                    eps_comp_name = row['eps_comp_name'][:50],
                    ccm_measure = row['ccm_measure'],
                )
            else:
                try:
                    EnquiryComponents.objects.create(
                        ec_sid = row['ec_sid'],
                        erp_sid = EnquiryRequestParts.objects.only('erp_sid').get(erp_sid=row['erp_sid']),
                        eps_ses_sid = row['eps_ses_sid'],
                        eps_ses_name = row['eps_ses_name'],
                        eps_ass_code = row['eps_ass_code'],
                        eps_ass_ver_no = row['eps_ass_ver_no'],
                        eps_com_id = row['eps_com_id'],
                        eps_qual_id = row['eps_qual_id'],
                        eps_qual_name = row['eps_qual_name'][:50],
                        eps_ass_name = row['eps_ass_name'][:50],
                        eps_comp_name = row['eps_comp_name'][:50],
                        ccm_measure = row['ccm_measure'],
                    )
                except:
                    pass
        
        df.apply(insert_to_model_ec, axis=1)

        #filename=os.path.join("\\\\filestorage\cie\Operations\Results Team\Enquiries About Results\\0.Nova Downloads\\Type Of Script.xlsx")
        workbook = load_workbook("\\\\filestorage\cie\Operations\Results Team\Enquiries About Results\\0.Nova Downloads\\Type Of Script.xlsx")
        sheet = workbook.active

        for row in sheet.iter_rows():
            ass_code = str(row[0].value).zfill(4)
            comp_id = str(row[2].value).zfill(2)
            script_type = row[5].value

            EnquiryComponents.objects.filter(eps_ass_code=ass_code,eps_com_id=comp_id).update(script_type=script_type)

        EarServerSettings.objects.update(delta_load_status='Table 3 of 15 loaded, Enquiry Components')
        print("EC loaded:" + str(datetime.datetime.now()))

            # # Get datalake data - Scaled Marks
        with pyodbc.connect("DSN=hive.ucles.internal", autocommit=True) as conn:
            df = pd.read_sql(f'''
            select 
            distinct 
                c.sessionassessmentcomponentid,
                a.assessmentcode as eps_ass_code,
                a.componentid as eps_com_id,
                c.centrenumber as eps_cnu_id,
                c.candidatenumber as eps_cand_no,
                s.sessionid as eps_ses_id,
                c.mark as raw_mark,
                c.assessormark as assessor_mark,
                c.finalmark as final_mark,
                c.examinernumber as exm_examiner_no,
                case
                when c.assessormark is not null then assessormark 
                else c.mark
                end as scaled_mark,
                case
                when assessormark is not null then "Scaled" 
                when assessormark is null then "No scaling"
                end as original_exm_scaled
            from cie.ods_candidatemarkelementmarks as c
            left join cie.ods_sessionassessmentcomponents as s
            on c.sessionassessmentcomponentid=s.sessionassessmentcomponentid
            left join cie.ods_assessmentcomponents as a
            on s.assessmentcomponentid=a.assessmentcomponentid
        
            inner join 
            (select 
            cer.cnu_id,
            ec.ccm_ass_code,
            ec.ccm_com_id,
            erp.caom_cand_no,
            cer.ses_sid
            from ar_meps_req_prd.centre_enquiry_requests cer
            left join ar_meps_req_prd.enquiry_request_parts erp
            on cer.sid=erp.cer_sid
            left join ar_meps_req_prd.enquiry_components ec
            on erp.sid=ec.erp_sid
            ) req
            on req.cnu_id = c.centrenumber
            and req.ccm_ass_code = a.assessmentcode
            and req.ccm_com_id = a.componentid
            and req.caom_cand_no = c.candidatenumber
            and req.ses_sid = s.sessionid
            
            where c.businessstreamid='02'
                and s.isdeletedfromsource!=1
                and c.isdeletedfromsource!=1
                and a.isdeletedfromsource!=1
                and c.examinernumber!=''
                and s.sessionid in ({session_id}) 
                                    ''', conn)
            
        print("Scaled Marks prepped:" + str(datetime.datetime.now()))
            
        ScaledMarks.objects.all().delete()

        print("Scaled Marks deleted:" + str(datetime.datetime.now()))

        def insert_to_model_erp(row):
            # try:
            ScaledMarks.objects.create(
                eps_ass_code = row['eps_ass_code'],
                eps_com_id = row['eps_com_id'],
                eps_cnu_id = row['eps_cnu_id'],
                eps_cand_no = row['eps_cand_no'],
                eps_ses_sid = row['eps_ses_id'],
                raw_mark = row['raw_mark'],
                assessor_mark  = row['assessor_mark'],
                final_mark = row['final_mark'],
                exm_examiner_no = row['exm_examiner_no'],
                scaled_mark = row['scaled_mark'],
                original_exm_scaled = row['original_exm_scaled'],
                )
            # except:
            #     pass

        df.apply(insert_to_model_erp, axis=1)

        print("Scaled Marks loaded:" + str(datetime.datetime.now()))
        EarServerSettings.objects.update(delta_load_status='Table 4 of 15 loaded, Scaled Marks')

        # # Get datalake data - Enquiry Batches
        with pyodbc.connect("DSN=hive.ucles.internal", autocommit=True) as conn:
            df = pd.read_sql(f'''
                select distinct
                sb.sid as eb_sid,
                sb.created_date as created_date,
                sb.enpe_eper_per_sid as eper_per_sid
                from ar_meps_req_prd.summary_batches sb
                left join ar_meps_req_prd.enquiry_component_elements ece
                on sb.sid = ece.eb_sid
                left join ar_meps_req_prd.enquiry_components ec
                on ec.sid = ece.ec_sid
                where ec.ccm_ses_sid in ({session_id}) 
                                    ''', conn)

        def insert_to_model_eb(row):
            if EnquiryBatches.objects.filter(eb_sid = row['eb_sid']).exists():
                EnquiryBatches.objects.filter(eb_sid = row['eb_sid']).update(
                    eb_sid = row['eb_sid'],
                    created_date = make_aware(parse(row['created_date'])),
                    enpe_eper_per_sid = row['eper_per_sid'],
                )
            else:
                try:
                    EnquiryBatches.objects.create(
                        eb_sid = row['eb_sid'],
                        created_date = make_aware(parse(row['created_date'])),
                        enpe_eper_per_sid = row['eper_per_sid'],
                    )
                except:
                    pass

        df.apply(insert_to_model_eb, axis=1)
        print("EB loaded:" + str(datetime.datetime.now()))
        EarServerSettings.objects.update(delta_load_status='Table 5 of 15 loaded, Enquiry Batches')

        #Get datalake data - Enquiry Component Elements
        with pyodbc.connect("DSN=hive.ucles.internal", autocommit=True) as conn:
            df = pd.read_sql(f'''
                select distinct
                ece.ec_sid as ec_sid,
                ece.status as ece_status,
                ece.eb_sid as eb_sid,
                ece.clerical_mark as clerical_mark,
                ece.mark_after_enquiry as mark_after_enquiry,
                ece.justification_code as justification_code,
                ece.omr_mark_changed_ind as omr_mark_changed_ind,
                ece.omr_mark_confirmed_ind as omr_mark_confirmed_ind,
                ece.clerical_mark_changed_ind as clerical_mark_changed_ind,
                ece.clerical_mark_confirmed_ind as clerical_mark_confirmed_ind,
                ece.me_id as me_id             
                from  ar_meps_req_prd.enquiry_component_elements ece
                left join ar_meps_req_prd.enquiry_components ec
                on ece.ec_sid = ec.sid
                where ec.ccm_ses_sid in ({session_id}) 
                                    ''', conn)

        def insert_to_model_ec(row):
            if EnquiryComponentElements.objects.filter(ec_sid=row['ec_sid']).exists():
                EnquiryComponentElements.objects.filter(ec_sid=row['ec_sid']).update(
                    ec_sid = EnquiryComponents.objects.only('ec_sid').get(ec_sid=row['ec_sid']),
                    ece_status = row['ece_status'][:20],
                    eb_sid = EnquiryBatches.objects.only('eb_sid').filter(eb_sid=row['eb_sid']).first(),
                    clerical_mark = row['clerical_mark'],
                    mark_after_enquiry = row['mark_after_enquiry'],
                    justification_code = row['justification_code'],
                    omr_mark_changed_ind = row['omr_mark_changed_ind'],
                    omr_mark_confirmed_ind = row['omr_mark_confirmed_ind'],
                    clerical_mark_changed_ind = row['clerical_mark_changed_ind'],
                    clerical_mark_confirmed_ind = row['clerical_mark_confirmed_ind'],
                    me_id = row['me_id'],
                )   
            else:
                try:
                    EnquiryComponentElements.objects.create(
                        ec_sid = EnquiryComponents.objects.only('ec_sid').get(ec_sid=row['ec_sid']),
                        ece_status = row['ece_status'][:20],
                        eb_sid = EnquiryBatches.objects.only('eb_sid').filter(eb_sid=row['eb_sid']).first(),
                        clerical_mark = row['clerical_mark'],
                        mark_after_enquiry = row['mark_after_enquiry'],
                        justification_code = row['justification_code'],
                        omr_mark_changed_ind = row['omr_mark_changed_ind'],
                        omr_mark_confirmed_ind = row['omr_mark_confirmed_ind'],
                        clerical_mark_changed_ind = row['clerical_mark_changed_ind'],
                        clerical_mark_confirmed_ind = row['clerical_mark_confirmed_ind'],
                        me_id = row['me_id'],
                    )   
                except:
                    pass


        df.apply(insert_to_model_ec, axis=1)

        print("ECE loaded:" + str(datetime.datetime.now()))
        EarServerSettings.objects.update(delta_load_status='Table 6 of 15 loaded, Enquiry Component Elements')


        # Get datalake data - ECH
        with pyodbc.connect("DSN=hive.ucles.internal", autocommit=True) as conn:
            df = pd.read_sql(f'''
                            SELECT distinct
                            cer.sid as cer_sid,
                            ec.sid as ec_sid,
                            wrm.ses_sid as ses_sid,
                            wrm.ass_code as ass_code,
                            wrm.com_id as com_id,
                            wrm.cnu_id as cnu_id,
                            wrm.cand_no as cand_no,
                            spp.examiner_no as exm_position,
                            wrm.kbr_code as kbr_code,
                            kbr.reason as kbr_reason,
                            wrm.mark as current_mark,
                            ec.ccm_measure as ear_mark,
                            ec.ccm_outcome as ear_mark_alt,
                            mld.batch as omr_batch,
                            mld.position as omr_position
                            from ar_meps_req_prd.enquiry_components ec
                            left join ar_meps_req_prd.enquiry_request_parts erp
                            on erp.sid = ec.erp_sid
                            left join ar_meps_req_prd.centre_enquiry_requests cer
                            on cer.sid = erp.cer_sid
                            left join ar_meps_mark_prd.working_raw_marks wrm
                            on wrm.ses_sid = ec.ccm_ses_sid
                            and wrm.ass_code = ec.ccm_ass_code
                            and wrm.com_id = ec.ccm_com_id
                            and wrm.cand_no = ec.ccm_cand_no
                            and wrm.cnu_id = ec.ccm_cnu_id
                            left join ar_meps_pan_prd.session_panel_positions spp
                            on wrm.spp_sid = spp.sid
                            left join ar_meps_mark_prd.keyed_batch_reason kbr
                            on kbr.code = wrm.kbr_code
                            left join ar_meps_mark_prd.mark_load_details mld
                            on wrm.mld_sid = mld.sid
                            where ec.ccm_ses_sid in ({session_id}) 
                                    ''', conn)

        def insert_to_model_enpee(row):
            try:
                kbr_reason = row['kbr_reason'][:50]
            except:
                kbr_reason = ''
            if EnquiryComponentsHistory.objects.filter(ec_sid=row['ec_sid']).exists():
                EnquiryComponentsHistory.objects.filter(ec_sid=row['ec_sid']).update(
                cer_sid = CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=row['cer_sid']),
                ec_sid = EnquiryComponents.objects.only('ec_sid').get(ec_sid=row['ec_sid']),
                eps_ses_sid = row['ses_sid'],
                eps_ass_code = row['ass_code'],
                eps_com_id = row['com_id'],
                eps_cnu_id = row['cnu_id'],
                eps_cand_no = row['cand_no'],
                exm_position = row['exm_position'],
                kbr_code = row['kbr_code'],
                kbr_reason = kbr_reason,
                current_mark = row['current_mark'],
                ear_mark = row['ear_mark'],
                ear_mark_alt = row['ear_mark_alt'],
                omr_batch = row['omr_batch'],
                omr_position = row['omr_position'],
                )

            else:
                try:
                    EnquiryComponentsHistory.objects.create(
                    cer_sid = CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=row['cer_sid']),
                    ec_sid = EnquiryComponents.objects.only('ec_sid').get(ec_sid=row['ec_sid']),
                    eps_ses_sid = row['ses_sid'],
                    eps_ass_code = row['ass_code'],
                    eps_com_id = row['com_id'],
                    eps_cnu_id = row['cnu_id'],
                    eps_cand_no = row['cand_no'],
                    exm_position = row['exm_position'],
                    kbr_code = row['kbr_code'],
                    kbr_reason = kbr_reason,
                    current_mark = row['current_mark'],
                    ear_mark = row['ear_mark'],
                    ear_mark_alt = row['ear_mark_alt'],
                    )
                except:
                    pass
            
        df.apply(insert_to_model_enpee, axis=1)

        print("ECH loaded:" + str(datetime.datetime.now()))
        EarServerSettings.objects.update(delta_load_status='Table 11 of 15 loaded, Enquiry Component History')

        # Get datalake data - ECEC
        with pyodbc.connect("DSN=hive.ucles.internal", autocommit=True) as conn:
            df = pd.read_sql(f'''
                        SELECT distinct
                        cer.sid as cer_sid,
                        ec.sid as ec_sid,
                        wrm.ses_sid as eps_ses_id,
                        wrm.ass_code as eps_ass_code,
                        wrm.com_id as eps_com_id,
                        wrm.cnu_id as eps_cnu_id,
                        wrm.cand_no as eps_cand_no,
                        spp.examiner_no as exm_position,
                        wrm.kbr_code as kbr_code,
                        kbr.reason as kbr_reason
                        from ar_meps_req_prd.enquiry_components ec
                        left join ar_meps_req_prd.enquiry_request_parts erp
                        on erp.sid = ec.erp_sid
                        left join ar_meps_req_prd.centre_enquiry_requests cer
                        on cer.sid = erp.cer_sid
                        left join (select ses_sid,ass_code,com_id,cnu_id,cand_no,kbr_code,mark,spp_sid from ar_meps_mark_prd.working_raw_marks where ses_sid in ({session_id}) 
                        union select ses_sid,ass_code,com_id,cnu_id,cand_no,kbr_code,mark,spp_sid from ar_meps_mark_prd.historical_raw_marks where ses_sid in ({session_id})) wrm
                        on wrm.ses_sid = ec.ccm_ses_sid
                        and wrm.ass_code = ec.ccm_ass_code
                        and wrm.com_id = ec.ccm_com_id
                        and wrm.cand_no = ec.ccm_cand_no
                        and wrm.cnu_id = ec.ccm_cnu_id
                        left join ar_meps_pan_prd.session_panel_positions spp
                        on wrm.spp_sid = spp.sid
                        left join ar_meps_mark_prd.keyed_batch_reason kbr
                        on kbr.code = wrm.kbr_code
                        where wrm.kbr_code in ('TL','GR')
                        and ec.ccm_ses_sid in ({session_id}) 
                                    ''', conn)

        def insert_to_model_enpee(row):
            if EnquiryComponentsExaminerChecks.objects.filter(ec_sid=row['ec_sid']).exists():
                EnquiryComponentsExaminerChecks.objects.filter(ec_sid=row['ec_sid']).update(
                cer_sid = CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=row['cer_sid']),
                ec_sid = EnquiryComponents.objects.only('ec_sid').get(ec_sid=row['ec_sid']),
                eps_ses_sid = row['eps_ses_id'],
                eps_ass_code = row['eps_ass_code'],
                eps_com_id = row['eps_com_id'],
                eps_cnu_id = row['eps_cnu_id'],
                eps_cand_no = row['eps_cand_no'],
                exm_position = row['exm_position'],
                kbr_code = row['kbr_code'],
                kbr_reason = row['kbr_reason'][:50],
                )

            else:
                try:
                    EnquiryComponentsExaminerChecks.objects.create(
                    cer_sid = CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=row['cer_sid']),
                    ec_sid = EnquiryComponents.objects.only('ec_sid').get(ec_sid=row['ec_sid']),
                    eps_ses_sid = row['eps_ses_id'],
                    eps_ass_code = row['eps_ass_code'],
                    eps_com_id = row['eps_com_id'],
                    eps_cnu_id = row['eps_cnu_id'],
                    eps_cand_no = row['eps_cand_no'],
                    exm_position = row['exm_position'],
                    kbr_code = row['kbr_code'],
                    kbr_reason = row['kbr_reason'][:50],
                    )
                except:
                    pass
            
        df.apply(insert_to_model_enpee, axis=1)

        print("EPNE2 loaded:" + str(datetime.datetime.now()))
        EarServerSettings.objects.update(delta_load_status='Table 12 of 15 loaded, Enquiry Component Examiner Checks')


        # # Get datalake data - Enquiry Grades
        with pyodbc.connect("DSN=hive.ucles.internal", autocommit=True) as conn:
            df = pd.read_sql(f'''
                select 
                cer.sid as enquiry_id,
                erp.caom_ass_code as eps_ass_code,
                erp.caom_cand_no as eps_cand_no,
                qgrp.title as previous_grade,
                qgrp.seq_no as previous_seq,
                qgrn.title as new_grade,
                qgrn.seq_no as new_seq
                            
                from ar_meps_req_prd.centre_enquiry_requests cer
                left join ar_meps_req_prd.enquiry_request_parts erp
                on cer.sid = erp.cer_sid
                left join ar_meps_ord_prd.audit_changes ac
                on cer.ses_sid = ac.ses_sid
                and cer.cnu_id = ac.co_centre_id
                and erp.caom_ass_code = ac.ass_code
                and erp.caom_cand_no = ac.coc_cand_no
                left join ar_meps_prod_prd.qualification_grades qgr
                on qgr.qgs_sid = erp.qgr_qgs_sid and qgr.seq_no = erp.qgr_seq_no_derived
                left join ar_meps_prod_prd.qualification_grades qgrp
                on erp.qgr_qgs_sid = qgrp.qgs_sid and qgrp.seq_no = ac.previous_value
                left join ar_meps_prod_prd.qualification_grades qgrn
                on erp.qgr_qgs_sid = qgrn.qgs_sid and qgrn.seq_no = ac.new_value
                where ac.record_type = "CAOM"
                --and ac.confirmation_status = "UNC"
                and ac.field_name in ("QGR_SEQ_NO_DERIVED","QGR_SEQ_NO_MANUAL")
                and ac.ses_sid in ({session_id}) 
                                    ''', conn)

        def insert_to_model_erp(row):
            if EnquiryGrades.objects.filter(enquiry_id=row['enquiry_id']).exists():
                EnquiryGrades.objects.filter(enquiry_id=row['enquiry_id']).update(
                    enquiry_id = CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=row['enquiry_id']),
                    eps_ass_code = row['eps_ass_code'],
                    eps_cand_no = row['eps_cand_no'],
                    previous_grade = row['previous_grade'],
                    previous_seq = row['previous_seq'],
                    new_grade = row['new_grade'],
                    new_seq = row['new_seq']
                )
            else:
                try:
                    EnquiryGrades.objects.create(
                        enquiry_id = CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=row['enquiry_id']),
                        eps_ass_code = row['eps_ass_code'],
                        eps_cand_no = row['eps_cand_no'],
                        previous_grade = row['previous_grade'],
                        previous_seq = row['previous_seq'],
                        new_grade = row['new_grade'],
                        new_seq = row['new_seq']
                    )
                except:
                    pass

        EnquiryGrades.objects.all().delete()

        df.apply(insert_to_model_erp, axis=1)

        EarServerSettings.objects.update(delta_load_status='Table 13 of 15 loaded, Enquiry Grades')
        print("EG loaded:" + str(datetime.datetime.now()))

        for enquiry in CentreEnquiryRequests.objects.all():
            try:
                erp_service = EnquiryRequestParts.objects.filter(cer_sid = enquiry.enquiry_id).first().service_code
            except:
                erp_service = 'Unknown'
            if 'P' in erp_service:
                deadline = enquiry.eps_creation_date + datetime.timedelta(days=18)
            elif enquiry.ministry_flag == "S":
                deadline = enquiry.eps_creation_date + datetime.timedelta(days=21)
            else:
                deadline = enquiry.eps_creation_date + datetime.timedelta(days=30)
            if not EnquiryDeadline.objects.filter(enquiry_id=enquiry).exists():
                EnquiryDeadline.objects.create(
                    enquiry_id = enquiry,
                    enquiry_deadline = deadline,
                    original_enquiry_deadline = deadline
                )

        print("ED loaded:" + str(datetime.datetime.now()))
        EarServerSettings.objects.update(delta_load_status='Table 14 of 15 loaded, Enquiry Deadlines')

        filename=os.path.join("\\\\filestorage\cie\Operations\Results Team\Enquiries About Results\\0.Nova Downloads\\Tolerances.xlsx")
        workbook = load_workbook(filename)
        sheet = workbook.active

        # Iterating through All rows with all columns...
        # for i in range(1, sheet.max_row+1):
        #     row = [cell.value for cell in sheet[i]] # sheet[n] gives nth row (list of cells)
        #     if str(row[0]) != 'None':
        #         if MarkTolerances.objects.filter(eps_ass_code = str(row[0]).zfill(4),eps_com_id = str(row[1]).zfill(2)).exists():
        #             MarkTolerances.objects.filter(eps_ass_code = str(row[0]).zfill(4),eps_com_id = str(row[1]).zfill(2)).update(mark_tolerance = row[2])
        #         else: 
        #             MarkTolerances.objects.create(
        #                 eps_ass_code = str(row[0]).zfill(4),
        #                 eps_com_id = str(row[1]).zfill(2),
        #                 mark_tolerance = row[2]
        #             )

        # print("Tols loaded:" + str(datetime.datetime.now()))
        EarServerSettings.objects.update(delta_load_status='Table 15 of 15 loaded, Enquiry Tolerances')

        for enquiry in CentreEnquiryRequests.objects.all():
            enquiry_id = enquiry.enquiry_id
            if not TaskManager.objects.filter(enquiry_id=enquiry_id,task_id_id = 'INITCH').exists():
                TaskManager.objects.create(
                    enquiry_id = CentreEnquiryRequests.objects.only('enquiry_id').get(enquiry_id=enquiry_id),
                    ec_sid = None,
                    task_id = TaskTypes.objects.get(task_id = 'INITCH'),
                    task_assigned_to = None,
                    task_assigned_date = None,
                    task_completion_date = None
                )

        print("IEC loaded:" + str(datetime.datetime.now()))

        EarServerSettings.objects.update(delta_load_status='Delta Load Complete, IECs ready to process at '+str(datetime.datetime.now().strftime("%d/%m/%Y, %H:%M:%S")))

    ### LOAD EXTRAS DATA 

        # # Get datalake data - Centre Enquiry Requests
        with pyodbc.connect("DSN=hive.ucles.internal", autocommit=True) as conn:
            df = pd.read_sql(f'''
                select distinct
                cer.sid as enquiry_id,
                cer.status as enquiry_status,
                cer.created_datetime as eps_creation_date,
                cer.completed_datetime as eps_completion_date,
                cer.acknowledge_letter_ind as eps_ack_letter_ind,
                cer.ses_sid as eps_ses_sid,
                cer.cnu_id as centre_id,
                cer.created_by as created_by,
                cer.cie_direct_id as cie_direct_id
                from ar_meps_req_prd.centre_enquiry_requests cer
                left join ar_meps_req_prd.enquiry_request_parts erp
                on erp.cer_sid = cer.sid
                where ses_sid in ({session_id}) 
                and erp.es_service_code not in ('1','1S','2','2P','2PS','2S','ASC','ASR','3')
                {enquiry_id_list}
                                    ''', conn)
            print(df)
        
        def insert_to_model_cer(row):
            if CentreEnquiryRequestsExtra.objects.filter(enquiry_id=row['enquiry_id']).exists():
                CentreEnquiryRequestsExtra.objects.filter(enquiry_id=row['enquiry_id']).update(
                    enquiry_id = row['enquiry_id'],
                    enquiry_status = row['enquiry_status'],
                    eps_creation_date = row['eps_creation_date'],
                    eps_completion_date = row['eps_completion_date'],
                    eps_ack_letter_ind = row['eps_ack_letter_ind'],
                    eps_ses_sid = row['eps_ses_sid'],
                    centre_id = row['centre_id'],
                    created_by = row['created_by'],
                    cie_direct_id = row['cie_direct_id'], 
                )
            else:
                CentreEnquiryRequestsExtra.objects.create(
                    enquiry_id = row['enquiry_id'],
                    enquiry_status = row['enquiry_status'],
                    eps_creation_date = row['eps_creation_date'],
                    eps_completion_date = row['eps_completion_date'],
                    eps_ack_letter_ind = row['eps_ack_letter_ind'],
                    eps_ses_sid = row['eps_ses_sid'],
                    centre_id = row['centre_id'],
                    created_by = row['created_by'],
                    cie_direct_id = row['cie_direct_id'],
                )

        df.apply(insert_to_model_cer, axis=1)
        print("CER loaded:" + str(datetime.datetime.now()))

        # # Get datalake data - Enquiry Request Parts
        with pyodbc.connect("DSN=hive.ucles.internal", autocommit=True) as conn:
            df = pd.read_sql(f'''
                select 
                erp.sid as erp_sid,
                erp.cer_sid as cer_sid,
                erp.es_service_code as service_code,
                erp.caom_ses_sid as eps_ses_sid,
                erp.caom_ass_code as eps_ass_code,
                erp.caom_asv_ver_no as eps_ass_ver_no,
                erp.caom_opt_code as eps_option_code,
                erp.caom_can_unique_identifier as eps_cand_unique_id,
                erp.caom_cand_no as eps_cand_id,
                erp.caom_cnu_id as eps_centre_id,
                if(erp.component_ind="Y","C","S") as eps_comp_ind,
                erp.caom_measure as eps_script_measure,
                erp.booked_in_error_ind as booked_in_error_ind,
                can.full_name as stud_name
                from ar_meps_req_prd.enquiry_request_parts erp
                left join cie.ca_candidates can
                on erp.caom_can_unique_identifier = can.unique_id
                where caom_ses_sid in ({session_id}) 
                                    ''', conn)

        def insert_to_model_erp(row):
            if EnquiryRequestPartsExtra.objects.filter(erp_sid = row['erp_sid']).exists():
                EnquiryRequestPartsExtra.objects.filter(erp_sid = row['erp_sid']).update(
                    erp_sid = row['erp_sid'],
                    cer_sid = CentreEnquiryRequestsExtra.objects.only('enquiry_id').get(enquiry_id=row['cer_sid']),
                    service_code = row['service_code'],
                    eps_ses_sid = row['eps_ses_sid'],
                    eps_ass_code = row['eps_ass_code'],
                    eps_ass_ver_no = row['eps_ass_ver_no'],
                    eps_option_code = row['eps_option_code'],
                    eps_cand_unique_id = row['eps_cand_unique_id'],
                    eps_cand_id = row['eps_cand_id'],
                    eps_centre_id = row['eps_centre_id'],
                    eps_comp_ind = row['eps_comp_ind'],
                    eps_script_measure = row['eps_script_measure'],
                    booked_in_error_ind = row['booked_in_error_ind'],
                    stud_name = row['stud_name'],
            )
            else:
                try:
                    EnquiryRequestPartsExtra.objects.create(
                        erp_sid = row['erp_sid'],
                        cer_sid = CentreEnquiryRequestsExtra.objects.only('enquiry_id').get(enquiry_id=row['cer_sid']),
                        service_code = row['service_code'],
                        eps_ses_sid = row['eps_ses_sid'],
                        eps_ass_code = row['eps_ass_code'],
                        eps_ass_ver_no = row['eps_ass_ver_no'],
                        eps_option_code = row['eps_option_code'],
                        eps_cand_unique_id = row['eps_cand_unique_id'],
                        eps_cand_id = row['eps_cand_id'],
                        eps_centre_id = row['eps_centre_id'],
                        eps_comp_ind = row['eps_comp_ind'],
                        eps_script_measure = row['eps_script_measure'],
                        booked_in_error_ind = row['booked_in_error_ind'],
                        stud_name = row['stud_name'],
                    )
                except:
                    pass

        df.apply(insert_to_model_erp, axis=1)


        print("ERP loaded:" + str(datetime.datetime.now()))

        # # Get datalake data - Enquiry Components
        with pyodbc.connect("DSN=hive.ucles.internal", autocommit=True) as conn:
            df = pd.read_sql(f'''
                    select distinct
                    ec.sid as ec_sid,
                    ec.erp_sid as erp_sid,
                    ec.ccm_ses_sid as eps_ses_sid,
                    ses.sessionname as eps_ses_name,
                    ec.ccm_ass_code as eps_ass_code,
                    ec.ccm_asv_ver_no as eps_ass_ver_no,
                    ec.ccm_com_id as eps_com_id,
                    prod.qualfication as eps_qual_id,
                    prod.qualification_short_text as eps_qual_name,
                    prod.assessment_short_text as eps_ass_name,
                    prod.component_text as eps_comp_name,
                    ec.ccm_measure as ccm_measure
                    from ar_meps_req_prd.enquiry_components ec
                    left join ar_meps_req_prd.enquiry_request_parts erp
                    on ec.erp_sid = erp.sid
                    left join cie.ca_products prod
                    on ec.ccm_ass_code = prod.assessment
                    and ec.ccm_asv_ver_no = prod.assessment_version_no
                    and ec.ccm_com_id = prod.component
                    and erp.caom_opt_code = prod.option_code
                    left join cie.ods_sessions ses
                    on ec.ccm_ses_sid = ses.sessionid
                    where ccm_ses_sid in ({session_id}) 
                                    ''', conn)

        def insert_to_model_ec(row):
            if EnquiryComponentsExtra.objects.filter(ec_sid = row['ec_sid']).exists():
                EnquiryComponentsExtra.objects.filter(ec_sid = row['ec_sid']).update(
                    ec_sid = row['ec_sid'],
                    erp_sid = EnquiryRequestPartsExtra.objects.only('erp_sid').get(erp_sid=row['erp_sid']),
                    eps_ses_sid = row['eps_ses_sid'],
                    eps_ses_name = row['eps_ses_name'],
                    eps_ass_code = row['eps_ass_code'],
                    eps_ass_ver_no = row['eps_ass_ver_no'],
                    eps_com_id = row['eps_com_id'],
                    eps_qual_id = row['eps_qual_id'],
                    eps_qual_name = row['eps_qual_name'],
                    eps_ass_name = row['eps_ass_name'],
                    eps_comp_name = row['eps_comp_name'],
                    ccm_measure = row['ccm_measure'],
                )
            else:
                try:
                    EnquiryComponentsExtra.objects.create(
                        ec_sid = row['ec_sid'],
                        erp_sid = EnquiryRequestPartsExtra.objects.only('erp_sid').get(erp_sid=row['erp_sid']),
                        eps_ses_sid = row['eps_ses_sid'],
                        eps_ses_name = row['eps_ses_name'],
                        eps_ass_code = row['eps_ass_code'],
                        eps_ass_ver_no = row['eps_ass_ver_no'],
                        eps_com_id = row['eps_com_id'],
                        eps_qual_id = row['eps_qual_id'],
                        eps_qual_name = row['eps_qual_name'],
                        eps_ass_name = row['eps_ass_name'],
                        eps_comp_name = row['eps_comp_name'],
                        ccm_measure = row['ccm_measure'],
                    )
                except:
                    pass
        
        df.apply(insert_to_model_ec, axis=1)

        filename=os.path.join("\\\\filestorage\cie\Operations\Results Team\Enquiries About Results\\0.Nova Downloads\\Type Of Script.xlsx")
        workbook = load_workbook(filename)
        sheet = workbook.active

        for row in sheet.iter_rows():
            ass_code = str(row[0].value).zfill(4)
            comp_id = row[2].value
            script_type = row[5].value

            EnquiryComponentsExtra.objects.filter(eps_ass_code=ass_code,eps_com_id=comp_id).update(script_type=script_type)


        print("EC loaded:" + str(datetime.datetime.now()))
        if os.getenv('DJANGO_PRODUCTION') == 'true':
            EarServerSettings.objects.update(delta_load_status='Delta Load completed at '+str(datetime.datetime.now().strftime("%d/%m/%Y, %H:%M:%S")))
            email = EmailMessage()
            email["From"] = "results.enquiries@cambridge.org"
            email["To"] = "results.enquiries@cambridge.org, jacob.bulman@cambridge.org,jonathon.east@cambridge.org,ben.herbert@cambridge.org,charlotte.weedon@cambridge.org,morgan.jones@cambridge.org,lab.d@cambridgeassessment.org.uk"
            email["Subject"] = "Morning Data Load - SUCCESS"
            email.set_content("Morning load was successful, IECs are ready to be completed.", subtype='html')

            sender = "results.enquiries@cambridge.org"
            smtp = smtplib.SMTP("smtp0.ucles.internal", port=25) 
            smtp.sendmail(sender, ["results.enquiries@cambridge.org", "jacob.bulman@cambridge.org","jonathon.east@cambridge.org","ben.herbert@cambridge.org","charlotte.weedon@cambridge.org","morgan.jones@cambridge.org","lab.d@cambridgeassessment.org.uk"], email.as_string())
            smtp.quit()


            end_time = datetime.datetime.now()
            print(end_time - start_time)
    except Exception as e:
        print(traceback.format_exc())
        if os.getenv('DJANGO_PRODUCTION') == 'true':
            EarServerSettings.objects.update(delta_load_status='Delta Load failed at '+str(datetime.datetime.now().strftime("%d/%m/%Y, %H:%M:%S")))
            email = EmailMessage()
            email["From"] = "results.enquiries@cambridge.org"
            email["To"] = "results.enquiries@cambridge.org, jacob.bulman@cambridge.org,jonathon.east@cambridge.org,ben.herbert@cambridge.org,charlotte.weedon@cambridge.org,morgan.jones@cambridge.org,lab.d@cambridgeassessment.org.uk"
            email["Subject"] = "Morning Data Load - ERROR"
            email.set_content(f"Morning load has failed, please contact the system administrator for further details. Error: {e}", subtype='html')

            sender = "results.enquiries@cambridge.org"
            smtp = smtplib.SMTP("smtp0.ucles.internal", port=25) 
            smtp.sendmail(sender, ["results.enquiries@cambridge.org", "jacob.bulman@cambridge.org","jonathon.east@cambridge.org","ben.herbert@cambridge.org","charlotte.weedon@cambridge.org","morgan.jones@cambridge.org","lab.d@cambridgeassessment.org.uk"], email.as_string())
            smtp.quit()

load_core_tables()
